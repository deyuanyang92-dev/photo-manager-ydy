"""CRUD, import, export, and filtered-id helpers for taxonomy records."""
from __future__ import annotations

import traceback
from typing import Any

from PyQt6.QtCore import QModelIndex
from PyQt6.QtWidgets import QDialog, QFileDialog, QMessageBox

from app.utils import ui
from app.views.taxonomy_dialogs import _RecordDialog


class TaxonomyRecordsWorkflowMixin:
    # ── Double-click edit ─────────────────────────────────────────────────────

    def _on_row_double_click(self, index: QModelIndex) -> None:
        if self._view != "original":
            return
        rec = self._model.record_at(index.row())
        if rec is None:
            return
        if not rec.get("recordId", "").startswith("user:"):
            QMessageBox.information(
                self, "只读",
                "种子库条目不可编辑（双击用户条目可编辑）"
            )
            return
        self._edit_record(rec)

    # ── CRUD ──────────────────────────────────────────────────────────────────

    def _on_add(self) -> None:
        if self._svc is None:
            return
        dlg = _RecordDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        self._svc.learn(dlg.get_record())
        self._page = 1
        self._load_page()

    def _edit_record(self, rec: dict[str, Any]) -> None:
        """Edit record. Mirrors web openTaxonomyTableModal('edit'):
        - user: records → update in place
        - seed: records → open dialog pre-filled with seed data, save creates a user override
        """
        if self._svc is None:
            return
        is_user = rec.get("recordId", "").startswith("user:")
        dlg = _RecordDialog(self, record=rec)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        if is_user:
            self._svc.update_user_record(rec["recordId"], dlg.get_record())
        else:
            # Seed record: create user override entry (mirrors web findUserEntryForCurrent + learn)
            self._svc.learn(dlg.get_record())
        self._load_page()

    def _delete_record(self, rec: dict[str, Any]) -> None:
        if self._svc is None:
            return
        if not rec.get("recordId", "").startswith("user:"):
            QMessageBox.warning(self, "只读", "种子库条目不可删除")
            return
        name = f"{rec.get('species', '')} ({rec.get('class', '')})"
        reply = QMessageBox.question(
            self, "确认删除", f"删除用户条目「{name}」？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self._svc.delete_user_record(rec["recordId"])
        self._load_page()

    # ── Export ────────────────────────────────────────────────────────────────

    def _records_for_export(self, selected_only: bool = False) -> list[dict[str, Any]]:
        if self._svc is None:
            return []
        source_filter = None
        if self._view == "worms":
            source_filter = "seed"
        all_recs, _ = self._svc.all_records(
            source_filter=source_filter, page=0, page_size=999999
        )
        if not selected_only:
            return all_recs
        if self._select_all_filtered:
            return list(getattr(self, "_filtered_records_cache", []) or [])
        selected_ids = set(self._model.checked_ids()) | set(self._selected_ids)
        return [
            record for record in all_recs
            if str(record.get("recordId") or "") in selected_ids
        ]

    def _on_export(self, fmt: str, *, selected_only: bool = False) -> None:
        records = self._records_for_export(selected_only=selected_only)
        if selected_only and not records:
            QMessageBox.information(self, "导出所选", "请先选择一条或多条分类记录。")
            return
        if fmt == "csv":
            self._export_csv(records)
        else:
            self._export_xlsx(records)

    def _export_csv(self, records: list[dict[str, Any]]) -> None:
        path = ui.get_save_file_name(
            self, "导出 CSV", "taxonomy_export.csv", "CSV 文件 (*.csv)"
        )
        if not path:
            return
        import csv
        cols = self._model.columns()
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            header = [c["label"] for c in cols] + ["来源"]
            writer.writerow(header)
            for rec in records:
                row = [rec.get(c["key"], "") for c in cols]
                row.append("用户" if rec.get("recordId", "").startswith("user:") else "种子")
                writer.writerow(row)
        QMessageBox.information(self, "导出完成", f"已导出 {len(records)} 条到\n{path}")

    def _export_xlsx(self, records: list[dict[str, Any]]) -> None:
        path = ui.get_save_file_name(
            self, "导出 Excel", "taxonomy_export.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "分类库"
            cols = self._model.columns()
            ws.append([c["label"] for c in cols] + ["来源"])
            for rec in records:
                row = [rec.get(c["key"], "") for c in cols]
                row.append("用户" if rec.get("recordId", "").startswith("user:") else "种子")
                ws.append(row)
            wb.save(path)
            QMessageBox.information(self, "导出完成", f"已导出 {len(records)} 条到\n{path}")
        except ImportError:
            QMessageBox.critical(self, "缺少依赖", "需要 openpyxl 库：pip install openpyxl")
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))

    # ── Import ────────────────────────────────────────────────────────────────

    def get_filtered_uids(self) -> list:
        """Return record identifiers of all records currently shown in the table (after filter applied).

        User records return their ``recordId``.  Seed records without one return
        ``seed:{page_offset+row}`` — the same convention used by _taxonomy_record_id.
        """
        uids = []
        for row in range(self._model.rowCount()):
            rec = self._model.record_at(row)
            if rec is None:
                continue
            rid = str(rec.get("recordId", "")).strip()
            if rid:
                uids.append(rid)
            else:
                page_offset = self._model._page_offset
                uids.append(f"seed:{page_offset + row}")
        return uids

    def _on_import(self) -> None:
        """Import taxonomy records from Excel or CSV.

        Mirrors the importInput handler in app.js (POST /api/taxonomy/import).
        Expected columns (case-insensitive, any order):
          class, order, family, species (required)
          classCn, orderCn, familyCn, speciesCn, genus, genusCn (optional)
        """
        if self._svc is None:
            return
        # §7 旧: QFileDialog 直调 —— 绕过 ui.py 的 DontUseNativeDialog+居中,
        #        WSLg/多屏可能弹到屏幕外看不见。
        path = ui.get_open_file_name(
            self,
            "选择 Excel / CSV 文件",
            "",
            "表格文件 (*.xlsx *.xls *.csv)",
        )
        if not path:
            return
        try:
            if path.lower().endswith(".csv"):
                self._import_csv(path)
            else:
                self._import_xlsx(path)
        except Exception as exc:
            # §7 旧: 把整段 traceback 糊到弹窗给用户 —— 生物学新手看天书。
            # 改走 ui.exception: 简短人话 + 可折叠详情 + 复制按钮; traceback 进日志。
            ui.exception(
                self, "导入失败", exc,
                text="表格无法导入。请检查文件是否损坏、表头是否包含必填列"
                     "(class / order / family / species)。",
            )

    def _import_xlsx(self, path: str) -> None:
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "缺少依赖", "需要 openpyxl 库：pip install openpyxl")
            return
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(self, "空文件", "Excel 文件中没有数据")
            return
        imported, skipped = self._import_rows(rows[0], rows[1:])
        self._load_page()
        QMessageBox.information(
            self, "导入完成", f"成功导入 {imported} 条，跳过 {skipped} 条。"
        )

    def _import_csv(self, path: str) -> None:
        import csv
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            QMessageBox.warning(self, "空文件", "CSV 文件中没有数据")
            return
        imported, skipped = self._import_rows(rows[0], rows[1:])
        self._load_page()
        QMessageBox.information(
            self, "导入完成", f"成功导入 {imported} 条，跳过 {skipped} 条。"
        )

    def _import_rows(
        self,
        header_row: Any,
        data_rows: list[Any],
    ) -> tuple[int, int]:
        _alias: dict[str, str] = {
            "class":     "class",   "纲":    "class",
            "order":     "order",   "目":    "order",
            "family":    "family",  "科":    "family",
            "species":   "species", "种":    "species",
            "classcn":   "classCn",  "纲中文": "classCn",
            "ordercn":   "orderCn",  "目中文": "orderCn",
            "familycn":  "familyCn", "科中文": "familyCn",
            "speciescn": "speciesCn","种中文": "speciesCn",
            "genus":     "genus",    "属":    "genus",
            "genuscn":   "genusCn",  "属中文": "genusCn",
        }

        col_map: dict[str, int] = {}
        for i, h in enumerate(header_row):
            if h:
                col_map[str(h).strip().lower()] = i

        field_idx: dict[str, int] = {}
        for raw, canon in _alias.items():
            if raw in col_map:
                field_idx[canon] = col_map[raw]

        def imported_taxonomy_cell_text(row: Any, field: str) -> str:
            i = field_idx.get(field)
            if i is None:
                return ""
            try:
                v = row[i]
            except IndexError:
                return ""
            return str(v).strip() if v is not None else ""

        imported = skipped = 0
        for row in data_rows:
            rec = {f: imported_taxonomy_cell_text(row, f) for f in (
                "class", "order", "family", "species",
                "classCn", "orderCn", "familyCn", "speciesCn",
                "genus", "genusCn",
            )}
            result = self._svc.learn(rec)  # type: ignore[union-attr]
            if result:
                imported += 1
            else:
                skipped += 1
        return imported, skipped
