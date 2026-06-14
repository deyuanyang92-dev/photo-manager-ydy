"""taxonomy_view.py — Taxonomy library management view.

Faithful PyQt6 replica of the web «内置分类库» page.

Web oracle:
  - pages_dom.json "内置分类库" section (all controls/layout keys)
  - styles.css  .taxon-table-* classes (line 5651 ff.)
  - app.js renderTaxonomyPage() (line 12060 ff.)
  - taxonomy_service.py (seed read-only + user CRUD)

view_id   = "taxonomy"
nav_title = "内置分类库"
nav_icon  = "🧬"
"""
from __future__ import annotations

import json
import traceback
from pathlib import Path
from typing import Any, Optional

from PyQt6.QtCore import (
    QAbstractTableModel,
    QModelIndex,
    QPersistentModelIndex,
    QPoint,
    Qt,
    QThread,
    pyqtSignal,
)
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpacerItem,
    QSpinBox,
    QSplitter,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from app.views.base_view import BaseView
from app.services.taxonomy_service import TaxonomyService
from app.services.worms_service import WormsService

if False:  # TYPE_CHECKING
    from app.app_context import AppContext

# ── Theme colours — resolved from the LIVE active theme ───────────────────────
# Previously these were hardcoded deep-teal constants, which force-painted the
# whole 内置分类库 page dark regardless of the chosen theme → under a light theme
# the table text / labels (which the theme paints dark) became invisible.  Now
# they are refreshed from the active theme tokens by _refresh_palette(), called
# at the top of _setup_ui() and at the start of every standalone widget/dialog
# defined in this file, so each f-string picks up the live palette.
_C_PANEL = "#10242a"
_C_INPUT = "#061c1e"
_C_TEXT = "#eef3ef"
_C_TEXT_SOFT = "#cfe0db"
_C_MUTED = "#87a2a1"
_C_DIM = "#5f7d7a"
_C_ACCENT = "#29b9ab"
_C_ACCENT_HI = "#31d4c4"
_C_DANGER = "#e66e63"
_C_BORDER = "rgba(145, 182, 181, 0.18)"
_C_ACCENT_SOFT = "rgba(41, 185, 171, 0.10)"
_C_DANGER_SOFT = "rgba(230, 110, 99, 0.10)"


def _refresh_palette() -> None:
    """Rebind the module `_C_*` colours to the current theme tokens."""
    global _C_PANEL, _C_INPUT, _C_TEXT, _C_TEXT_SOFT
    global _C_MUTED, _C_DIM, _C_ACCENT, _C_ACCENT_HI
    global _C_DANGER, _C_BORDER, _C_ACCENT_SOFT, _C_DANGER_SOFT
    from app.config.theme import TOKENS
    g = TOKENS.get
    _C_PANEL = g("panel", _C_PANEL)
    _C_INPUT = g("input_bg", _C_INPUT)
    _C_TEXT = g("text", _C_TEXT)
    _C_TEXT_SOFT = g("text_soft", _C_TEXT_SOFT)
    _C_MUTED = g("muted", _C_MUTED)
    _C_DIM = g("muted_dim", _C_DIM)
    _C_ACCENT = g("accent", _C_ACCENT)
    _C_ACCENT_HI = g("accent_hover", _C_ACCENT_HI)
    _C_DANGER = g("danger", _C_DANGER)
    _C_BORDER = g("border", _C_BORDER)
    _C_ACCENT_SOFT = g("accent_soft", _C_ACCENT_SOFT)
    _C_DANGER_SOFT = g("danger_soft", _C_DANGER_SOFT)


# ── Paths ─────────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent           # app/views/
_PROJECT_ROOT = _HERE.parent.parent               # photo-platform-ydy-v3/
_DATA_DIR = _PROJECT_ROOT / "data"

_DEFAULT_SEED_PATH = _DATA_DIR / "taxonomy_seed.json"
_DEFAULT_USER_PATH = _DATA_DIR / "user_taxonomy.json"

# Page size mirrors the web default (server.js limit)
_PAGE_SIZE = 50

# ── Column definitions ────────────────────────────────────────────────────────
# (display_label, record_key, show_in_original, show_in_worms)
# These mirror getVisibleTaxonColumns() logic.

_ALL_COLS: list[dict[str, Any]] = [
    {"label": "纲(中)",   "key": "classCn",       "level": "taxonGroup", "lang": "cn"},
    {"label": "纲(拉丁)",  "key": "class",         "level": "taxonGroup", "lang": "latin"},
    {"label": "目(中)",   "key": "orderCn",        "level": "order",     "lang": "cn"},
    {"label": "目(拉丁)",  "key": "order",         "level": "order",     "lang": "latin"},
    {"label": "科(中)",   "key": "familyCn",       "level": "family",    "lang": "cn"},
    {"label": "科(拉丁)",  "key": "family",        "level": "family",    "lang": "latin"},
    {"label": "属(中)",   "key": "genusCn",        "level": "genus",     "lang": "cn"},
    {"label": "属(拉丁)",  "key": "genus",         "level": "genus",     "lang": "latin"},
    {"label": "种(中)",   "key": "speciesCn",      "level": "species",   "lang": "cn"},
    {"label": "种(拉丁)",  "key": "species",       "level": "species",   "lang": "latin"},
]

# Level keys available as column-group chips
_LEVEL_CHIPS: list[tuple[str, str]] = [
    ("order",   "目"),
    ("family",  "科"),
    ("genus",   "属"),
    ("species", "种"),
]
_LANG_CHIPS: list[tuple[str, str]] = [
    ("cn",    "中文"),
    ("latin", "拉丁名"),
]

# ── Column index constants ────────────────────────────────────────────────────
_COL_CHECK = 0   # checkbox column (taxon-th-check)
_COL_NUM   = 1   # row number column (taxon-th-num / #)
# dynamic data columns start at _COL_DATA_START
_COL_DATA_START = 2


# ── Table model ───────────────────────────────────────────────────────────────

class _TaxonTableModel(QAbstractTableModel):
    """Flat list model; columns are selected dynamically via vis_levels/vis_langs.

    Column layout (mirrors DOM):
      0 — checkbox (taxon-th-check)
      1 — # row number (taxon-th-num)
      2..N — dynamic data columns
      N+1 — 来源 (source)
      N+2 — 操作 (action placeholder — delegate renders buttons)
    """

    checked_changed = pyqtSignal()   # emitted whenever the checkbox set changes

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._records: list[dict[str, Any]] = []
        self._vis_levels: dict[str, bool] = {k: True for k, _ in _LEVEL_CHIPS}
        # taxonGroup (纲) always visible; individual level toggles affect order/family/genus/species
        self._vis_langs: dict[str, bool] = {"cn": True, "latin": True}
        self._columns: list[dict[str, Any]] = []
        self._checked: set[str] = set()    # recordIds that are checked
        self._page_offset: int = 0         # for row # display
        self._rebuild_columns()

    # -- Column management -------------------------------------------------

    def _rebuild_columns(self) -> None:
        """Recompute the ordered list of visible data columns."""
        visible = []
        for col in _ALL_COLS:
            level = col["level"]
            lang = col["lang"]
            if level == "taxonGroup":
                if self._vis_langs.get(lang, True):
                    visible.append(col)
            elif level in self._vis_levels:
                if self._vis_levels[level] and self._vis_langs.get(lang, True):
                    visible.append(col)
            elif self._vis_langs.get(lang, True):
                visible.append(col)
        self._columns = visible

    def set_vis_level(self, level: str, show: bool) -> None:
        self._vis_levels[level] = show
        self.beginResetModel()
        self._rebuild_columns()
        self.endResetModel()

    def set_vis_lang(self, lang: str, show: bool) -> None:
        self._vis_langs[lang] = show
        self.beginResetModel()
        self._rebuild_columns()
        self.endResetModel()

    def columns(self) -> list[dict[str, Any]]:
        return list(self._columns)

    # -- Data loading -------------------------------------------------------

    def set_records(self, records: list[dict[str, Any]], page_offset: int = 0) -> None:
        self.beginResetModel()
        self._records = list(records)
        self._page_offset = page_offset
        # NOTE: checked state persists across pages — selecting entries on page 1,
        # paging away, and clicking "WoRMS 更新所选" must still target page-1 rows.
        # The set is keyed by recordId (globally unique); 取消选择 clears it.
        self.endResetModel()

    def record_at(self, row: int) -> Optional[dict[str, Any]]:
        if 0 <= row < len(self._records):
            return self._records[row]
        return None

    def set_page_offset(self, offset: int) -> None:
        self._page_offset = offset

    def checked_ids(self) -> list[str]:
        return list(self._checked)

    def clear_checked(self) -> None:
        self._checked.clear()
        self.dataChanged.emit(
            self.index(0, _COL_CHECK),
            self.index(max(0, len(self._records) - 1), _COL_CHECK),
        )
        self.checked_changed.emit()

    def set_all_page_checked(self, checked: bool) -> None:
        if checked:
            self._checked.update(
                r.get("recordId", "") for r in self._records
                if r.get("recordId")
            )
        else:
            page_ids = {r.get("recordId", "") for r in self._records}
            self._checked -= page_ids
        self.dataChanged.emit(
            self.index(0, _COL_CHECK),
            self.index(max(0, len(self._records) - 1), _COL_CHECK),
        )
        self.checked_changed.emit()

    def is_page_all_checked(self) -> bool:
        if not self._records:
            return False
        return all(
            r.get("recordId", "") in self._checked
            for r in self._records
            if r.get("recordId")
        )

    # -- Qt model interface -----------------------------------------------

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._records)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        # checkbox + # + dynamic cols + 来源 + 操作
        return _COL_DATA_START + len(self._columns) + 2

    def headerData(
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> Any:
        if orientation != Qt.Orientation.Horizontal:
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            if section == _COL_CHECK:
                return ""    # checkbox painted by delegate / header checkbox
            if section == _COL_NUM:
                return "#"
            data_idx = section - _COL_DATA_START
            if 0 <= data_idx < len(self._columns):
                return self._columns[data_idx]["label"]
            # 来源 column
            after_data = section - _COL_DATA_START - len(self._columns)
            if after_data == 0:
                return "来源"
            if after_data == 1:
                return "操作"
        return None

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> Any:
        if not index.isValid():
            return None
        row, col = index.row(), index.column()
        if row >= len(self._records):
            return None
        rec = self._records[row]
        is_user = rec.get("recordId", "").startswith("user:")
        rec_id = rec.get("recordId", "")

        # ── Checkbox column ────────────────────────────────────────────
        if col == _COL_CHECK:
            if role == Qt.ItemDataRole.CheckStateRole:
                return Qt.CheckState.Checked if rec_id in self._checked else Qt.CheckState.Unchecked
            return None

        # ── Row number column ──────────────────────────────────────────
        if col == _COL_NUM:
            if role == Qt.ItemDataRole.DisplayRole:
                return str(self._page_offset + row + 1)
            if role == Qt.ItemDataRole.ForegroundRole:
                return QColor(_C_DIM)
            return None

        # ── Dynamic data columns ───────────────────────────────────────
        data_idx = col - _COL_DATA_START
        n_data = len(self._columns)

        if 0 <= data_idx < n_data:
            if role == Qt.ItemDataRole.DisplayRole:
                val = rec.get(self._columns[data_idx]["key"], "")
                return str(val) if val else ""
            if role == Qt.ItemDataRole.ForegroundRole:
                return QColor(_C_ACCENT) if is_user else QColor(_C_MUTED)
            if role == Qt.ItemDataRole.BackgroundRole:
                if is_user:
                    c = QColor(_C_ACCENT)
                    c.setAlpha(10)
                    return c
            if role == Qt.ItemDataRole.UserRole:
                return rec
            return None

        # ── 来源 column ────────────────────────────────────────────────
        after_data = col - _COL_DATA_START - n_data
        if after_data == 0:
            if role == Qt.ItemDataRole.DisplayRole:
                return "用户" if is_user else "种子"
            if role == Qt.ItemDataRole.ForegroundRole:
                return QColor(_C_ACCENT) if is_user else QColor(_C_DIM)
            return None

        # ── 操作 column ────────────────────────────────────────────────
        if after_data == 1:
            if role == Qt.ItemDataRole.UserRole:
                return rec
            return None

        return None

    def flags(self, index: QModelIndex) -> Qt.ItemFlag:
        base = super().flags(index)
        if index.column() == _COL_CHECK:
            return base | Qt.ItemFlag.ItemIsUserCheckable
        return base

    def setData(self, index: QModelIndex, value: Any, role: int = Qt.ItemDataRole.EditRole) -> bool:
        if not index.isValid():
            return False
        if index.column() == _COL_CHECK and role == Qt.ItemDataRole.CheckStateRole:
            rec = self._records[index.row()]
            rec_id = rec.get("recordId", "")
            if value == Qt.CheckState.Checked:
                self._checked.add(rec_id)
            else:
                self._checked.discard(rec_id)
            self.dataChanged.emit(index, index, [role])
            self.checked_changed.emit()
            return True
        return False


# ── Add/Edit dialog ───────────────────────────────────────────────────────────

_DIALOG_FIELDS: list[tuple[str, str, bool]] = [
    ("class",     "纲 / 门（Latin）",   True),
    ("order",     "目（Latin）",         True),
    ("family",    "科（Latin）",         True),
    ("species",   "种（Latin）",         True),
    ("classCn",   "纲中文",              False),
    ("orderCn",   "目中文",              False),
    ("familyCn",  "科中文",              False),
    ("speciesCn", "种中文",              False),
    ("genus",     "属（Latin）",         False),
    ("genusCn",   "属中文",              False),
]


class _RecordDialog(QDialog):
    """Form dialog for adding or editing a user taxonomy record.

    Mirrors openTaxonomyTableModal() in app.js.
    Shows a '查看历史' button when the record has a history[] list.
    """

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        record: Optional[dict[str, Any]] = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("编辑分类条目" if record else "新增分类条目")
        self.setMinimumWidth(420)
        self._record = record or {}
        self._inputs: dict[str, QLineEdit] = {}
        self._btn_history: QPushButton = QPushButton("查看历史")
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 12)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        for key, label_text, required in _DIALOG_FIELDS:
            inp = QLineEdit()
            inp.setText(self._record.get(key, ""))
            if required:
                inp.setPlaceholderText("必填")
            form.addRow(label_text, inp)
            self._inputs[key] = inp

        layout.addLayout(form)

        # History button — visible only when record has history entries
        history = self._record.get("history", [])
        self._btn_history.setObjectName("Outline")
        self._btn_history.setVisible(bool(history))
        self._btn_history.clicked.connect(self._show_history)
        layout.addWidget(self._btn_history)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _validate_and_accept(self) -> None:
        for key, label_text, required in _DIALOG_FIELDS:
            if required and not self._inputs[key].text().strip():
                QMessageBox.warning(self, "必填项", f"「{label_text}」不能为空")
                self._inputs[key].setFocus()
                return
        self.accept()

    def _show_history(self) -> None:
        history = self._record.get("history", [])
        if not history:
            return
        dlg = _HistoryDialog(self, history=history)
        restored = dlg.exec_and_get()
        if restored is not None:
            # Apply restored snapshot to the form inputs
            for k, v in restored.items():
                inp = self._inputs.get(k)
                if inp is not None:
                    inp.setText(str(v))

    def get_record(self) -> dict[str, Any]:
        return {k: inp.text().strip() for k, inp in self._inputs.items()}


class _HistoryDialog(QDialog):
    """Shows history entries for a user record and allows 1-level rollback.

    Each entry in history[] has: { "at": ISO8601, "before": {10 fields} }
    Selecting an entry and clicking "回滚到此版本" fills the parent form
    with the snapshot values (the parent dialog still needs OK to persist).
    """

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        history: Optional[list[dict[str, Any]]] = None,
    ) -> None:
        super().__init__(parent)
        _refresh_palette()
        self.setWindowTitle("编辑历史")
        self.setMinimumWidth(560)
        self.setMinimumHeight(320)
        self._history = list(reversed(history or []))  # newest first
        self._selected_before: Optional[dict[str, Any]] = None
        self._build_ui()

    def _build_ui(self) -> None:
        from PyQt6.QtWidgets import QListWidget, QListWidgetItem

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 12)
        layout.setSpacing(10)

        info = QLabel("选择历史版本后点击「回滚」，将把该快照填入编辑框（需再点确定保存）。")
        info.setWordWrap(True)
        info.setStyleSheet(f"color:{_C_MUTED}; font-size:11px;")
        layout.addWidget(info)

        self._list = QListWidget()
        self._list.setAlternatingRowColors(True)
        for entry in self._history:
            at = entry.get("at", "")
            before = entry.get("before", {})
            species = before.get("species", "")
            family  = before.get("family", "")
            label   = f"{at[:19].replace('T', ' ')}  →  {species} ({family})"
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole, before)
            self._list.addItem(item)
        layout.addWidget(self._list, 1)

        buttons = QDialogButtonBox()
        btn_rollback = buttons.addButton("回滚到此版本", QDialogButtonBox.ButtonRole.AcceptRole)
        btn_cancel   = buttons.addButton("取消",         QDialogButtonBox.ButtonRole.RejectRole)
        btn_rollback.clicked.connect(self._on_rollback)
        btn_cancel.clicked.connect(self.reject)
        layout.addWidget(buttons)

    def _on_rollback(self) -> None:
        item = self._list.currentItem()
        if item is None:
            QMessageBox.information(self, "提示", "请先选择一条历史记录")
            return
        self._selected_before = item.data(Qt.ItemDataRole.UserRole)
        self.accept()

    def exec_and_get(self) -> Optional[dict[str, Any]]:
        """Run dialog modally; return the snapshot dict if user confirmed, else None."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._selected_before
        return None


# ── Facet filter panel (mirrors renderTaxonFacetMenu in app.js) ──────────────

class _TaxonFacetPanel(QFrame):
    """Per-column facet filter popup (mirrors renderTaxonFacetMenu in app.js)."""

    filter_applied = pyqtSignal(str, object)
    sort_requested = pyqtSignal(str, str)

    def __init__(self, column_key: str, column_label: str, all_records: list[dict[str, Any]], current_predicate: Optional[dict[str, Any]] = None, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent, Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool)
        _refresh_palette()
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
        self.setObjectName("TaxonFacetPanel")
        self._col_key = column_key
        self._col_label = column_label
        self._all_records = all_records
        self._draft: Optional[dict[str, Any]] = dict(current_predicate) if current_predicate else {"mode": "all"}
        self._search_text = ""
        self._build_ui()
        self._fill_values()
        self.setStyleSheet(f"QFrame#TaxonFacetPanel {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER}; border-radius: 8px; }}")
        self.setMinimumWidth(280)

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 14, 14, 10)
        layout.setSpacing(8)
        layout.addWidget(self._make_label(f"{self._col_label} 筛选", "13px", "600", _C_TEXT))
        sort_row = QHBoxLayout()
        for txt, d in [("升序", "asc"), ("降序", "desc")]:
            b = QPushButton(txt); b.setObjectName("Outline"); b.setFixedHeight(24)
            b.clicked.connect(lambda _=False, dd=d: self.sort_requested.emit(self._col_key, dd))
            sort_row.addWidget(b)
        sort_row.addStretch(); layout.addLayout(sort_row)
        self._search_inp = QLineEdit(); self._search_inp.setPlaceholderText("在此列搜索值..."); self._search_inp.setFixedHeight(28)
        self._search_inp.textChanged.connect(self._on_search_changed); layout.addWidget(self._search_inp)
        sel_row = QHBoxLayout(); sel_row.setSpacing(6)
        for txt, fn in [("全选", self._on_select_all), ("全不选", self._on_select_none)]:
            b = QPushButton(txt); b.setObjectName("Outline"); b.setFixedHeight(22); b.clicked.connect(fn); sel_row.addWidget(b)
        self._btn_found = QPushButton("选搜索结果"); self._btn_found.setObjectName("Outline"); self._btn_found.setFixedHeight(22)
        self._btn_found.setEnabled(False); self._btn_found.clicked.connect(self._on_select_found); sel_row.addWidget(self._btn_found); sel_row.addStretch(); layout.addLayout(sel_row)
        self._meta_label = QLabel(""); self._meta_label.setStyleSheet(f"color: {_C_MUTED}; font-size: 11px; background: transparent;"); layout.addWidget(self._meta_label)
        self._values_list = QListWidget(); self._values_list.setMaximumHeight(200)
        self._values_list.setStyleSheet(f"QListWidget {{ background: {_C_INPUT}; border: 1px solid {_C_BORDER}; border-radius: 4px; }} QListWidget::item {{ color: {_C_TEXT}; padding: 3px 6px; }}")
        layout.addWidget(self._values_list, 1)
        act_row = QHBoxLayout(); act_row.setSpacing(6)
        btn_ok = QPushButton("确定"); btn_ok.setObjectName("Primary"); btn_ok.setFixedHeight(28); btn_ok.clicked.connect(self._on_apply); act_row.addWidget(btn_ok)
        btn_cancel = QPushButton("取消"); btn_cancel.setObjectName("Outline"); btn_cancel.setFixedHeight(28); btn_cancel.clicked.connect(self.close); act_row.addWidget(btn_cancel)
        btn_clear = QPushButton("清除筛选"); btn_clear.setObjectName("Outline"); btn_clear.setFixedHeight(28); btn_clear.clicked.connect(self._on_clear); act_row.addWidget(btn_clear)
        layout.addLayout(act_row)

    @staticmethod
    def _make_label(text: str, size: str, weight: str, color: str) -> QLabel:
        lbl = QLabel(text); lbl.setStyleSheet(f"font-size: {size}; font-weight: {weight}; color: {color}; background: transparent;"); return lbl

    def _unique_values(self) -> list[tuple[str, int]]:
        counts: dict[str, int] = {}
        q = self._search_text.strip().lower()
        for rec in self._all_records:
            v = str(rec.get(self._col_key, "") or "")
            if q and q not in v.lower(): continue
            counts[v] = counts.get(v, 0) + 1
        return sorted(counts.items(), key=lambda x: (-x[1], x[0]))

    def _fill_values(self) -> None:
        items = self._unique_values()
        self._values_list.clear()
        total_unique = len({str(r.get(self._col_key, "")) for r in self._all_records})
        self._meta_label.setText(f"匹配 {len(items)} / {total_unique} 个值")
        try: self._values_list.itemChanged.disconnect()
        except (RuntimeError, TypeError): pass
        for value, count in items:
            item = QListWidgetItem(f"{value or '(空白)'}  ({count})")
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            item.setCheckState(Qt.CheckState.Checked if self._value_checked(value) else Qt.CheckState.Unchecked)
            item.setData(Qt.ItemDataRole.UserRole, value)
            self._values_list.addItem(item)
        self._values_list.itemChanged.connect(self._on_item_changed)

    def _value_checked(self, value: str) -> bool:
        """Mirrors taxonFacetValueChecked(menu, value) in app.js."""
        draft = self._draft or {"mode": "all"}
        mode = draft.get("mode", "all")
        if mode == "include": return value in (draft.get("values") or [])
        if mode == "exclude": return value not in (draft.get("excluded") or [])
        if mode == "search":
            q = draft.get("search", "")
            return ((not q) or (q.lower() in value.lower())) and value not in (draft.get("excluded") or [])
        return True

    def _on_search_changed(self, text: str) -> None:
        self._search_text = text; self._btn_found.setEnabled(bool(text.strip())); self._fill_values()

    def _on_item_changed(self, item: QListWidgetItem) -> None:
        """Mirrors toggleTaxonFacetValue in app.js."""
        value: str = item.data(Qt.ItemDataRole.UserRole)
        checked = item.checkState() == Qt.CheckState.Checked
        draft = self._draft or {"mode": "all"}; mode = draft.get("mode", "all")
        if mode == "all":
            if not checked: self._draft = {"mode": "exclude", "excluded": [value]}
            return
        if mode == "include":
            vals: list = list(draft.get("values") or [])
            if checked and value not in vals: vals.append(value)
            elif not checked: vals = [v for v in vals if v != value]
            self._draft = {"mode": "include", "values": vals}; return
        excl: list = list(draft.get("excluded") or [])
        if not checked and value not in excl: excl.append(value)
        elif checked: excl = [v for v in excl if v != value]
        self._draft = {**draft, "excluded": excl}

    def _on_select_all(self) -> None: self._draft = {"mode": "all"}; self._fill_values()
    def _on_select_none(self) -> None: self._draft = {"mode": "include", "values": []}; self._fill_values()

    def _on_select_found(self) -> None:
        q = self._search_text.strip()
        if q: self._draft = {"mode": "search", "search": q, "excluded": []}; self._fill_values()

    def _on_apply(self) -> None:
        draft = self._draft or {"mode": "all"}
        self.filter_applied.emit(self._col_key, None if draft.get("mode") == "all" else dict(draft))
        self.close()

    def _on_clear(self) -> None: self.filter_applied.emit(self._col_key, None); self.close()

    def show_below(self, ref_widget: QWidget) -> None:
        self.move(ref_widget.mapToGlobal(QPoint(0, ref_widget.height() + 2))); self.show(); self.raise_(); self.activateWindow()


# ── WoRMS background worker ───────────────────────────────────────────────────

class _WormsSearchWorker(QThread):
    """Background thread for WoRMS name search + classification lookup."""

    results_ready = pyqtSignal(list)
    chain_ready = pyqtSignal(list)
    error_occurred = pyqtSignal(str)

    def __init__(self, worms_svc: "WormsService", query: str, like: bool = True, aphia_id: Optional[int] = None, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._svc = worms_svc; self._query = query; self._like = like; self._aphia_id = aphia_id

    def run(self) -> None:
        try:
            if self._aphia_id is not None:
                self.chain_ready.emit(self._svc.flatten_classification(self._svc.classification(self._aphia_id)))
            else:
                self.results_ready.emit(self._svc.search(self._query, like=self._like))
        except Exception as exc:
            self.error_occurred.emit(str(exc))


class _WormsJobWorker(QThread):
    """Drive a WoRMS batch-validation job to completion off the UI thread.

    Mirrors oracle ``runWormsJob``: for each record at the cursor, resolve it,
    match it against WoRMS, record the result, repeat until the job is no
    longer ``running`` (completed / paused / cancelled).  All progress is
    delivered via Qt signals — the worker never touches widgets.  Pausing is
    cooperative: the main thread flips the job status to ``paused`` and the
    worker exits at the next loop boundary.
    """

    progress = pyqtSignal(int, int, dict)   # cursor, total, counts
    finished_job = pyqtSignal(dict)         # final/last-seen job dict
    failed = pyqtSignal(str)

    def __init__(self, worms_svc: "WormsService", job_id: str, resolver, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._svc = worms_svc
        self._job_id = job_id
        self._resolve = resolver   # record_id -> Optional[dict]

    def run(self) -> None:
        try:
            while not self.isInterruptionRequested():
                job = self._svc.get_job(self._job_id)
                if not job or job.get("status") != "running":
                    if job:
                        self.finished_job.emit(job)
                    return
                record_ids = job.get("record_ids") or []
                cursor = job.get("cursor", 0)
                if cursor >= len(record_ids):
                    done = self._svc.update_job_status(self._job_id, "completed") or job
                    self.finished_job.emit(done)
                    return
                record = self._resolve(record_ids[cursor])
                # Unresolvable recordId → "stale" (oracle runWormsJob fallback).
                status = self._svc.match_one(record) if record else "stale"
                updated = self._svc.record_job_result(self._job_id, status) or job
                self.progress.emit(
                    updated.get("cursor", cursor + 1),
                    len(record_ids),
                    updated.get("counts", {}),
                )
                if updated.get("status") == "completed":
                    self.finished_job.emit(updated)
                    return
        except Exception as exc:
            self.failed.emit(str(exc))


# ── WoRMS match dialog (mirrors renderWormsMatchModal in app.js) ──────────────

class _WormsMatchDialog(QDialog):
    """Search WoRMS and select the correct candidate (mirrors renderWormsMatchModal)."""

    def __init__(self, row: dict[str, Any], worms_svc: "WormsService", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        _refresh_palette()
        self._row = row; self._svc = worms_svc; self._selected: Optional[dict[str, Any]] = None
        self._chain: list[dict[str, Any]] = []; self._worker: Optional[_WormsSearchWorker] = None; self._result: Optional[dict[str, Any]] = None
        self.setWindowTitle("WoRMS 匹配物种"); self.setMinimumWidth(680); self.setMinimumHeight(420)
        self._build_ui(); self._do_search()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self); layout.setContentsMargins(18, 18, 18, 14); layout.setSpacing(10)
        t = QLabel("WoRMS 匹配物种"); t.setStyleSheet(f"font-size: 16px; font-weight: 600; color: {_C_TEXT};"); layout.addWidget(t)
        o = QLabel(f"原始种名：{self._row.get('species', '')}"); o.setStyleSheet(f"color: {_C_MUTED}; font-size: 12px;"); layout.addWidget(o)
        sr = QHBoxLayout(); sr.setSpacing(8)
        self._search_input = QLineEdit(); self._search_input.setPlaceholderText("输入科学名"); self._search_input.setText(self._row.get("species", "")); self._search_input.returnPressed.connect(self._do_search); sr.addWidget(self._search_input, 1)
        self._fuzzy_check = QCheckBox("模糊匹配"); self._fuzzy_check.setStyleSheet(f"color: {_C_MUTED};"); sr.addWidget(self._fuzzy_check)
        bs = QPushButton("搜索"); bs.setObjectName("Outline"); bs.setFixedWidth(60); bs.clicked.connect(self._do_search); sr.addWidget(bs); layout.addLayout(sr)
        self._error_label = QLabel(""); self._error_label.setStyleSheet(f"color: {_C_DANGER}; font-size: 11px;"); self._error_label.hide(); layout.addWidget(self._error_label)
        self._loading_label = QLabel("正在查询 WoRMS..."); self._loading_label.setStyleSheet(f"color: {_C_ACCENT}; font-size: 12px;"); self._loading_label.hide(); layout.addWidget(self._loading_label)
        body = QSplitter(Qt.Orientation.Horizontal); body.setHandleWidth(6)
        self._results_list = QListWidget(); self._results_list.setStyleSheet(f"QListWidget {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER}; border-radius: 4px; }} QListWidget::item {{ color: {_C_TEXT}; padding: 6px 8px; }} QListWidget::item:selected {{ background: {_C_ACCENT}; }}")
        self._results_list.currentItemChanged.connect(self._on_result_selected); body.addWidget(self._results_list)
        df = QFrame(); df.setStyleSheet(f"QFrame {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER}; border-radius: 4px; }}")
        dl = QVBoxLayout(df); dl.setContentsMargins(10, 10, 10, 10); dl.setSpacing(4)
        dt = QLabel("采用后保存的 WoRMS 分类链"); dt.setStyleSheet(f"font-size: 11px; font-weight: 600; color: {_C_MUTED}; background: transparent;"); dl.addWidget(dt)
        self._chain_label = QLabel("选择候选后预览标准分类阶元"); self._chain_label.setStyleSheet(f"color: {_C_DIM}; font-size: 11px; background: transparent;"); self._chain_label.setWordWrap(True); self._chain_label.setAlignment(Qt.AlignmentFlag.AlignTop); dl.addWidget(self._chain_label)
        self._chain_loading_label = QLabel("加载分类链..."); self._chain_loading_label.setStyleSheet(f"color: {_C_ACCENT}; font-size: 11px; background: transparent;"); self._chain_loading_label.hide(); dl.addWidget(self._chain_loading_label); dl.addStretch()
        body.addWidget(df); body.setSizes([340, 300]); layout.addWidget(body, 1)
        ar = QHBoxLayout(); ar.setSpacing(8)
        ms = self._row.get("mappingStatus", ""); bl = "重新匹配并保存" if ms and ms != "unprocessed" else "采用并保存"
        self._btn_save = QPushButton(bl); self._btn_save.setObjectName("Primary"); self._btn_save.setEnabled(False); self._btn_save.clicked.connect(self._on_save); ar.addWidget(self._btn_save)
        bn = QPushButton("标记未找到"); bn.setObjectName("Outline"); bn.clicked.connect(self._on_no_match); ar.addWidget(bn)
        bc = QPushButton("取消"); bc.setObjectName("Outline"); bc.clicked.connect(self.reject); ar.addWidget(bc); ar.addStretch(); layout.addLayout(ar)

    def _do_search(self) -> None:
        query = self._search_input.text().strip()
        if not query: return
        if self._worker and self._worker.isRunning(): self._worker.terminate()
        self._selected = None; self._chain = []; self._error_label.hide(); self._loading_label.show(); self._results_list.clear(); self._chain_label.setText("选择候选后预览标准分类阶元"); self._btn_save.setEnabled(False)
        self._worker = _WormsSearchWorker(self._svc, query, like=self._fuzzy_check.isChecked(), parent=self)
        self._worker.results_ready.connect(self._on_results_ready); self._worker.error_occurred.connect(self._on_search_error); self._worker.start()

    def _on_results_ready(self, hits: list[dict[str, Any]]) -> None:
        self._loading_label.hide(); self._results_list.clear()
        if not hits:
            item = QListWidgetItem("未找到候选，请修改关键词或启用模糊匹配。"); item.setFlags(Qt.ItemFlag.NoItemFlags); item.setForeground(QColor(_C_DIM)); self._results_list.addItem(item); return
        for rec in hits:
            name = rec.get("valid_name") or rec.get("scientificname") or ""; aphia = rec.get("valid_AphiaID") or rec.get("AphiaID", ""); status = rec.get("status", "")
            chain_str = " > ".join(p for p in [rec.get("class"), rec.get("order"), rec.get("family"), rec.get("genus")] if p)
            item = QListWidgetItem(f"{name}\n{status} · AphiaID {aphia}\n{chain_str}"); item.setData(Qt.ItemDataRole.UserRole, rec); self._results_list.addItem(item)

    def _on_search_error(self, msg: str) -> None:
        self._loading_label.hide(); self._error_label.setText(f"搜索失败：{msg}"); self._error_label.show()

    def _on_result_selected(self, current: Optional[QListWidgetItem], _: Optional[QListWidgetItem]) -> None:
        if current is None: return
        rec = current.data(Qt.ItemDataRole.UserRole)
        if rec is None: return
        self._selected = rec; aphia_id = rec.get("valid_AphiaID") or rec.get("AphiaID")
        if not aphia_id: return
        self._chain_label.hide(); self._chain_loading_label.show(); self._btn_save.setEnabled(False)
        w = _WormsSearchWorker(self._svc, "", aphia_id=int(aphia_id), parent=self)
        w.chain_ready.connect(self._on_chain_ready); w.error_occurred.connect(self._on_chain_error); w.start()

    def _on_chain_ready(self, chain: list[dict[str, Any]]) -> None:
        self._chain = chain; self._chain_loading_label.hide(); self._chain_label.show()
        self._chain_label.setText("\n".join(f"{n.get('rank','')}  {n.get('scientificname','')}" for n in chain) or "（无分类链数据）")
        self._btn_save.setEnabled(True)

    def _on_chain_error(self, msg: str) -> None:
        self._chain_loading_label.hide(); self._chain_label.show(); self._chain_label.setText(f"分类链加载失败：{msg}"); self._btn_save.setEnabled(bool(self._selected))

    def _on_save(self) -> None:
        if self._selected is None: return
        aphia = self._selected.get("valid_AphiaID") or self._selected.get("AphiaID")
        self._result = {"aphia_id": int(aphia) if aphia else None, "worms_record": self._selected, "chain": self._chain}; self.accept()

    def _on_no_match(self) -> None: self._result = {"no_match": True}; self.accept()

    def get_result(self) -> Optional[dict[str, Any]]: return self._result


# ── WoRMS review dialog (mirrors renderTaxonReviewModal in app.js) ─────────────

class _TaxonReviewDialog(QDialog):
    """Review auto-found WoRMS candidates (mirrors renderTaxonReviewModal in app.js)."""

    def __init__(self, row: dict[str, Any], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        _refresh_palette()
        self._row = row; self._result: Optional[dict[str, Any]] = None
        self.setWindowTitle("审核 WoRMS 匹配"); self.setMinimumWidth(440); self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self); layout.setContentsMargins(18, 18, 18, 14); layout.setSpacing(10)
        t = QLabel("审核 WoRMS 匹配"); t.setStyleSheet(f"font-size: 16px; font-weight: 600; color: {_C_TEXT};"); layout.addWidget(t)
        o = QLabel(f"原始种名：{self._row.get('species', '')}"); o.setStyleSheet(f"color: {_C_MUTED}; font-size: 12px;"); layout.addWidget(o)
        candidates = self._row.get("mappingCandidates") or []
        if not candidates:
            e = QLabel("没有自动候选，可标记为未找到后重新运行更新。"); e.setStyleSheet(f"color: {_C_DIM}; font-size: 12px;"); e.setWordWrap(True); layout.addWidget(e)
        else:
            for cand in candidates:
                rf = QFrame(); rf.setStyleSheet(f"QFrame {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER}; border-radius: 4px; }}")
                rl = QHBoxLayout(rf); rl.setContentsMargins(10, 8, 10, 8); rl.setSpacing(8)
                aphia = cand.get("valid_AphiaID") or cand.get("AphiaID") or ""; name = cand.get("valid_name") or cand.get("scientificname") or ""
                info = QLabel(f"{name} · AphiaID {aphia}"); info.setStyleSheet(f"color: {_C_TEXT}; font-size: 12px; background: transparent;"); rl.addWidget(info, 1)
                bu = QPushButton("采用"); bu.setObjectName("Primary"); bu.setFixedWidth(56)
                bu.clicked.connect(lambda _=False, c=cand: self._on_use(c)); rl.addWidget(bu); layout.addWidget(rf)
        ar = QHBoxLayout(); ar.setSpacing(8)
        bn = QPushButton("标记未找到"); bn.setObjectName("Outline"); bn.clicked.connect(self._on_no_match); ar.addWidget(bn)
        bc = QPushButton("关闭"); bc.setObjectName("Outline"); bc.clicked.connect(self.reject); ar.addWidget(bc); ar.addStretch(); layout.addLayout(ar)

    def _on_use(self, cand: dict[str, Any]) -> None:
        aphia = cand.get("valid_AphiaID") or cand.get("AphiaID")
        self._result = {"aphia_id": int(aphia) if aphia else None}; self.accept()

    def _on_no_match(self) -> None: self._result = {"no_match": True}; self.accept()

    def get_result(self) -> Optional[dict[str, Any]]: return self._result


# ── Column-group chip button (taxon-col-chip style) ───────────────────────────

class _ChipButton(QPushButton):
    """Toggle chip matching .taxon-col-chip in styles.css."""

    def __init__(self, text: str, checked: bool = True, parent: Optional[QWidget] = None) -> None:
        super().__init__(text, parent)
        _refresh_palette()
        self.setCheckable(True)
        self.setChecked(checked)
        self._refresh_style()
        self.toggled.connect(lambda _: self._refresh_style())

    def _refresh_style(self) -> None:
        if self.isChecked():
            self.setStyleSheet(
                f"QPushButton {{ background: {_C_ACCENT}; color: {_C_INPUT}; font-weight: 600;"
                f" border: none; border-radius: 6px; padding: 4px 10px; font-size: 12px; }}"
                f"QPushButton:hover {{ background: {_C_ACCENT_HI}; }}"
            )
        else:
            self.setStyleSheet(
                f"QPushButton {{ background: {_C_PANEL}; color: {_C_MUTED};"
                f" border: 1px solid {_C_BORDER}; border-radius: 6px;"
                f" padding: 4px 10px; font-size: 12px; }}"
                f"QPushButton:hover {{ color: {_C_TEXT}; border-color: {_C_ACCENT}; }}"
            )


# ── View-switch button (taxon-view-btn style) ─────────────────────────────────

class _ViewTabButton(QPushButton):
    """Segmented-switch tab button matching .taxon-view-btn."""

    def __init__(self, text: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(text, parent)
        _refresh_palette()
        self.setCheckable(True)
        self._refresh_style()
        self.toggled.connect(lambda _: self._refresh_style())

    def _refresh_style(self) -> None:
        if self.isChecked():
            self.setStyleSheet(
                f"QPushButton {{ background: {_C_ACCENT}; color: {_C_INPUT}; font-weight: 600;"
                f" border: none; border-radius: 6px; padding: 6px 14px; font-size: 12px; }}"
            )
        else:
            self.setStyleSheet(
                f"QPushButton {{ background: transparent; color: {_C_MUTED};"
                f" border: none; border-radius: 6px; padding: 6px 14px; font-size: 12px; }}"
                f"QPushButton:hover {{ color: {_C_TEXT}; background: {_C_ACCENT_SOFT}; }}"
            )


# ── 操作 column delegate — renders inline "编辑" / "删除" buttons ──────────────

class _ActionDelegate(QStyledItemDelegate):
    """Renders per-row 编辑 + 删除 buttons in the 操作 column.

    Signals (emitted with the record dict as argument):
      edit_requested(dict)
      delete_requested(dict)
    """

    edit_requested: pyqtSignal = pyqtSignal(dict)
    delete_requested: pyqtSignal = pyqtSignal(dict)

    _BTN_W = 42
    _BTN_H = 22
    _GAP   = 4

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)

    def _btn_rects(self, option: QStyleOptionViewItem) -> tuple[Any, Any]:
        """Return (edit_rect, del_rect) for the given option geometry."""
        from PyQt6.QtCore import QRect
        r = option.rect
        y = r.y() + (r.height() - self._BTN_H) // 2
        x_edit = r.x() + 4
        x_del  = x_edit + self._BTN_W + self._GAP
        return (
            QRect(x_edit, y, self._BTN_W, self._BTN_H),
            QRect(x_del,  y, self._BTN_W, self._BTN_H),
        )

    def paint(self, painter: Any, option: QStyleOptionViewItem, index: QModelIndex) -> None:  # type: ignore[override]
        rec = index.data(Qt.ItemDataRole.UserRole)
        if rec is None:
            return
        is_user = rec.get("recordId", "").startswith("user:")
        edit_rect, del_rect = self._btn_rects(option)

        from PyQt6.QtGui import QPainter, QBrush, QPen, QFont
        from PyQt6.QtCore import QRect, Qt as _Qt

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        def _draw_btn(rect: Any, text: str, fg: str, border: str) -> None:
            painter.setPen(QPen(QColor(border)))
            painter.setBrush(QBrush(QColor(0, 0, 0, 0)))
            painter.drawRoundedRect(rect, 4, 4)
            painter.setPen(QPen(QColor(fg)))
            f = painter.font()
            f.setPixelSize(11)
            painter.setFont(f)
            painter.drawText(rect, _Qt.AlignmentFlag.AlignCenter, text)

        # Edit button — always shown in original view
        _draw_btn(edit_rect, "编辑", _C_MUTED, _C_BORDER)

        # Delete button — only for user records
        if is_user:
            _draw_btn(del_rect, "删除", _C_DANGER, _C_DANGER_SOFT)

        painter.restore()

    def editorEvent(
        self,
        event: Any,
        model: Any,
        option: QStyleOptionViewItem,
        index: QModelIndex,
    ) -> bool:
        from PyQt6.QtCore import QEvent
        if event.type() != QEvent.Type.MouseButtonRelease:
            return False
        rec = index.data(Qt.ItemDataRole.UserRole)
        if rec is None:
            return False
        is_user = rec.get("recordId", "").startswith("user:")
        edit_rect, del_rect = self._btn_rects(option)
        pos = event.pos()
        if edit_rect.contains(pos):
            self.edit_requested.emit(rec)
            return True
        if is_user and del_rect.contains(pos):
            self.delete_requested.emit(rec)
            return True
        return False

    def sizeHint(self, option: QStyleOptionViewItem, index: QModelIndex) -> Any:
        from PyQt6.QtCore import QSize
        return QSize(self._BTN_W * 2 + self._GAP + 12, 30)


# ── Main view ─────────────────────────────────────────────────────────────────

class TaxonomyView(BaseView):
    """Taxonomy library management view (view_id="taxonomy").

    Faithful replica of the web 内置分类库 page.

    Layout (matches DOM):

        [QScrollArea — controls area, prevents top from squishing table]
          taxon-table-header
            taxon-table-title-row  (h2 + stats + taxon-view-switch + 图表-btn)
            taxon-col-controls     (类群 chips + 语言 chips)  [original only]
          taxon-table-toolbar
            taxon-table-filter-bar (col-select + search-input + 搜索 + 清除)
            taxon-table-actions    (+ 新增 | selection note | 全选 | 取消 |
                                    WoRMS更新所选 | WoRMS更新筛选 | 导出Excel |
                                    导出CSV | 导入Excel/CSV)
        taxon-table-wrap
          QTableView  (☑ # <dynamic cols> 来源 操作)
          [loading overlay label when _loading=True]
        taxon-table-pager
          上一页 | 第N/M页（共K条） | 跳到 [spin] | 下一页  | footer stats
    """

    view_id = "taxonomy"
    nav_title = "内置分类库"
    nav_icon = "🧬"

    def __init__(self, ctx: "AppContext") -> None:
        # View state — mirrors state.taxonTable in app.js
        self._view: str = "original"          # "original" | "worms" | "compare"
        self._show_chart: bool = False
        self._page: int = 1
        self._total: int = 0
        self._selected_ids: list[str] = []
        self._select_all_filtered: bool = False
        self._filter_col: str = ""
        self._filter_text: str = ""
        self._loading: bool = False
        self._svc: Optional[TaxonomyService] = None
        # New state for facet/sort/WoRMS
        self._worms_svc: Optional[WormsService] = None
        self._job_worker: Optional["_WormsJobWorker"] = None
        self._chart_dialog: Optional[QDialog] = None
        self._col_filters: dict[str, Optional[dict[str, Any]]] = {}
        self._sort_col: str = ""
        self._sort_dir: str = "asc"
        self._facet_panel: Optional[_TaxonFacetPanel] = None
        super().__init__(ctx)

    # ── BaseView._setup_ui ────────────────────────────────────────────

    def _setup_ui(self) -> None:
        _refresh_palette()
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Scrollable controls area (header + toolbar) ───────────────
        # Wrapped in QScrollArea so it never squishes the table.
        ctrl_container = QWidget()
        ctrl_container.setObjectName("TaxonCtrlContainer")
        ctrl_container.setStyleSheet(
            "QWidget#TaxonCtrlContainer { background: transparent; }"
        )
        ctrl_v = QVBoxLayout(ctrl_container)
        ctrl_v.setContentsMargins(0, 0, 0, 0)
        ctrl_v.setSpacing(0)

        # ── Header (taxon-table-header) ───────────────────────────────
        header_frame = QFrame()
        header_frame.setObjectName("TaxonHeader")
        header_frame.setStyleSheet(
            f"QFrame#TaxonHeader {{ border: none;"
            f" border-bottom: 1px solid {_C_BORDER}; }}"
        )
        header_layout = QVBoxLayout(header_frame)
        header_layout.setContentsMargins(24, 12, 24, 0)
        header_layout.setSpacing(8)

        # taxon-table-title-row
        title_row = QHBoxLayout()
        title_row.setSpacing(14)

        page_title = QLabel("内置分类库")
        page_title.setObjectName("TaxonPageTitle")
        page_title.setStyleSheet(
            f"QLabel {{ font-size: 18px; font-weight: 600; color: {_C_TEXT};"
            f" background: transparent; }}"
        )
        title_row.addWidget(page_title)

        self._stats_label = QLabel("共 0 条")
        self._stats_label.setObjectName("Muted")
        self._stats_label.setStyleSheet(
            f"QLabel {{ font-size: 13px; color: {_C_MUTED}; background: transparent; }}"
        )
        title_row.addWidget(self._stats_label)

        title_row.addStretch()

        # taxon-view-switch (3 tabs + "图表" toggle) in a capsule frame
        switch_wrapper = QFrame()
        switch_wrapper.setStyleSheet(
            f"QFrame {{ background: {_C_PANEL}; border-radius: 8px; padding: 2px; }}"
        )
        switch_inner = QHBoxLayout(switch_wrapper)
        switch_inner.setContentsMargins(3, 3, 3, 3)
        switch_inner.setSpacing(2)

        self._btn_original = _ViewTabButton("原始分类")
        self._btn_worms    = _ViewTabButton("WoRMS 分类")
        self._btn_compare  = _ViewTabButton("对照视图")
        self._btn_original.setChecked(True)

        for btn, view_key in [
            (self._btn_original, "original"),
            (self._btn_worms,    "worms"),
            (self._btn_compare,  "compare"),
        ]:
            btn.clicked.connect(lambda checked, k=view_key: self._on_view_switch(k))
            switch_inner.addWidget(btn)

        title_row.addWidget(switch_wrapper)

        # "图表" toggle button — only active in "original" view (mirrors app.js)
        self._btn_chart = _ViewTabButton("图表")
        self._btn_chart.setChecked(False)
        self._btn_chart.clicked.connect(self._on_chart_toggle)
        title_row.addWidget(self._btn_chart)

        header_layout.addLayout(title_row)

        # taxon-col-controls (only in "original" view)
        self._col_ctrl_frame = QFrame()
        col_ctrl_layout = QHBoxLayout(self._col_ctrl_frame)
        col_ctrl_layout.setContentsMargins(0, 2, 0, 6)
        col_ctrl_layout.setSpacing(16)

        # 类群 group
        level_group = QHBoxLayout()
        level_group.setSpacing(6)
        level_label = QLabel("类群")
        level_label.setStyleSheet(
            f"QLabel {{ color: {_C_DIM}; font-size: 11px; font-weight: 600;"
            f" background: transparent; }}"
        )
        level_group.addWidget(level_label)

        self._level_chips: dict[str, _ChipButton] = {}
        for level_key, level_text in _LEVEL_CHIPS:
            chip = _ChipButton(level_text, checked=True)
            chip.toggled.connect(
                lambda checked, k=level_key: self._on_level_chip(k, checked)
            )
            level_group.addWidget(chip)
            self._level_chips[level_key] = chip
        col_ctrl_layout.addLayout(level_group)

        # 语言 group
        lang_group = QHBoxLayout()
        lang_group.setSpacing(6)
        lang_label = QLabel("语言")
        lang_label.setStyleSheet(
            f"QLabel {{ color: {_C_DIM}; font-size: 11px; font-weight: 600;"
            f" background: transparent; }}"
        )
        lang_group.addWidget(lang_label)

        self._lang_chips: dict[str, _ChipButton] = {}
        for lang_key, lang_text in _LANG_CHIPS:
            chip = _ChipButton(lang_text, checked=True)
            chip.toggled.connect(
                lambda checked, k=lang_key: self._on_lang_chip(k, checked)
            )
            lang_group.addWidget(chip)
            self._lang_chips[lang_key] = chip
        col_ctrl_layout.addLayout(lang_group)
        col_ctrl_layout.addStretch()

        # ── 字体 / 行高 调节 ─────────────────────────────────────────
        font_group = QHBoxLayout()
        font_group.setSpacing(4)
        font_lbl = QLabel("字号")
        font_lbl.setStyleSheet(
            f"QLabel {{ color: {_C_DIM}; font-size: 11px; font-weight: 600;"
            f" background: transparent; }}"
        )
        font_group.addWidget(font_lbl)

        self._font_size: int = 12
        btn_font_minus = QPushButton("−")
        btn_font_minus.setFixedSize(22, 22)
        btn_font_minus.setStyleSheet(
            f"QPushButton {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER};"
            f" border-radius: 4px; color: {_C_TEXT_SOFT}; font-size: 13px; font-weight: bold; }}"
            f"QPushButton:hover {{ border-color: {_C_ACCENT}; color: {_C_ACCENT}; }}"
        )
        self._font_size_lbl = QLabel("12")
        self._font_size_lbl.setFixedWidth(22)
        self._font_size_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._font_size_lbl.setStyleSheet(f"color: {_C_TEXT_SOFT}; font-size: 11px; background: transparent;")
        btn_font_plus = QPushButton("+")
        btn_font_plus.setFixedSize(22, 22)
        btn_font_plus.setStyleSheet(btn_font_minus.styleSheet())

        btn_font_minus.clicked.connect(lambda: self._adjust_font(-1))
        btn_font_plus.clicked.connect(lambda: self._adjust_font(1))

        font_group.addWidget(btn_font_minus)
        font_group.addWidget(self._font_size_lbl)
        font_group.addWidget(btn_font_plus)
        col_ctrl_layout.addLayout(font_group)

        header_layout.addWidget(self._col_ctrl_frame)
        ctrl_v.addWidget(header_frame)

        # ── Toolbar (taxon-table-toolbar) ─────────────────────────────
        toolbar_frame = QFrame()
        toolbar_frame.setObjectName("TaxonToolbar")
        toolbar_frame.setStyleSheet(
            f"QFrame#TaxonToolbar {{ border: none;"
            f" border-bottom: 1px solid {_C_BORDER}; }}"
        )
        toolbar_layout = QVBoxLayout(toolbar_frame)
        toolbar_layout.setContentsMargins(24, 8, 24, 8)
        toolbar_layout.setSpacing(8)

        # filter bar (taxon-table-filter-bar)
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)

        self._col_select = QComboBox()
        self._col_select.setFixedWidth(116)
        for val, label in [
            ("",           "全部列"),
            ("classCn",    "纲(中)"),
            ("class",      "纲(拉丁)"),
            ("orderCn",    "目(中)"),
            ("order",      "目(拉丁)"),
            ("familyCn",   "科(中)"),
            ("family",     "科(拉丁)"),
            ("genusCn",    "属(中)"),
            ("genus",      "属(拉丁)"),
            ("speciesCn",  "种(中)"),
            ("species",    "种(拉丁)"),
        ]:
            self._col_select.addItem(label, val)
        filter_row.addWidget(self._col_select)

        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("搜索…")
        self._search_input.setMinimumWidth(180)
        self._search_input.returnPressed.connect(self._on_search)
        filter_row.addWidget(self._search_input)

        btn_search = QPushButton("搜索")
        btn_search.setObjectName("Outline")
        btn_search.setFixedWidth(60)
        btn_search.clicked.connect(self._on_search)
        filter_row.addWidget(btn_search)

        btn_clear = QPushButton("清除")
        btn_clear.setObjectName("Outline")
        btn_clear.setFixedWidth(60)
        btn_clear.clicked.connect(self._on_clear_filter)
        filter_row.addWidget(btn_clear)

        self._filter_active_label = QLabel("")
        self._filter_active_label.setStyleSheet(
            f"QLabel {{ color: {_C_ACCENT}; font-size: 11px; background: transparent; }}"
        )
        self._filter_active_label.hide()
        filter_row.addWidget(self._filter_active_label)

        filter_row.addStretch()
        toolbar_layout.addLayout(filter_row)

        # action bar (taxon-table-actions)
        action_row = QHBoxLayout()
        action_row.setSpacing(8)

        self._btn_add = QPushButton("+ 新增条目")
        self._btn_add.setObjectName("Primary")
        self._btn_add.setFixedWidth(96)
        self._btn_add.clicked.connect(self._on_add)
        action_row.addWidget(self._btn_add)

        # Thin separator
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet(f"QFrame {{ color: {_C_BORDER}; }}")
        action_row.addWidget(sep)

        self._selection_note = QLabel("已选 0 条")
        self._selection_note.setStyleSheet(
            f"QLabel {{ color: {_C_MUTED}; font-size: 11px; background: transparent; }}"
        )
        action_row.addWidget(self._selection_note)

        btn_sel_all = QPushButton("全选筛选结果")
        btn_sel_all.setObjectName("Outline")
        btn_sel_all.clicked.connect(self._on_select_all_filtered)
        action_row.addWidget(btn_sel_all)

        btn_desel = QPushButton("取消选择")
        btn_desel.setObjectName("Outline")
        btn_desel.clicked.connect(self._on_deselect)
        action_row.addWidget(btn_desel)

        # Thin separator
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.VLine)
        sep2.setStyleSheet(f"QFrame {{ color: {_C_BORDER}; }}")
        action_row.addWidget(sep2)

        self._btn_worms_sel = QPushButton("WoRMS 更新所选")
        self._btn_worms_sel.setObjectName("Primary")
        self._btn_worms_sel.setEnabled(False)
        self._btn_worms_sel.clicked.connect(
            lambda: self._on_worms_update(selected_only=True)
        )
        action_row.addWidget(self._btn_worms_sel)

        btn_worms_filt = QPushButton("WoRMS 更新筛选结果")
        btn_worms_filt.setObjectName("Outline")
        btn_worms_filt.clicked.connect(
            lambda: self._on_worms_update(selected_only=False)
        )
        action_row.addWidget(btn_worms_filt)

        # Thin separator
        sep3 = QFrame()
        sep3.setFrameShape(QFrame.Shape.VLine)
        sep3.setStyleSheet(f"QFrame {{ color: {_C_BORDER}; }}")
        action_row.addWidget(sep3)

        btn_export_xlsx = QPushButton("导出 Excel")
        btn_export_xlsx.setObjectName("Outline")
        btn_export_xlsx.clicked.connect(lambda: self._on_export("xlsx"))
        action_row.addWidget(btn_export_xlsx)

        btn_export_csv = QPushButton("导出 CSV")
        btn_export_csv.setObjectName("Outline")
        btn_export_csv.clicked.connect(lambda: self._on_export("csv"))
        action_row.addWidget(btn_export_csv)

        # 导入 Excel/CSV — file-picker button (taxon-import-label in DOM)
        btn_import = QPushButton("导入 Excel/CSV")
        btn_import.setObjectName("Outline")
        btn_import.clicked.connect(self._on_import)
        action_row.addWidget(btn_import)

        action_row.addStretch()
        toolbar_layout.addLayout(action_row)
        ctrl_v.addWidget(toolbar_frame)

        # Scroll area wrapping the controls
        ctrl_scroll = QScrollArea()
        ctrl_scroll.setWidget(ctrl_container)
        ctrl_scroll.setWidgetResizable(True)
        ctrl_scroll.setFrameShape(QFrame.Shape.NoFrame)
        ctrl_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        ctrl_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        # Cap scroll area height so it never steals table space
        ctrl_scroll.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum
        )
        ctrl_scroll.setMaximumHeight(200)
        ctrl_scroll.setStyleSheet(
            "QScrollArea { background: transparent; border: none; }"
        )
        root.addWidget(ctrl_scroll)

        # ── Table area (taxon-table-wrap) ─────────────────────────────
        self._model = _TaxonTableModel(self)

        self._table = QTableView()
        self._table.setModel(self._model)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(False)  # server-side sort via API
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setStretchLastSection(False)
        self._table.horizontalHeader().setSectionsMovable(True)   # drag to reorder columns
        self._table.horizontalHeader().setMinimumSectionSize(40)
        self._table.verticalHeader().setVisible(False)
        self._table.doubleClicked.connect(self._on_row_double_click)
        self._table.selectionModel().selectionChanged.connect(self._on_selection_changed)
        # Checkbox toggles update the "已选 N 条" note + enable WoRMS-update button,
        # independent of row selection (clicking a checkbox does not select a full row).
        self._model.checked_changed.connect(self._update_selection_note)
        # Row height (default 32px, scales with font size)
        self._table.verticalHeader().setDefaultSectionSize(32)

        # Column widths for checkbox / # columns
        self._table.setColumnWidth(_COL_CHECK, 32)
        self._table.setColumnWidth(_COL_NUM,   40)
        # Action column delegate
        self._action_delegate = _ActionDelegate(self._table)
        self._action_delegate.edit_requested.connect(self._edit_record)
        self._action_delegate.delete_requested.connect(self._delete_record)

        # Ctrl+C → copy selected cells in Excel format (tab-separated cols, newline rows)
        self._table.keyPressEvent = self._table_key_press

        # Right-click context menu (mirrors openTaxonRowMenu in app.js)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_row_context_menu)

        # Column header right-click → facet filter (mirrors openTaxonFacetMenu)
        self._table.horizontalHeader().setContextMenuPolicy(
            Qt.ContextMenuPolicy.CustomContextMenu
        )
        self._table.horizontalHeader().customContextMenuRequested.connect(
            self._on_header_context_menu
        )

        # Table in a wrapper frame for the loading overlay
        table_wrapper = QFrame()
        table_wrapper.setObjectName("TaxonTableWrap")
        table_wrapper.setStyleSheet(
            "QFrame#TaxonTableWrap { border: none; }"
        )
        table_wrap_layout = QVBoxLayout(table_wrapper)
        table_wrap_layout.setContentsMargins(0, 0, 0, 0)
        table_wrap_layout.setSpacing(0)

        # Loading indicator (taxon-table-loading)
        self._loading_label = QLabel("加载中…")
        self._loading_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._loading_label.setStyleSheet(
            f"QLabel {{ color: {_C_ACCENT}; font-size: 14px; background: transparent;"
            f" padding: 24px; }}"
        )
        self._loading_label.hide()

        table_wrap_layout.addWidget(self._loading_label)
        table_wrap_layout.addWidget(self._table, 1)

        root.addWidget(table_wrapper, 1)

        # ── WoRMS job panel (mirrors renderTaxonJobPanel in app.js) ──────────
        self._job_panel_frame = QFrame()
        self._job_panel_frame.setObjectName("TaxonJobPanel")
        self._job_panel_frame.setStyleSheet(
            f"QFrame#TaxonJobPanel {{ background: {_C_PANEL};"
            f" border: 1px solid {_C_BORDER}; border-radius: 6px;"
            f" margin: 4px 24px; }}"
        )
        _jpanel_layout = QHBoxLayout(self._job_panel_frame)
        _jpanel_layout.setContentsMargins(12, 8, 12, 8)
        _jpanel_layout.setSpacing(12)
        self._job_title_label = QLabel("WoRMS 任务")
        self._job_title_label.setStyleSheet(f"font-weight: 600; color: {_C_ACCENT}; font-size: 12px; background: transparent;")
        _jpanel_layout.addWidget(self._job_title_label)
        self._job_progress_label = QLabel("")
        self._job_progress_label.setStyleSheet(f"color: {_C_MUTED}; font-size: 11px; background: transparent;")
        _jpanel_layout.addWidget(self._job_progress_label)
        self._job_bar = QProgressBar()
        self._job_bar.setFixedHeight(8)
        self._job_bar.setTextVisible(False)
        self._job_bar.setStyleSheet(
            f"QProgressBar {{ background: {_C_INPUT}; border: none; border-radius: 4px; }}"
            f"QProgressBar::chunk {{ background: {_C_ACCENT}; border-radius: 4px; }}"
        )
        self._job_bar.setFixedWidth(120)
        _jpanel_layout.addWidget(self._job_bar)
        self._job_counts_label = QLabel("")
        self._job_counts_label.setStyleSheet(f"color: {_C_MUTED}; font-size: 11px; background: transparent;")
        _jpanel_layout.addWidget(self._job_counts_label)
        self._btn_job_pause = QPushButton("暂停")
        self._btn_job_pause.setObjectName("Outline")
        self._btn_job_pause.setFixedHeight(24)
        self._btn_job_pause.hide()
        _jpanel_layout.addWidget(self._btn_job_pause)
        self._btn_job_resume = QPushButton("继续")
        self._btn_job_resume.setObjectName("Outline")
        self._btn_job_resume.setFixedHeight(24)
        self._btn_job_resume.hide()
        _jpanel_layout.addWidget(self._btn_job_resume)
        self._btn_job_retry = QPushButton("重试失败")
        self._btn_job_retry.setObjectName("Outline")
        self._btn_job_retry.setFixedHeight(24)
        self._btn_job_retry.hide()
        _jpanel_layout.addWidget(self._btn_job_retry)
        _jpanel_layout.addStretch()
        self._btn_job_pause.clicked.connect(self._on_job_pause)
        self._btn_job_resume.clicked.connect(self._on_job_resume)
        self._btn_job_retry.clicked.connect(self._on_job_retry)
        self._job_panel_frame.hide()
        root.addWidget(self._job_panel_frame)

        # ── Pager (taxon-table-pager) ─────────────────────────────────
        pager_frame = QFrame()
        pager_frame.setObjectName("TaxonPager")
        pager_frame.setStyleSheet(
            f"QFrame#TaxonPager {{ border: none;"
            f" border-top: 1px solid {_C_BORDER}; }}"
        )
        pager_layout = QHBoxLayout(pager_frame)
        pager_layout.setContentsMargins(24, 10, 24, 10)
        pager_layout.setSpacing(12)
        pager_frame.setFixedHeight(48)

        self._btn_prev = QPushButton("上一页")
        self._btn_prev.setObjectName("Outline")
        self._btn_prev.setFixedWidth(76)
        self._btn_prev.clicked.connect(self._on_prev_page)
        pager_layout.addWidget(self._btn_prev)

        self._page_info = QLabel("第 1 / 1 页（共 0 条）")
        self._page_info.setObjectName("TaxonPageInfo")
        self._page_info.setStyleSheet(
            f"QLabel {{ color: {_C_MUTED}; font-size: 12px; background: transparent; }}"
        )
        pager_layout.addWidget(self._page_info)

        jump_label = QLabel("跳到")
        jump_label.setStyleSheet(
            f"QLabel {{ color: {_C_DIM}; font-size: 12px; background: transparent; }}"
        )
        pager_layout.addWidget(jump_label)

        self._page_jump = QSpinBox()
        self._page_jump.setMinimum(1)
        self._page_jump.setMaximum(9999)
        self._page_jump.setValue(1)
        self._page_jump.setFixedWidth(68)
        self._page_jump.editingFinished.connect(self._on_jump_page)
        pager_layout.addWidget(self._page_jump)

        self._btn_next = QPushButton("下一页")
        self._btn_next.setObjectName("Outline")
        self._btn_next.setFixedWidth(76)
        self._btn_next.clicked.connect(self._on_next_page)
        pager_layout.addWidget(self._btn_next)

        pager_layout.addStretch()

        # stats summary at right of pager
        self._footer_label = QLabel("")
        self._footer_label.setStyleSheet(
            f"QLabel {{ color: {_C_DIM}; font-size: 11px; background: transparent; }}"
        )
        pager_layout.addWidget(self._footer_label)

        root.addWidget(pager_frame)

        # ── Init service ──────────────────────────────────────────────
        self._try_init_service()

    # ── Service init ──────────────────────────────────────────────────────────

    def _try_init_service(self) -> None:
        candidates = [
            (
                Path(__file__).parent.parent.parent.parent
                / "photo-platform-ydy"
                / "prototype-photo-gui"
                / "data"
                / "taxonomy_seed.json",
                _PROJECT_ROOT / "data" / "user_taxonomy.json",
            ),
            (_DEFAULT_SEED_PATH, _DEFAULT_USER_PATH),
        ]
        for seed_p, user_p in candidates:
            if seed_p.exists():
                self._svc = TaxonomyService(seed_p, user_p)
                return
        _DATA_DIR.mkdir(parents=True, exist_ok=True)
        self._svc = TaxonomyService(_DEFAULT_SEED_PATH, _DEFAULT_USER_PATH)

    # ── BaseView contract ─────────────────────────────────────────────────────

    def on_activate(self) -> None:
        if self._svc:
            self._svc.reload()
        self._load_page()

    def stop_background_work(self) -> None:
        """Interrupt the WoRMS batch-job worker so it cannot keep a QThread +
        its DB reads alive past app exit (the must-reboot lock-leak path)."""
        w = getattr(self, "_job_worker", None)
        if w is not None and w.isRunning():
            w.requestInterruption()
            w.wait(2000)

    # ── View-switch ───────────────────────────────────────────────────────────

    def _on_view_switch(self, view_key: str) -> None:
        self._view = view_key
        self._page = 1
        self._selected_ids.clear()
        self._select_all_filtered = False
        self._model.clear_checked()

        # Sync button checked states (avoid recursive toggling)
        self._btn_original.blockSignals(True)
        self._btn_worms.blockSignals(True)
        self._btn_compare.blockSignals(True)
        self._btn_original.setChecked(view_key == "original")
        self._btn_worms.setChecked(view_key == "worms")
        self._btn_compare.setChecked(view_key == "compare")
        self._btn_original.blockSignals(False)
        self._btn_worms.blockSignals(False)
        self._btn_compare.blockSignals(False)

        # col controls and add/chart only in original view
        in_original = view_key == "original"
        self._col_ctrl_frame.setVisible(in_original)
        self._btn_add.setVisible(in_original)
        self._btn_chart.setVisible(in_original)

        # Action column delegate — only in original view
        action_col = _COL_DATA_START + len(self._model.columns()) + 1
        if in_original:
            self._table.setItemDelegateForColumn(action_col, self._action_delegate)
        else:
            self._table.setItemDelegateForColumn(action_col, None)

        self._load_page()

    # ── Chart toggle ──────────────────────────────────────────────────────────

    def _on_chart_toggle(self) -> None:
        """Toggle chart dialog (mirrors renderTaxonChart in app.js)."""
        self._show_chart = self._btn_chart.isChecked()
        if self._show_chart:
            self._open_chart_dialog()
        elif self._chart_dialog is not None:
            self._chart_dialog.close()

    def _chart_entries(self) -> list[tuple[str, int]]:
        """Return Top-12 order buckets (mirrors web renderTaxonChart)."""
        if self._svc is None:
            return []
        rows, total = self._svc.all_records(page=0, page_size=1_000_000)
        if len(rows) < total:
            rows, _ = self._svc.all_records(page=0, page_size=max(total, 1))
        buckets: dict[str, int] = {}
        for rec in rows:
            if self._filter_text and not self._record_matches_filter(rec):
                continue
            label = str(rec.get("orderCn") or rec.get("order") or "—").strip() or "—"
            buckets[label] = buckets.get(label, 0) + 1
        return sorted(buckets.items(), key=lambda item: (-item[1], item[0]))[:12]

    def _record_matches_filter(self, rec: dict[str, Any]) -> bool:
        needle = self._filter_text.strip().lower()
        if not needle:
            return True
        if self._filter_col:
            return needle in str(rec.get(self._filter_col, "")).lower()
        return any(needle in str(rec.get(col["key"], "")).lower() for col in _ALL_COLS)

    def _open_chart_dialog(self) -> None:
        entries = self._chart_entries()
        if self._chart_dialog is not None:
            self._chart_dialog.close()
        dlg = QDialog(self)
        dlg.setWindowTitle("分类群分布图表")
        dlg.setMinimumWidth(520)
        dlg.setMinimumHeight(360)
        dlg.finished.connect(self._on_chart_dialog_finished)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(18, 18, 18, 14)
        layout.setSpacing(12)
        title = QLabel("按目统计")
        title.setStyleSheet(f"font-size: 16px; font-weight: 600; color: {_C_TEXT};")
        layout.addWidget(title)
        subtitle = QLabel("显示当前筛选条件下数量最多的前 12 个目。")
        subtitle.setStyleSheet(f"font-size: 12px; color: {_C_MUTED};")
        layout.addWidget(subtitle)
        body = QFrame()
        body.setStyleSheet(
            f"QFrame {{ background: {_C_PANEL}; border: 1px solid {_C_BORDER}; border-radius: 8px; }}"
        )
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(14, 14, 14, 14)
        body_layout.setSpacing(10)
        if entries:
            max_count = max(count for _, count in entries)
            for label, count in entries:
                row_layout = QHBoxLayout()
                row_layout.setSpacing(10)
                name = QLabel(label)
                name.setMinimumWidth(130)
                name.setStyleSheet(f"color: {_C_TEXT}; font-size: 12px;")
                row_layout.addWidget(name)
                bar = QProgressBar()
                bar.setRange(0, max_count)
                bar.setValue(count)
                bar.setTextVisible(False)
                bar.setFixedHeight(12)
                bar.setStyleSheet(
                    f"QProgressBar {{ background: {_C_INPUT}; border: none; border-radius: 6px; }}"
                    f"QProgressBar::chunk {{ background: {_C_ACCENT}; border-radius: 6px; }}"
                )
                row_layout.addWidget(bar, 1)
                value = QLabel(str(count))
                value.setMinimumWidth(36)
                value.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                value.setStyleSheet(f"color: {_C_MUTED}; font-size: 12px;")
                row_layout.addWidget(value)
                body_layout.addLayout(row_layout)
        else:
            empty = QLabel("暂无可统计记录")
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty.setStyleSheet(f"color: {_C_MUTED}; padding: 28px;")
            body_layout.addWidget(empty)
        layout.addWidget(body, 1)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(dlg.close)
        layout.addWidget(buttons)
        self._chart_dialog = dlg
        dlg.show()

    def _on_chart_dialog_finished(self) -> None:
        self._chart_dialog = None
        self._show_chart = False
        if hasattr(self, "_btn_chart"):
            self._btn_chart.blockSignals(True)
            self._btn_chart.setChecked(False)
            self._btn_chart.blockSignals(False)

    # ── Column chip handlers ──────────────────────────────────────────────────

    def _on_level_chip(self, level_key: str, show: bool) -> None:
        self._model.set_vis_level(level_key, show)
        self._update_action_delegate_column()

    def _on_lang_chip(self, lang_key: str, show: bool) -> None:
        self._model.set_vis_lang(lang_key, show)
        self._update_action_delegate_column()

    def _update_action_delegate_column(self) -> None:
        """Re-attach delegate to the correct 操作 column after column visibility change."""
        if self._view != "original":
            return
        action_col = _COL_DATA_START + len(self._model.columns()) + 1
        self._table.setItemDelegateForColumn(action_col, self._action_delegate)
        self._table.resizeColumnToContents(action_col)

    def _table_key_press(self, event: "QKeyEvent") -> None:
        """Handle Ctrl+C on the table: copy selected cells in Excel/CSV format."""
        from PyQt6.QtCore import QItemSelection
        from PyQt6.QtGui import QKeyEvent, QKeySequence
        from PyQt6.QtWidgets import QAbstractItemView

        if event.matches(QKeySequence.StandardKey.Copy):
            indexes = self._table.selectionModel().selectedIndexes()
            if not indexes:
                return
            # Sort by row then column
            indexes = sorted(indexes, key=lambda i: (i.row(), i.column()))
            rows: dict[int, list] = {}
            for idx in indexes:
                rows.setdefault(idx.row(), []).append(idx)
            lines = []
            for row_idxs in rows.values():
                parts = []
                for idx in sorted(row_idxs, key=lambda i: i.column()):
                    val = self._model.data(idx, Qt.ItemDataRole.DisplayRole)
                    parts.append(str(val) if val is not None else "")
                lines.append("\t".join(parts))
            from PyQt6.QtWidgets import QApplication
            QApplication.clipboard().setText("\n".join(lines))
            return
        # Fall through to default handling
        QTableView.keyPressEvent(self._table, event)

    def _adjust_font(self, delta: int) -> None:
        """Increase or decrease table font size. Mirrors Excel-style zoom."""
        self._font_size = max(8, min(20, self._font_size + delta))
        self._font_size_lbl.setText(str(self._font_size))
        f = self._table.font()
        f.setPointSize(self._font_size)
        self._table.setFont(f)
        self._table.horizontalHeader().setFont(f)
        # Row height scales with font (approx 2× font pt size)
        row_h = max(22, int(self._font_size * 2.2))
        self._table.verticalHeader().setDefaultSectionSize(row_h)
        self._table.viewport().update()

    # ── Search / filter ───────────────────────────────────────────────────────

    def _on_search(self) -> None:
        self._filter_text = self._search_input.text().strip()
        self._filter_col = self._col_select.currentData() or ""
        self._page = 1
        self._selected_ids.clear()
        self._select_all_filtered = False
        self._model.clear_checked()
        self._load_page()

    def _on_clear_filter(self) -> None:
        self._search_input.clear()
        self._col_select.setCurrentIndex(0)
        self._filter_text = ""
        self._filter_col = ""
        self._page = 1
        self._selected_ids.clear()
        self._select_all_filtered = False
        self._model.clear_checked()
        self._filter_active_label.hide()
        self._load_page()

    # ── Pagination ────────────────────────────────────────────────────────────

    def _on_prev_page(self) -> None:
        if self._page > 1:
            self._page -= 1
            self._load_page()

    def _on_next_page(self) -> None:
        total_pages = max(1, (self._total + _PAGE_SIZE - 1) // _PAGE_SIZE)
        if self._page < total_pages:
            self._page += 1
            self._load_page()

    def _on_jump_page(self) -> None:
        p = self._page_jump.value()
        total_pages = max(1, (self._total + _PAGE_SIZE - 1) // _PAGE_SIZE)
        p = max(1, min(p, total_pages))
        if p != self._page:
            self._page = p
            self._load_page()

    # ── Data loading ──────────────────────────────────────────────────────────

    def _load_page(self) -> None:
        if self._svc is None:
            return

        self._loading = True
        self._loading_label.show()
        self._table.hide()

        # Fetch all records for client-side filtering and sorting
        source_filter = None
        if self._view == "worms":
            source_filter = "seed"

        _fetch_size = max(self._svc.seed_count() + self._svc.user_count(), _PAGE_SIZE)
        all_recs, _ = self._svc.all_records(source_filter=source_filter, page=0, page_size=_fetch_size)

        # Apply text filter client-side
        if self._filter_text:
            q = self._filter_text.lower()
            col_key = self._filter_col
            if col_key:
                all_recs = [r for r in all_recs if q in str(r.get(col_key, "")).lower()]
            else:
                all_recs = [r for r in all_recs if any(q in str(v).lower() for v in r.values())]

        # Apply per-column facet filters (mirrors web taxon filter predicates)
        for col_key, pred in self._col_filters.items():
            if not pred:
                continue
            mode = pred.get("mode", "all")
            if mode == "all":
                continue
            if mode == "include":
                include_vals = set(pred.get("values") or [])
                all_recs = [r for r in all_recs if str(r.get(col_key, "")) in include_vals]
            elif mode == "exclude":
                excluded = set(pred.get("excluded") or [])
                all_recs = [r for r in all_recs if str(r.get(col_key, "")) not in excluded]
            elif mode == "search":
                sq = (pred.get("search") or "").lower()
                excluded = set(pred.get("excluded") or [])
                all_recs = [
                    r for r in all_recs
                    if (not sq or sq in str(r.get(col_key, "")).lower())
                    and str(r.get(col_key, "")) not in excluded
                ]

        # Apply column sort
        if self._sort_col:
            all_recs = sorted(
                all_recs,
                key=lambda r: str(r.get(self._sort_col, "")).lower(),
                reverse=(self._sort_dir == "desc"),
            )

        self._total = len(all_recs)

        # Update filter active label
        has_active_filter = bool(self._filter_text) or bool(self._col_filters)
        if has_active_filter:
            self._filter_active_label.setText(f"已筛选 {self._total} 条")
            self._filter_active_label.show()
        else:
            self._filter_active_label.hide()

        # Client-side pagination
        page_offset = (self._page - 1) * _PAGE_SIZE
        records = all_recs[page_offset : page_offset + _PAGE_SIZE]

        # Back-fill WoRMS mapping status onto the visible rows so review entries
        # surface in the row context menu (mirrors web per-row mappingStatus).
        records = self._annotate_mappings(records)

        self._model.set_records(records, page_offset=page_offset)

        # Re-attach action delegate after model reset (column count may change)
        if self._view == "original":
            action_col = _COL_DATA_START + len(self._model.columns()) + 1
            self._table.setItemDelegateForColumn(action_col, self._action_delegate)

        self._loading = False
        self._loading_label.hide()
        self._table.show()

        self._update_pager()
        self._update_selection_note()

        seed_n = self._svc.seed_count()
        user_n = self._svc.user_count()
        self._stats_label.setText(f"共 {self._total} 条")
        self._footer_label.setText(f"种子库 {seed_n} 条 | 用户 {user_n} 条")

        # Refresh WoRMS job panel (mirrors renderTaxonJobPanel in web)
        self._refresh_job_panel()

    def _update_pager(self) -> None:
        total_pages = max(1, (self._total + _PAGE_SIZE - 1) // _PAGE_SIZE)
        self._page_info.setText(
            f"第 {self._page} / {total_pages} 页（共 {self._total} 条）"
        )
        self._page_jump.setMaximum(total_pages)
        self._page_jump.setValue(self._page)
        self._btn_prev.setEnabled(self._page > 1)
        self._btn_next.setEnabled(self._page < total_pages)

    def _update_selection_note(self) -> None:
        checked_ids = self._model.checked_ids()
        if self._select_all_filtered:
            note = f"已选择全部筛选结果（{self._total} 条）"
        else:
            note = f"已选 {len(checked_ids)} 条"
        self._selection_note.setText(note)
        self._btn_worms_sel.setEnabled(bool(checked_ids) or self._select_all_filtered)

    # ── Selection ─────────────────────────────────────────────────────────────

    def _on_selection_changed(self) -> None:
        idxs = self._table.selectionModel().selectedRows()
        self._selected_ids = []
        for idx in idxs:
            rec = self._model.record_at(idx.row())
            if rec and rec.get("recordId"):
                self._selected_ids.append(rec["recordId"])
        self._select_all_filtered = False
        self._update_selection_note()

    def _on_select_all_filtered(self) -> None:
        self._select_all_filtered = True
        self._model.set_all_page_checked(True)
        self._update_selection_note()

    def _on_deselect(self) -> None:
        self._selected_ids.clear()
        self._select_all_filtered = False
        self._model.clear_checked()
        self._table.clearSelection()
        self._update_selection_note()

    # ── Row context menu (mirrors openTaxonRowMenu / renderTaxonRowMenu) ──────

    def _on_row_context_menu(self, pos: QPoint) -> None:
        index = self._table.indexAt(pos)
        if not index.isValid():
            return
        rec = self._model.record_at(index.row())
        if rec is None:
            return
        menu = QMenu(self._table)
        menu.setStyleSheet(f"QMenu {{ background: {_C_PANEL}; color: {_C_TEXT}; border: 1px solid {_C_BORDER}; border-radius: 6px; }} QMenu::item {{ padding: 6px 18px; font-size: 12px; }} QMenu::item:selected {{ background: {_C_ACCENT_SOFT}; }} QMenu::separator {{ background: {_C_BORDER}; height: 1px; margin: 4px 0; }}")
        title_action = menu.addAction(rec.get("species") or rec.get("class") or "当前记录")
        title_action.setEnabled(False)
        menu.addSeparator()
        wm = menu.addAction("WoRMS 匹配当前物种")
        wm.triggered.connect(lambda: self._on_worms_match_row(rec))
        mapping_candidates = rec.get("mappingCandidates") or []
        if mapping_candidates:
            ra = menu.addAction(f"审核 WoRMS 候选（{len(mapping_candidates)} 个）")
            ra.triggered.connect(lambda: self._on_review_worms_row(rec))
        checked_ids = self._model.checked_ids() or self._selected_ids
        if len(checked_ids) > 1 and rec.get("recordId") in checked_ids:
            menu.addSeparator()
            ba = menu.addAction(f"WoRMS 更新已选 {len(checked_ids)} 条")
            ba.triggered.connect(lambda: self._on_worms_update(selected_only=True))
        if self._select_all_filtered:
            menu.addSeparator()
            fa = menu.addAction(f"WoRMS 更新全部筛选结果 {self._total} 条")
            fa.triggered.connect(lambda: self._on_worms_update(selected_only=False))
        if rec.get("recordId", "").startswith("user:"):
            menu.addSeparator()
            ea = menu.addAction("编辑"); ea.triggered.connect(lambda: self._edit_record(rec))
            da = menu.addAction("删除"); da.triggered.connect(lambda: self._delete_record(rec))
        menu.exec(self._table.viewport().mapToGlobal(pos))

    # ── Column header facet filter (mirrors openTaxonFacetMenu) ──────────────

    def _on_header_context_menu(self, pos: QPoint) -> None:
        self._open_facet_for_column(self._table.horizontalHeader().logicalIndexAt(pos))

    def _open_facet_for_column(self, logical_col: int) -> None:
        if self._svc is None:
            return
        data_idx = logical_col - _COL_DATA_START
        cols = self._model.columns()
        if data_idx < 0 or data_idx >= len(cols):
            return
        col_def = cols[data_idx]
        all_recs, total = self._svc.all_records(page=0, page_size=1_000_000)
        if len(all_recs) < total:
            all_recs, _ = self._svc.all_records(page=0, page_size=max(total, 1))
        if self._facet_panel is not None:
            self._facet_panel.close()
        panel = _TaxonFacetPanel(col_def["key"], col_def["label"], all_recs, current_predicate=self._col_filters.get(col_def["key"]), parent=self)
        panel.filter_applied.connect(self._on_facet_filter_applied)
        panel.sort_requested.connect(self._on_facet_sort)
        header = self._table.horizontalHeader()
        x = header.sectionViewportPosition(logical_col)
        panel.move(self._table.mapToGlobal(QPoint(x, header.height())))
        panel.show(); panel.raise_(); panel.activateWindow()
        self._facet_panel = panel

    def _on_facet_filter_applied(self, col_key: str, predicate: Optional[dict[str, Any]]) -> None:
        if predicate is None: self._col_filters.pop(col_key, None)
        else: self._col_filters[col_key] = predicate
        self._page = 1; self._selected_ids.clear(); self._select_all_filtered = False
        self._model.clear_checked(); self._facet_panel = None; self._load_page()

    def _on_facet_sort(self, col_key: str, direction: str) -> None:
        self._sort_col = col_key; self._sort_dir = direction; self._page = 1; self._load_page()

    # ── WoRMS service helpers ─────────────────────────────────────────────────

    def _worms_data_dir(self) -> Path:
        project_dir = getattr(self.ctx, "current_project_dir", None)
        if project_dir: return Path(project_dir) / "_data"
        return Path.home() / ".photo_workbench" / "data"

    def _ensure_worms_svc(self) -> Optional[WormsService]:
        if self._worms_svc is not None: return self._worms_svc
        data_dir = self._worms_data_dir(); data_dir.mkdir(parents=True, exist_ok=True)
        self._worms_svc = WormsService(cache_path=str(data_dir / "worms_cache.json"), jobs_path=str(data_dir / "worms_jobs.json"))
        return self._worms_svc

    # ── WoRMS match / review / resolve ────────────────────────────────────────

    # ── WoRMS update (mirrors startTaxonomyWormsJob in app.js) ───────────────

    def _on_worms_update(self, selected_only: bool) -> None:
        if self._job_worker is not None and self._job_worker.isRunning():
            QMessageBox.information(self, "WoRMS 更新", "已有 WoRMS 任务在运行，请等待完成或在进度区暂停。")
            return
        record_ids = self._worms_update_record_ids(selected_only)
        if not record_ids:
            QMessageBox.information(self, "WoRMS 更新", "没有可更新的分类条目。")
            return
        source = "selected" if selected_only and not self._select_all_filtered else "filtered"
        scope = f"已选 {len(record_ids)} 条" if source == "selected" else f"筛选结果 {len(record_ids)} 条"
        confirm = QMessageBox.question(
            self, "WoRMS 更新",
            f"即将对{scope}发起 WoRMS 校验更新。\n"
            "将逐条访问 WoRMS（约每条 0.6 秒），结果记录为校验状态供审核，"
            "不会改写原始条目；可在进度条处随时暂停。\n\n是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return
        service = self._ensure_worms_svc()
        if service is None:
            QMessageBox.warning(self, "WoRMS 更新", "WoRMS 服务不可用。")
            return
        try:
            job = service.create_job(record_ids, source=source)
        except Exception as exc:
            QMessageBox.warning(self, "WoRMS 更新", f"创建任务失败：{exc}")
            return
        setattr(self.ctx, "pending_worms_job_id", job.id)
        self._start_job_worker(job.id)

    def _build_record_resolver(self):
        """Return a recordId→record lookup over the FULL library.

        Selections may span pages, so the resolver must see every record, not
        just the visible page.  Records without a recordId (seed rows) are not
        checkbox-selectable and are simply absent → worker treats them as stale.
        """
        if self._svc is None:
            return lambda rid: None
        total = self._svc.seed_count() + self._svc.user_count()
        rows, _ = self._svc.all_records(page=0, page_size=max(total, 1))
        index = {r.get("recordId", ""): r for r in rows if r.get("recordId")}
        return lambda rid: index.get(rid)

    def _start_job_worker(self, job_id: str) -> None:
        service = self._ensure_worms_svc()
        if service is None:
            return
        worker = _WormsJobWorker(service, job_id, self._build_record_resolver(), parent=self)
        worker.progress.connect(self._on_job_progress)
        worker.finished_job.connect(self._on_job_finished)
        worker.failed.connect(self._on_job_failed)
        self._job_worker = worker
        self._refresh_job_panel()   # show panel immediately at 0/N
        worker.start()

    def _on_job_progress(self, cursor: int, total: int, counts: dict) -> None:
        self._refresh_job_panel()

    def _on_job_finished(self, job: dict) -> None:
        self._refresh_job_panel()
        # Reload so per-row mapping status surfaces (审核 entry appears on review rows).
        self._load_page()
        if job.get("status") == "completed":
            counts = job.get("counts") or {}
            _CL = {"matched": "匹配", "renamed": "改名", "review": "待审",
                   "not_found": "未找到", "error": "错误", "stale": "跳过"}
            summary = " · ".join(f"{_CL.get(k, k)} {v}" for k, v in counts.items() if v) or "无变化"
            QMessageBox.information(self, "WoRMS 更新", f"WoRMS 校验完成：{summary}")

    def _on_job_failed(self, msg: str) -> None:
        self._refresh_job_panel()
        QMessageBox.warning(self, "WoRMS 更新", f"任务出错：{msg}")

    def _on_job_pause(self) -> None:
        service = self._ensure_worms_svc()
        if service is None:
            return
        active = next((j for j in service.list_jobs() if j.get("status") == "running"), None)
        if active:
            service.update_job_status(active["id"], "paused")   # worker exits at next tick
        self._refresh_job_panel()

    def _on_job_resume(self) -> None:
        service = self._ensure_worms_svc()
        if service is None:
            return
        paused = next((j for j in service.list_jobs() if j.get("status") == "paused"), None)
        if not paused:
            return
        service.update_job_status(paused["id"], "running")
        self._start_job_worker(paused["id"])

    def _on_job_retry(self) -> None:
        service = self._ensure_worms_svc()
        if service is None:
            return
        target = next((j for j in service.list_jobs() if (j.get("counts") or {}).get("error")), None)
        if not target:
            return
        retried = service.retry_failed_job(target["id"])
        if retried:
            self._start_job_worker(retried["id"])

    def _annotate_mappings(self, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        worms_svc = self._ensure_worms_svc()
        if worms_svc is None:
            return records
        try:
            mappings = worms_svc.list_mappings()
        except Exception:
            return records
        if not mappings:
            return records
        out: list[dict[str, Any]] = []
        for rec in records:
            m = mappings.get(rec.get("recordId", ""))
            if m:
                rec = dict(rec)
                rec["mappingStatus"] = m.get("status", "")
                rec["mappingCandidates"] = m.get("candidates", [])
            out.append(rec)
        return out

    def _worms_update_record_ids(self, selected_only: bool) -> list[str]:
        if self._svc is None: return []
        if selected_only and not self._select_all_filtered:
            return list(dict.fromkeys(rid for rid in self._model.checked_ids() if rid))
        source_filter = "seed" if self._view == "worms" else None
        rows, total = self._svc.all_records(source_filter=source_filter, page=0, page_size=1_000_000)
        if len(rows) < total: rows, _ = self._svc.all_records(source_filter=source_filter, page=0, page_size=max(total, 1))
        ids: list[str] = []
        for idx, rec in enumerate(rows):
            if self._filter_text and not self._record_matches_filter(rec): continue
            ids.append(self._taxonomy_record_id(rec, idx, source_filter))
        return ids

    def _taxonomy_record_id(self, rec: dict[str, Any], index: int, source_filter: Optional[str] = None) -> str:
        rid = str(rec.get("recordId") or "").strip()
        if rid: return rid
        return f"{source_filter or ('user' if str(rec.get('recordId', '')).startswith('user:') else 'seed')}:{index}"

    def _navigate_to_worms(self) -> None:
        win = self.window()
        nav = getattr(win, "navigate_to", None)
        if callable(nav): nav("worms")
        wv = getattr(win, "_views", {}).get("worms") if hasattr(win, "_views") else None
        ref = getattr(wv, "_refresh_jobs", None)
        if callable(ref): ref()

    def _on_worms_match_row(self, rec: dict[str, Any]) -> None:
        worms_svc = self._ensure_worms_svc()
        if worms_svc is None: QMessageBox.warning(self, "WoRMS", "WoRMS 服务不可用"); return
        dlg = _WormsMatchDialog(rec, worms_svc, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            result = dlg.get_result()
            if result: self._on_resolve_mapping(rec, result)

    def _on_review_worms_row(self, rec: dict[str, Any]) -> None:
        dlg = _TaxonReviewDialog(rec, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            result = dlg.get_result()
            if result: self._on_resolve_mapping(rec, result)

    def _on_resolve_mapping(self, rec: dict[str, Any], result: dict[str, Any]) -> None:
        """Apply WoRMS resolution decision (mirrors resolveTaxonMapping in app.js)."""
        worms_svc = self._ensure_worms_svc()
        if worms_svc is None: return
        record_id = rec.get("recordId", "")
        if not record_id: return
        try:
            if result.get("no_match"):
                worms_svc.resolve_mapping(record_id, None, no_match=True)
            else:
                aphia_id = result.get("aphia_id")
                worms_svc.resolve_mapping(record_id, int(aphia_id) if aphia_id else None, worms_record=result.get("worms_record") or {}, chain=result.get("chain") or [])
            QMessageBox.information(self, "WoRMS", "审核结果已保存")
        except Exception as exc:
            QMessageBox.warning(self, "WoRMS 错误", f"审核失败：{exc}")
        self._load_page()

    # ── Job panel (mirrors renderTaxonJobPanel in app.js) ─────────────────────

    def _refresh_job_panel(self) -> None:
        worms_svc = self._ensure_worms_svc()
        if worms_svc is None: self._job_panel_frame.hide(); return
        jobs = worms_svc.list_jobs()
        if not jobs: self._job_panel_frame.hide(); return
        active = next((j for j in jobs if j.get("status") in ("running", "paused")), None)
        job = active or jobs[0]
        record_ids = job.get("record_ids") or []; total = len(record_ids); cursor = job.get("cursor", 0); status = job.get("status", ""); source = job.get("source", "")
        _SL = {"running": "运行中", "paused": "已暂停", "completed": "已完成", "cancelled": "已取消"}
        self._job_title_label.setText(f"WoRMS 任务 · {'选中条目' if source == 'selected' else '筛选结果'}")
        self._job_progress_label.setText(f"{cursor} / {total} · {_SL.get(status, status)}")
        self._job_bar.setRange(0, max(total, 1)); self._job_bar.setValue(cursor)
        counts = job.get("counts") or {}
        _CL = {"matched": "匹配", "renamed": "改名", "review": "待审", "not_found": "未找到", "error": "错误", "stale": "旧缓存"}
        self._job_counts_label.setText("  ".join(f"{_CL.get(k, k)} {v}" for k, v in counts.items() if v))
        self._btn_job_pause.setVisible(status == "running"); self._btn_job_resume.setVisible(status == "paused")
        self._btn_job_retry.setVisible(bool(counts.get("error"))); self._job_panel_frame.show()

    # ── Double-click edit ─────────────────────────────────────────────────────

    def _on_row_double_click(self, index: QModelIndex) -> None:
        if self._view != "original":
            return
        rec = self._model.record_at(index.row())
        if rec is None:
            return
        if not rec.get("recordId", "").startswith("user:"):
            QMessageBox.information(
                self, "只读",
                "种子库条目不可编辑（双击用户条目可编辑）"
            )
            return
        self._edit_record(rec)

    # ── CRUD ──────────────────────────────────────────────────────────────────

    def _on_add(self) -> None:
        if self._svc is None:
            return
        dlg = _RecordDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        self._svc.learn(dlg.get_record())
        self._page = 1
        self._load_page()

    def _edit_record(self, rec: dict[str, Any]) -> None:
        """Edit record. Mirrors web openTaxonomyTableModal('edit'):
        - user: records → update in place
        - seed: records → open dialog pre-filled with seed data, save creates a user override
        """
        if self._svc is None:
            return
        is_user = rec.get("recordId", "").startswith("user:")
        dlg = _RecordDialog(self, record=rec)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        if is_user:
            self._svc.update(rec["recordId"], dlg.get_record())
        else:
            # Seed record: create user override entry (mirrors web findUserEntryForCurrent + learn)
            self._svc.learn(dlg.get_record())
        self._load_page()

    def _delete_record(self, rec: dict[str, Any]) -> None:
        if self._svc is None:
            return
        if not rec.get("recordId", "").startswith("user:"):
            QMessageBox.warning(self, "只读", "种子库条目不可删除")
            return
        name = f"{rec.get('species', '')} ({rec.get('class', '')})"
        reply = QMessageBox.question(
            self, "确认删除", f"删除用户条目「{name}」？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self._svc.delete(rec["recordId"])
        self._load_page()

    # ── Export ────────────────────────────────────────────────────────────────

    def _on_export(self, fmt: str) -> None:
        if self._svc is None:
            return
        source_filter = None
        if self._view == "worms":
            source_filter = "seed"
        all_recs, _ = self._svc.all_records(
            source_filter=source_filter, page=0, page_size=999999
        )
        if fmt == "csv":
            self._export_csv(all_recs)
        else:
            self._export_xlsx(all_recs)

    def _export_csv(self, records: list[dict[str, Any]]) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 CSV", "taxonomy_export.csv", "CSV 文件 (*.csv)"
        )
        if not path:
            return
        import csv
        cols = self._model.columns()
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            header = [c["label"] for c in cols] + ["来源"]
            writer.writerow(header)
            for rec in records:
                row = [rec.get(c["key"], "") for c in cols]
                row.append("用户" if rec.get("recordId", "").startswith("user:") else "种子")
                writer.writerow(row)
        QMessageBox.information(self, "导出完成", f"已导出 {len(records)} 条到\n{path}")

    def _export_xlsx(self, records: list[dict[str, Any]]) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", "taxonomy_export.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "分类库"
            cols = self._model.columns()
            ws.append([c["label"] for c in cols] + ["来源"])
            for rec in records:
                row = [rec.get(c["key"], "") for c in cols]
                row.append("用户" if rec.get("recordId", "").startswith("user:") else "种子")
                ws.append(row)
            wb.save(path)
            QMessageBox.information(self, "导出完成", f"已导出 {len(records)} 条到\n{path}")
        except ImportError:
            QMessageBox.critical(self, "缺少依赖", "需要 openpyxl 库：pip install openpyxl")
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))

    # ── Import ────────────────────────────────────────────────────────────────

    def get_filtered_uids(self) -> list:
        """Return record identifiers of all records currently shown in the table (after filter applied).

        User records return their ``recordId``.  Seed records without one return
        ``seed:{page_offset+row}`` — the same convention used by _taxonomy_record_id.
        """
        uids = []
        for row in range(self._model.rowCount()):
            rec = self._model.record_at(row)
            if rec is None:
                continue
            rid = str(rec.get("recordId", "")).strip()
            if rid:
                uids.append(rid)
            else:
                page_offset = self._model._page_offset
                uids.append(f"seed:{page_offset + row}")
        return uids

    def _on_import(self) -> None:
        """Import taxonomy records from Excel or CSV.

        Mirrors the importInput handler in app.js (POST /api/taxonomy/import).
        Expected columns (case-insensitive, any order):
          class, order, family, species (required)
          classCn, orderCn, familyCn, speciesCn, genus, genusCn (optional)
        """
        if self._svc is None:
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel / CSV 文件",
            "",
            "表格文件 (*.xlsx *.xls *.csv)",
        )
        if not path:
            return
        try:
            if path.lower().endswith(".csv"):
                self._import_csv(path)
            else:
                self._import_xlsx(path)
        except Exception as exc:
            QMessageBox.critical(
                self, "导入失败",
                f"{type(exc).__name__}: {exc}\n\n{traceback.format_exc()}"
            )

    def _import_xlsx(self, path: str) -> None:
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "缺少依赖", "需要 openpyxl 库：pip install openpyxl")
            return
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(self, "空文件", "Excel 文件中没有数据")
            return
        imported, skipped = self._import_rows(rows[0], rows[1:])
        self._load_page()
        QMessageBox.information(
            self, "导入完成", f"成功导入 {imported} 条，跳过 {skipped} 条。"
        )

    def _import_csv(self, path: str) -> None:
        import csv
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            QMessageBox.warning(self, "空文件", "CSV 文件中没有数据")
            return
        imported, skipped = self._import_rows(rows[0], rows[1:])
        self._load_page()
        QMessageBox.information(
            self, "导入完成", f"成功导入 {imported} 条，跳过 {skipped} 条。"
        )

    def _import_rows(
        self,
        header_row: Any,
        data_rows: list[Any],
    ) -> tuple[int, int]:
        _alias: dict[str, str] = {
            "class":     "class",   "纲":    "class",
            "order":     "order",   "目":    "order",
            "family":    "family",  "科":    "family",
            "species":   "species", "种":    "species",
            "classcn":   "classCn",  "纲中文": "classCn",
            "ordercn":   "orderCn",  "目中文": "orderCn",
            "familycn":  "familyCn", "科中文": "familyCn",
            "speciescn": "speciesCn","种中文": "speciesCn",
            "genus":     "genus",    "属":    "genus",
            "genuscn":   "genusCn",  "属中文": "genusCn",
        }

        col_map: dict[str, int] = {}
        for i, h in enumerate(header_row):
            if h:
                col_map[str(h).strip().lower()] = i

        field_idx: dict[str, int] = {}
        for raw, canon in _alias.items():
            if raw in col_map:
                field_idx[canon] = col_map[raw]

        def _cell(row: Any, field: str) -> str:
            i = field_idx.get(field)
            if i is None:
                return ""
            try:
                v = row[i]
            except IndexError:
                return ""
            return str(v).strip() if v is not None else ""

        imported = skipped = 0
        for row in data_rows:
            rec = {f: _cell(row, f) for f in (
                "class", "order", "family", "species",
                "classCn", "orderCn", "familyCn", "speciesCn",
                "genus", "genusCn",
            )}
            result = self._svc.learn(rec)  # type: ignore[union-attr]
            if result:
                imported += 1
            else:
                skipped += 1
        return imported, skipped
