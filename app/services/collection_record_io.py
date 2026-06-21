"""collection_record_io.py — 采集记录 Excel/CSV 模板导出 + 导入（步骤 5）.

用户工作流：导出一个 Excel 模板（含已有记录，便于离线/在 Excel 里批量填）→
离线填写 → 导回软件。复用 openpyxl（requirements 已含）。

**按采区分两套国标表样**（GB/T 12763.6-2007）：
  - 潮间带 H.39（样方号 / 气温 / 标本瓶数 …）
  - 潮下带 H.30（航次 / 船号 / 放绳长度 / 网型 / 拖网距离 …）
列序来自 ``collection_record_service.ZONE_COLUMNS``（单一真相源）。

*zone* 取值：
  - ``'intertidal'`` / ``'subtidal'`` → 单 sheet，只含该区记录 + 该区列序；
  - ``'all'`` → 两 sheet（潮间带 / 潮下带）；历史未分类（zone 为空）记录归入潮间带 sheet。

导入按表头自动判区（``crs.infer_zone_from_headers``），逐 sheet 处理多 sheet 工作簿。
写入走 ``collection_record_service.upsert_record``（带 zone）。

纯逻辑、无 Qt，便于测试。
"""

from __future__ import annotations

import csv
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path

from app.services import collection_record_service as crs

# 必填四键——导入时缺任一则跳过该行。
_KEY_FIELDS: tuple[str, ...] = ("province", "site", "station", "collection_date")

# sheet 名 ↔ zone
_ZONE_TITLES: dict[str, str] = {"intertidal": "潮间带", "subtidal": "潮下带"}

# 旧表头别名 → 字段 key：兼容字段优化前导出的模板。
# （UI 表头已改 底质/气象/表层水温，但 DB key 不变；旧「生境/天气/水温」仍要能读。）
_LEGACY_HEADER_ALIASES: dict[str, str] = {
    "水温": "water_temp",
    "采集方法": "method",
    "生境": "habitat",        # 旧表头「生境」→ habitat（现显示「底质」）
    "天气": "weather",        # 旧表头「天气」→ weather（现显示「气象」）
}


def _alias_map() -> dict[str, str]:
    """Build {header → field key}: accept 中文表头 / 英文 key（大小写不敏感）."""
    amap: dict[str, str] = dict(_LEGACY_HEADER_ALIASES)
    for zh, key in crs._HEADER_TO_KEY.items():
        amap[zh] = key
    for key in crs._COLUMNS:
        amap[key] = key
        amap[key.lower()] = key
    return amap


@dataclass
class ImportReport:
    ok: bool = True
    imported: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def _record_in_zone(rec: dict, zone: str) -> bool:
    """记录是否归入 *zone*。subtidal 严格匹配；intertidal 含历史未分类（空）。"""
    rz = (rec.get("zone") or "").strip()
    if zone == "subtidal":
        return rz == "subtidal"
    return rz in ("", "intertidal")   # intertidal 含 NULL（用户主用）


# ── Export ──────────────────────────────────────────────────────────────────
def export_template(
    db: sqlite3.Connection,
    path: str,
    *,
    zone: str = "all",
    province: str = "",
    site: str = "",
    blank_rows: int = 20,
) -> int:
    """Write an .xlsx template at *path*; return the written-record count.

    *zone* ∈ {'intertidal','subtidal','all'}：单区单 sheet / 全部两 sheet。
    含该区已有记录（便于离线批量编辑）+ *blank_rows* 空行（预填继承的地区/样地）。
    历史未分类（zone 空）记录归入潮间带 sheet。
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    if zone not in ("intertidal", "subtidal", "all"):
        raise ValueError(f"zone 非法：{zone!r}，应为 intertidal/subtidal/all")

    records = crs.list_records(db) if db is not None else []
    zones = ("intertidal", "subtidal") if zone == "all" else (zone,)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 去掉默认 sheet，下面按 zone 建

    hdr_fill = PatternFill("solid", fgColor="2C5F8A")
    hdr_font = Font(color="FFFFFF", bold=True)

    total = 0
    for z in zones:
        cols = crs.columns_for_zone(z)
        if not cols:
            continue
        ws = wb.create_sheet(_ZONE_TITLES[z])
        headers = [zh for _k, zh in cols]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font

        zrecs = [r for r in records if _record_in_zone(r, z)]
        for rec in zrecs:
            ws.append([_cell(rec.get(k)) for k, _zh in cols])
        # 空行预填 地区/样地（离线录入免重敲）
        for _ in range(max(0, blank_rows)):
            ws.append([
                province if k == "province" else (site if k == "site" else "")
                for k, _zh in cols
            ])
        total += len(zrecs)

    if not wb.worksheets:   # 兜底：zone 非法导致空工作簿时留一张
        ws = wb.create_sheet("采集记录")

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return total


def _cell(val) -> str:
    return "" if val in (None,) else str(val)


# ── Import ──────────────────────────────────────────────────────────────────
def import_file(db: sqlite3.Connection, path: str) -> ImportReport:
    """Import .xlsx or .csv at *path* into collection_records (upsert per row).

    多 sheet xlsx 逐 sheet 处理；每 sheet 按表头自动判区（潮间带/潮下带），
    fallback 潮间带。CSV 单表同理。
    """
    if db is None:
        return ImportReport(ok=False, errors=["没有打开的项目"])
    p = Path(path)
    try:
        if p.suffix.lower() in (".xlsx", ".xlsm"):
            sheets = _read_xlsx_sheets(path)
        else:
            sheets = [_read_csv(path)]
    except Exception as exc:  # noqa: BLE001
        return ImportReport(ok=False, errors=[f"读取失败：{exc}"])

    rep = ImportReport()
    for _name, header, rows in sheets:
        sub = _import_rows(db, header, rows)
        rep.imported += sub.imported
        rep.skipped += sub.skipped
        rep.errors.extend(sub.errors)
        if not sub.ok and sub.errors and not rep.imported and not rep.skipped:
            # 表头全无法识别等致命错误，向上透传 ok=False（仅在尚无任何成功时）
            rep.ok = False
    return rep


def _read_xlsx(path: str) -> tuple[list[str], list[list]]:
    """读首个 sheet（向后兼容旧测试/单 sheet 模板）。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [("" if c is None else str(c)).strip() for c in next(rows_iter)]
    except StopIteration:
        return [], []
    data = [list(r) for r in rows_iter]
    return header, data


def _read_xlsx_sheets(path: str) -> list[tuple[str, list[str], list[list]]]:
    """读全部 sheet：返回 [(sheet名, header, rows), ...]。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[str], list[list]]] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [("" if c is None else str(c)).strip() for c in next(rows_iter)]
        except StopIteration:
            out.append((ws.title, [], []))
            continue
        data = [list(r) for r in rows_iter]
        out.append((ws.title, header, data))
    return out


def _read_csv(path: str) -> tuple[str, list[str], list[list]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return ("CSV", [], [])
    header = [c.strip() for c in rows[0]]
    return ("CSV", header, rows[1:])


def _import_rows(db, header: list[str], rows: list[list]) -> ImportReport:
    amap = _alias_map()
    col_to_key: dict[int, str] = {}
    for i, h in enumerate(header):
        key = amap.get(h) or amap.get(h.lower())
        if key:
            col_to_key[i] = key

    rep = ImportReport()
    if not col_to_key:
        rep.ok = False
        rep.errors.append("表头无法识别（需含 地区/样地/站位/采集日期 等列）")
        return rep

    # 本 sheet 的采区：按表头判，fallback 潮间带
    zone = crs.infer_zone_from_headers(header) or "intertidal"

    for r_idx, raw in enumerate(rows, start=2):
        data: dict = {"zone": zone}
        for i, key in col_to_key.items():
            val = raw[i] if i < len(raw) else None
            data[key] = ("" if val is None else str(val)).strip()
        # Skip wholly-empty rows silently.
        if not any(data.get(k) for k in col_to_key.values()):
            continue
        # Need the 4 key fields to identify a record.
        if not all(data.get(k) for k in _KEY_FIELDS):
            rep.skipped += 1
            continue
        try:
            crs.upsert_record(db, data)
            rep.imported += 1
        except Exception as exc:  # noqa: BLE001
            rep.errors.append(f"第 {r_idx} 行：{exc}")
    return rep
