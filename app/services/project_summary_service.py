"""project_summary_service.py — 跨工作区汇总导出（只读，无新建表）.

给定一组工作区目录（每个含 ``<dir>/_data/project.db``）和一个 root 目录（仅用于
计算 断面/来源 标签），合并它们的数据并写出 Excel / HTML 报告。

ZERO 新表：只读已有表（specimens / collection_records / grouping）。复用：
  - ``db_manager.open_project_db_private(dir, create=False)`` 严格打开（缺库即
    跳过），用完即 close —— 跨工作区读规约：不缓存子库连接、不跑 ensure_schema
  - ``Specimen.from_row`` 构造标本
  - ``export_service.export_excel`` 的 ``extra_leading`` 前置列 + 其样式辅助
  - ``collection_record_service.list_records`` 取采集记录
  - ``project_tree_service.scan_tree`` 做「文件夹存在但非工作区」质控

纯逻辑、无 Qt。一个缺失/锁定的工作区被静默跳过，绝不中断整个导出。
"""
from __future__ import annotations

import json
import os
import sqlite3
from collections import Counter
from pathlib import Path, PureWindowsPath
from typing import Iterator, Optional

import openpyxl

# from app.db.db_manager import open_project_db  # §7 旧: 缓存连接跨工作区 → 锁泄漏
from app.db.db_manager import open_project_db_private
from app.models.specimen import Specimen
from app.services.collection_record_service import list_records
from app.services.export_service import (
    _ALT_FILL,
    _apply_header_row,
    _auto_col_widths,
    export_excel,
)
from app.services.project_paths import ProjectUnavailableError
from app.services.project_tree_service import scan_tree
from app.utils.naming import parse_uid

# 采集站位汇总的列（断面在前，由调用处单独前置）。与 schema.sql 的真实列对齐。
_COLLECTION_COLUMNS: tuple[str, ...] = (
    "province", "site", "station", "collection_date",
    "station_label", "lon", "lat", "geo_area", "water_body",
    "cruise", "vessel",
    "tidal_zone", "depth",
    "habitat", "tide", "salinity", "water_temp", "bottom_temp",
    "dissolved_oxygen", "ph", "weather",
    "sample_type", "method", "sampler_model", "sampler_spec", "sample_area",
    "replicates", "sieve_mesh", "sample_no",
    "collector", "recorder", "checker", "photographer", "identifier",
    "collection_time", "photo_date", "photo_location",
    "remark",
)

_GROUPING_COMPOSED_STATUSES: tuple[str, ...] = ("composed", "organized")
_GROUPING_KEY_SEP = "\0"

# Table-data cache for the interactive summary page.  Export functions remain
# uncached because they write files and are normally one-shot operations.
_SPECIMEN_ROWS_CACHE: dict[tuple, dict] = {}


# ── helpers ────────────────────────────────────────────────────────────────────

def _label(ws_dir: str, root: str) -> str:
    """工作区的 断面/来源 标签 = ``relpath(ws_dir, root)``。

    若 relpath 为 "." → 用 ``basename(ws_dir)``（root 本身就是工作区）。
    若 ws_dir 不在 root 之下（relpath 形如 "../.."，模式 B：无关的近期项目）→
    退回 ``basename(ws_dir)``。
    """
    try:
        rel = os.path.relpath(ws_dir, root)
    except ValueError:
        # 不同盘符（Windows）等无法计算相对路径的情形
        return os.path.basename(os.path.normpath(ws_dir))
    if rel == "." or rel.startswith(".."):
        return os.path.basename(os.path.normpath(ws_dir))
    return rel


def _iter_dbs(dirs: list[str]) -> Iterator[tuple[str, sqlite3.Connection]]:
    """逐个 yield ``(ws_dir, conn)``；缺失/锁定的库静默跳过，绝不中断全流程。

    私有连接 + 用完即 close（CLAUDE.md 跨工作区读规约）：缓存连接会把每个
    子工作区的文件锁扣到退出（Windows 目录删不掉/移不动），且
    ``open_project_db`` 每次 open 都跑 ensure_schema —— 「只读汇总」会静默
    迁移/写入子库。plain 私有连接跳过 ensure_schema，真正只读；老库缺列由
    ``Specimen.from_row`` 的 ``d.get(k)`` 容错，缺表由各调用点的
    ``except sqlite3.Error`` 跳过。
    """
    for d in dirs:
        try:
            # conn = open_project_db(d, create=False)  # §7 旧: 缓存连接 + 隐式 ensure_schema
            conn = open_project_db_private(d, create=False)
        except (ProjectUnavailableError, sqlite3.Error):
            continue
        try:
            yield d, conn
        finally:
            conn.close()


def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    try:
        return {str(row["name"]) for row in conn.execute(f"PRAGMA table_info({table})")}
    except sqlite3.Error:
        return set()


def _json_path_list(raw) -> list[str]:
    if raw in (None, ""):
        return []
    try:
        data = json.loads(raw) if isinstance(raw, str) else raw
    except (TypeError, ValueError):
        return []
    if not isinstance(data, list):
        return []
    return [str(item).strip() for item in data if str(item or "").strip()]


def _path_basename(path: str) -> str:
    value = str(path or "").strip()
    if not value:
        return ""
    if "\\" in value:
        return PureWindowsPath(value).name or value
    return Path(value).name or value


def _joined_basenames(paths: list[str]) -> str:
    names = [_path_basename(path) for path in paths if str(path or "").strip()]
    return "；".join(name for name in names if name)


def _empty_grouping_summary(project_code: str = "") -> dict:
    return {
        "count": 0,
        "group_count": 0,
        "status": "无成果",
        "projCode": project_code,
        "jpg_count": 0,
        "tiff_count": 0,
        "zip_count": 0,
        "tiff_names": "",
        "tiff_paths": "",
        "zip_names": "",
        "zip_paths": "",
    }


def specimen_grouping_key(workspace_dir: str, uid: str) -> str:
    """Return a stable grouping key for one specimen row.

    ``uid`` is only unique inside one workspace.  Cross-workspace summary tables
    therefore key grouping facts by ``workspace_dir + uid`` to avoid silent
    collisions when two projects contain the same specimen number.
    """
    return f"{workspace_dir or ''}{_GROUPING_KEY_SEP}{uid or ''}"


def _cache_signature(dirs: list[str], root: str) -> tuple:
    parts: list[tuple] = []
    for d in dirs:
        ws = str(Path(d).resolve())
        db_path = Path(ws) / "_data" / "project.db"
        wal_path = Path(str(db_path) + "-wal")

        def _file_stat_signature(path: Path):
            try:
                s = path.stat()
                return (s.st_mtime_ns, s.st_size)
            except OSError:
                return None

        parts.append((ws, _file_stat_signature(db_path), _file_stat_signature(wal_path)))
    return (str(Path(root).resolve()) if root else "", tuple(parts))


def clear_specimen_summary_cache() -> None:
    """Clear cached interactive summary rows.

    Tests and maintenance commands can call this after deliberately mutating
    many project DBs in-place.
    """
    _SPECIMEN_ROWS_CACHE.clear()


def _copy_summary_payload(payload: dict) -> dict:
    return {
        "specimens": list(payload.get("specimens", [])),
        "grouping": {k: dict(v) for k, v in payload.get("grouping", {}).items()},
        "projects": [dict(p) for p in payload.get("projects", [])],
        "dashboard": dict(payload.get("dashboard") or {}),
    }


def _project_meta(conn: sqlite3.Connection) -> dict:
    try:
        row = conn.execute(
            "SELECT value_json FROM project_settings WHERE setting_key='project_meta'"
        ).fetchone()
    except sqlite3.Error:
        return {}
    if not row:
        return {}
    try:
        data = json.loads(row[0])
    except (ValueError, TypeError):
        return {}
    return data if isinstance(data, dict) else {}


def _project_descriptor(conn: sqlite3.Connection, ws_dir: str, root: str) -> dict:
    label = _label(ws_dir, root)
    name = label
    code = ""

    meta = _project_meta(conn)
    if meta:
        name = str(meta.get("name") or name)
        code = str(meta.get("project_code") or meta.get("code") or code)

    # Older converted databases may carry a projects table.  Treat it as a
    # best-effort source and keep the newer project_settings value authoritative.
    try:
        rows = conn.execute("SELECT directory, name, project_code FROM projects").fetchall()
    except sqlite3.Error:
        rows = []
    for row in rows:
        d = dict(row)
        directory = d.get("directory") or ""
        if directory and os.path.normpath(directory) != os.path.normpath(ws_dir):
            continue
        name = str(d.get("name") or name)
        code = code or str(d.get("project_code") or "")
        break

    return {
        "directory": ws_dir,
        "name": name or Path(ws_dir).name,
        "project_code": code,
        "label": label,
    }


def _current_photo_assignment_counts(conn: sqlite3.Connection) -> dict[str, int]:
    photo_cols = _table_columns(conn, "photos")
    assignment_cols = _table_columns(conn, "photo_assignments")
    if not {"current_assignment_id"} <= photo_cols:
        return {}
    if not {"assignment_id", "specimen_uid"} <= assignment_cols:
        return {}

    current_expr = "AND COALESCE(a.is_current, 1)=1" if "is_current" in assignment_cols else ""
    try:
        rows = conn.execute(
            f"""
            SELECT a.specimen_uid AS uid, COUNT(*) AS n
              FROM photos p
              JOIN photo_assignments a ON a.assignment_id = p.current_assignment_id
             WHERE a.specimen_uid IS NOT NULL
               AND TRIM(a.specimen_uid) <> ''
               {current_expr}
             GROUP BY a.specimen_uid
            """
        ).fetchall()
    except sqlite3.Error:
        return {}
    return {str(row["uid"]): int(row["n"] or 0) for row in rows}


def collect_grouping_summary_for_connection(
    conn: sqlite3.Connection,
    project_code: str = "",
) -> dict[str, dict]:
    """Return per-specimen grouping facts for the interactive summary table."""
    out: dict[str, dict] = {}
    cols = _table_columns(conn, "grouping")
    if "uid" not in cols:
        assigned_counts = _current_photo_assignment_counts(conn)
        return {
            uid: {**_empty_grouping_summary(project_code), "jpg_count": count}
            for uid, count in assigned_counts.items()
        }

    wanted = ["uid", "status", "jpg_paths", "composed_tiff_path", "archive_zip"]
    select = ", ".join(
        name if name in cols else f"NULL AS {name}"
        for name in wanted
    )
    try:
        rows = conn.execute(f"SELECT {select} FROM grouping").fetchall()
    except sqlite3.Error:
        return out

    scratch: dict[str, dict] = {}
    for row in rows:
        d = dict(row)
        uid = str(d.get("uid") or "").strip()
        if not uid:
            continue
        bucket = scratch.setdefault(uid, {
            "total": 0,
            "done": 0,
            "jpg_paths": set(),
            "tiff_paths": [],
            "zip_paths": [],
        })
        bucket["total"] += 1
        if str(d.get("status") or "") in _GROUPING_COMPOSED_STATUSES:
            bucket["done"] += 1
        for path in _json_path_list(d.get("jpg_paths")):
            bucket["jpg_paths"].add(path)
        tiff_path = str(d.get("composed_tiff_path") or "").strip()
        if tiff_path and tiff_path not in bucket["tiff_paths"]:
            bucket["tiff_paths"].append(tiff_path)
        zip_path = str(d.get("archive_zip") or "").strip()
        if zip_path and zip_path not in bucket["zip_paths"]:
            bucket["zip_paths"].append(zip_path)

    assigned_counts = _current_photo_assignment_counts(conn)
    for uid, assigned_count in assigned_counts.items():
        scratch.setdefault(uid, {
            "total": 0,
            "done": 0,
            "jpg_paths": set(),
            "tiff_paths": [],
            "zip_paths": [],
        })["assigned_count"] = assigned_count

    for uid, bucket in scratch.items():
        total = int(bucket.get("total") or 0)
        done = int(bucket.get("done") or 0)
        if done == 0:
            status = "待合成" if total else "无成果"
        elif done == total:
            status = "已合成"
        else:
            status = "部分合成"
        jpg_count = max(
            len(bucket.get("jpg_paths") or []),
            int(bucket.get("assigned_count") or 0),
        )
        tiff_paths = list(bucket.get("tiff_paths") or [])
        zip_paths = list(bucket.get("zip_paths") or [])
        out[uid] = {
            "count": done,
            "group_count": total,
            "status": status,
            "projCode": project_code,
            "jpg_count": jpg_count,
            "tiff_count": len(tiff_paths),
            "zip_count": len(zip_paths),
            "tiff_names": _joined_basenames(tiff_paths),
            "tiff_paths": "；".join(tiff_paths),
            "zip_names": _joined_basenames(zip_paths),
            "zip_paths": "；".join(zip_paths),
        }
    return out


def _grouping_summary_for_db(conn: sqlite3.Connection, project_code: str) -> dict[str, dict]:
    return collect_grouping_summary_for_connection(conn, project_code)


def collect_specimen_summary_rows(
    dirs: list[str],
    root: str,
    *,
    use_cache: bool = True,
) -> dict:
    """Return interactive summary rows across workspaces.

    Result shape:
      ``{"specimens": list[Specimen], "grouping": dict, "projects": list[dict],
      "dashboard": dict}``

    Missing or locked project DBs are skipped, matching the export behavior.
    """
    norm_dirs = []
    seen: set[str] = set()
    for d in dirs:
        if not d:
            continue
        try:
            resolved = str(Path(d).resolve())
        except OSError:
            resolved = str(d)
        if resolved not in seen:
            seen.add(resolved)
            norm_dirs.append(resolved)

    sig = _cache_signature(norm_dirs, root)
    if use_cache and sig in _SPECIMEN_ROWS_CACHE:
        return _copy_summary_payload(_SPECIMEN_ROWS_CACHE[sig])

    specimens: list[Specimen] = []
    grouping: dict[str, dict] = {}
    projects: list[dict] = []

    for ws_dir, conn in _iter_dbs(norm_dirs):
        ws_dir = str(Path(ws_dir).resolve())
        project = _project_descriptor(conn, ws_dir, root)
        projects.append(project)
        project_code = str(project.get("project_code") or "")
        grouped_by_uid = _grouping_summary_for_db(conn, project_code)

        try:
            rows = conn.execute("SELECT * FROM specimens ORDER BY uid").fetchall()
        except sqlite3.Error:
            continue
        for row in rows:
            sp = Specimen.from_row(row)
            sp.owner_project_dir = ws_dir
            specimens.append(sp)

            uid = sp.uid or ""
            key = specimen_grouping_key(ws_dir, uid)
            grouping[key] = dict(
                grouped_by_uid.get(
                    uid,
                    _empty_grouping_summary(project_code),
                )
            )

    dashboard = collect_project_dashboard(norm_dirs, root) if norm_dirs else {}
    payload = {
        "specimens": specimens,
        "grouping": grouping,
        "projects": projects,
        "dashboard": dashboard,
    }
    if use_cache:
        _SPECIMEN_ROWS_CACHE[sig] = _copy_summary_payload(payload)
    return _copy_summary_payload(payload)


def _text(value) -> str:
    return str(value or "").strip()


def _first_text(data: dict, *keys: str) -> str:
    for key in keys:
        value = data.get(key)
        if value not in (None, ""):
            return _text(value)
    return ""


def _normalise_identification(data: dict | str) -> dict:
    if isinstance(data, str):
        return {"scientific_name": data.strip()}
    if not isinstance(data, dict):
        return {}
    return {
        "scientific_name": _first_text(
            data, "scientific_name", "scientificName", "scientificname", "species"
        ),
        "scientific_name_cn": _first_text(
            data, "scientific_name_cn", "scientificNameCn", "speciesCn", "name_cn"
        ),
        "taxon_group": _first_text(data, "taxon_group", "taxonGroup", "class", "class_name"),
        "order_name": _first_text(data, "order_name", "order", "orderName"),
        "family": _first_text(data, "family", "family_name"),
        "genus": _first_text(data, "genus", "genus_name"),
        "notes": _first_text(data, "notes", "note", "remark", "remarks"),
    }


def _has_taxon_identity(row: dict) -> bool:
    return any(
        _text(row.get(key))
        for key in (
            "scientific_name",
            "scientific_name_cn",
            "taxon_group",
            "order_name",
            "family",
            "genus",
        )
    )


def _raw_identifications(raw: dict) -> list[dict]:
    if not isinstance(raw, dict):
        return []
    for key in (
        "identifications",
        "taxa",
        "taxonIdentifications",
        "mixedTaxa",
        "speciesList",
    ):
        value = raw.get(key)
        if isinstance(value, list):
            rows = [_normalise_identification(item) for item in value]
            return [row for row in rows if _has_taxon_identity(row)]
    return []


def _additional_raw_identifications(raw: dict) -> list[dict]:
    if not isinstance(raw, dict):
        return []
    for key in ("additional_identifications", "additionalIdentifications"):
        value = raw.get(key)
        if isinstance(value, list):
            rows = [_normalise_identification(item) for item in value]
            return [row for row in rows if _has_taxon_identity(row)]
    return []


def _primary_identification(sp: Specimen) -> dict:
    raw = sp.raw
    return {
        "scientific_name": _text(sp.scientific_name)
        or _first_text(raw, "scientificName", "scientific_name", "species"),
        "scientific_name_cn": _text(sp.scientific_name_cn)
        or _first_text(raw, "scientificNameCn", "scientific_name_cn", "speciesCn"),
        "taxon_group": _text(sp.taxon_group)
        or _first_text(raw, "taxonGroup", "taxon_group", "class"),
        "order_name": _text(sp.order_name)
        or _first_text(raw, "order", "order_name", "orderName"),
        "family": _text(sp.family) or _first_text(raw, "family"),
        "genus": _text(sp.genus) or _first_text(raw, "genus"),
        "notes": _text(sp.notes) or _first_text(raw, "notes", "remark"),
    }


def _dedupe_identifications(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for row in rows:
        if not _has_taxon_identity(row):
            continue
        key = (
            _text(row.get("scientific_name")).casefold(),
            _text(row.get("scientific_name_cn")),
            _text(row.get("taxon_group")).casefold(),
            _text(row.get("order_name")).casefold(),
            _text(row.get("family")).casefold(),
            _text(row.get("genus")).casefold(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _specimen_identifications(sp: Specimen) -> list[dict]:
    explicit = _raw_identifications(sp.raw)
    additional = _additional_raw_identifications(sp.raw)
    if explicit:
        return _dedupe_identifications(explicit + additional)
    primary = _primary_identification(sp)
    rows = [primary] if _has_taxon_identity(primary) else []
    rows.extend(additional)
    return _dedupe_identifications(rows)


def _collection_event_contexts(conn: sqlite3.Connection) -> dict[tuple[str, str, str, str], dict]:
    cols = _table_columns(conn, "collection_records")
    if not cols:
        return {}
    wanted = (
        "province", "site", "station", "collection_date", "station_label",
        "lon", "lat", "geo_area", "water_body", "collection_time",
        "habitat", "method", "remark",
    )
    select = ", ".join(
        name if name in cols else f"NULL AS {name}"
        for name in wanted
    )
    try:
        rows = conn.execute(f"SELECT {select} FROM collection_records").fetchall()
    except sqlite3.Error:
        return {}

    out: dict[tuple[str, str, str, str], dict] = {}
    for row in rows:
        d = dict(row)
        province = _text(d.get("province"))
        site = _text(d.get("site"))
        station = _text(d.get("station"))
        collection_date = _text(d.get("collection_date"))
        value = {
            "station_label": _text(d.get("station_label")),
            "lon": d.get("lon") if d.get("lon") not in (None, "") else "",
            "lat": d.get("lat") if d.get("lat") not in (None, "") else "",
            "geo_area": _text(d.get("geo_area")),
            "water_body": _text(d.get("water_body")),
            "collection_time": _text(d.get("collection_time")),
            "habitat": _text(d.get("habitat")),
            "method": _text(d.get("method")),
            "remark": _text(d.get("remark")),
        }
        keys = [
            (province, site, station, collection_date),
            (province, site, station, ""),
        ]
        for key in keys:
            out.setdefault(key, value)
    return out


def _collection_context(
    contexts: dict[tuple[str, str, str, str], dict],
    province: str,
    site: str,
    station: str,
    collection_date: str,
) -> dict:
    return (
        contexts.get((province, site, station, collection_date))
        or contexts.get((province, site, station, ""))
        or {}
    )


def collect_taxon_checklist(dirs: list[str], root: str) -> list[dict]:
    """Aggregate identified taxa by workspace/site/station.

    This is a taxonomy/ecology checklist. Unidentified mixed samples and
    workflow-only grouping rows belong in ``collect_sample_processing_summary``.
    """
    buckets: dict[tuple, dict] = {}

    for ws_dir, conn in _iter_dbs(dirs):
        label = _label(ws_dir, root)
        event_contexts = _collection_event_contexts(conn)
        try:
            rows = conn.execute("SELECT * FROM specimens ORDER BY station, uid").fetchall()
        except sqlite3.Error:
            continue
        for row in rows:
            sp = Specimen.from_row(row)
            province = _text(sp.province)
            site = _text(sp.site)
            station = _text(sp.station)
            sample_label = _text(sp.id)
            uid = _text(sp.uid)
            collection_date = _text(sp.collection_date)
            event = _collection_context(
                event_contexts, province, site, station, collection_date
            )
            lon = sp.lon if sp.lon not in (None, "") else event.get("lon", "")
            lat = sp.lat if sp.lat not in (None, "") else event.get("lat", "")
            geo_area = _text(sp.geo_area) or _text(event.get("geo_area"))

            for ident in _specimen_identifications(sp):
                scientific_name = _text(ident.get("scientific_name"))
                scientific_name_cn = _text(ident.get("scientific_name_cn"))
                taxon_key = scientific_name or scientific_name_cn or "未鉴定/混合样"
                key = (
                    label,
                    province,
                    site,
                    station,
                    collection_date,
                    str(lon),
                    str(lat),
                    taxon_key,
                )
                bucket = buckets.setdefault(key, {
                    "workspace_label": label,
                    "province": province,
                    "site": site,
                    "station": station,
                    "station_label": _text(event.get("station_label")),
                    "collection_date": collection_date,
                    "collection_time": _text(event.get("collection_time")),
                    "lon": lon,
                    "lat": lat,
                    "geo_area": geo_area,
                    "water_body": _text(event.get("water_body")),
                    "habitat": _text(event.get("habitat")),
                    "method": _text(event.get("method")),
                    "taxon_group": _text(ident.get("taxon_group")),
                    "order_name": _text(ident.get("order_name")),
                    "family": _text(ident.get("family")),
                    "genus": _text(ident.get("genus")),
                    "scientific_name": scientific_name,
                    "scientific_name_cn": scientific_name_cn or taxon_key,
                    "sample_labels": set(),
                    "uids": set(),
                    "notes": set(),
                })
                for field in ("taxon_group", "order_name", "family", "genus"):
                    if not bucket.get(field) and ident.get(field):
                        bucket[field] = _text(ident.get(field))
                if sample_label:
                    bucket["sample_labels"].add(sample_label)
                if uid:
                    bucket["uids"].add(uid)
                note = _text(ident.get("notes"))
                if note:
                    bucket["notes"].add(note)

    out: list[dict] = []
    for bucket in buckets.values():
        uid_list = sorted(bucket.pop("uids"))
        sample_labels = sorted(bucket.pop("sample_labels"))
        notes = sorted(bucket.pop("notes"))
        row = dict(bucket)
        row.update({
            "sample_labels": "；".join(sample_labels),
            "uids": "；".join(uid_list),
            "evidence_count": len(uid_list),
            "notes": "；".join(notes),
        })
        out.append(row)

    return sorted(
        out,
        key=lambda r: (
            r.get("workspace_label", ""),
            r.get("site", ""),
            r.get("station", ""),
            r.get("collection_date", ""),
            r.get("scientific_name") or r.get("scientific_name_cn") or "",
        ),
    )


def collect_station_species_summary(dirs: list[str], root: str) -> list[dict]:
    """Backward-compatible wrapper for the taxonomic checklist."""
    return collect_taxon_checklist(dirs, root)


_TAXON_CHECKLIST_HEADERS: tuple[tuple[str, str], ...] = (
    ("断面/来源", "workspace_label"),
    ("地区", "province"),
    ("样地", "site"),
    ("站位", "station"),
    ("站位说明", "station_label"),
    ("采集日期", "collection_date"),
    ("采集时间", "collection_time"),
    ("经度", "lon"),
    ("纬度", "lat"),
    ("采集地理区", "geo_area"),
    ("水体", "water_body"),
    ("生境", "habitat"),
    ("采样方法", "method"),
    ("类群", "taxon_group"),
    ("目", "order_name"),
    ("科", "family"),
    ("属", "genus"),
    ("物种拉丁名", "scientific_name"),
    ("物种中文名", "scientific_name_cn"),
    ("证据记录数", "evidence_count"),
    ("物种编号", "sample_labels"),
    ("凭证/样品编号", "uids"),
    ("备注", "notes"),
)


def export_taxon_checklist(
    dirs: list[str],
    root: str,
    out_path: str | None = None,
) -> Path:
    """Write a true taxonomic checklist for field survey review."""
    rows = collect_taxon_checklist(dirs, root)
    if out_path is None:
        out = _default_out(root, "分类名录", "xlsx")
    else:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

    headers = [label for label, _key in _TAXON_CHECKLIST_HEADERS]
    data_rows = [
        [row.get(key, "") for _label, key in _TAXON_CHECKLIST_HEADERS]
        for row in rows
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分类名录"
    ws.freeze_panes = "A2"
    _apply_header_row(ws, headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, headers, data_rows)
    wb.save(str(out))
    return out.resolve()


def export_station_species_summary(
    dirs: list[str],
    root: str,
    out_path: str | None = None,
) -> Path:
    """Backward-compatible wrapper for exporting the taxonomic checklist."""
    return export_taxon_checklist(dirs, root, out_path)


def _grouping_processing_counts(conn: sqlite3.Connection) -> dict[str, dict]:
    cols = _table_columns(conn, "grouping")
    if "uid" not in cols:
        return {}
    wanted = [
        "uid", "status", "jpg_paths", "composed_tiff_path", "archive_zip",
    ]
    select = ", ".join(
        name if name in cols else f"NULL AS {name}"
        for name in wanted
    )
    try:
        rows = conn.execute(f"SELECT {select} FROM grouping").fetchall()
    except sqlite3.Error:
        return {}

    out: dict[str, dict] = {}
    for row in rows:
        uid = _text(row["uid"])
        if not uid:
            continue
        bucket = out.setdefault(uid, {
            "group_count": 0,
            "done_group_count": 0,
            "pending_group_count": 0,
            "group_jpg_count": 0,
            "tiff_count": 0,
            "zip_count": 0,
        })
        bucket["group_count"] += 1
        status = _text(row["status"])
        if status in _GROUPING_COMPOSED_STATUSES:
            bucket["done_group_count"] += 1
        else:
            bucket["pending_group_count"] += 1
        bucket["group_jpg_count"] += len(_json_path_list(row["jpg_paths"]))
        if _text(row["composed_tiff_path"]):
            bucket["tiff_count"] += 1
        if _text(row["archive_zip"]):
            bucket["zip_count"] += 1
    return out


def _current_assignment_counts(conn: sqlite3.Connection) -> dict[str, int]:
    photo_cols = _table_columns(conn, "photos")
    assignment_cols = _table_columns(conn, "photo_assignments")
    if not {"current_assignment_id"} <= photo_cols:
        return {}
    if not {"assignment_id", "specimen_uid"} <= assignment_cols:
        return {}
    try:
        rows = conn.execute(
            """
            SELECT a.specimen_uid AS uid, COUNT(*) AS n
              FROM photos p
              JOIN photo_assignments a ON a.assignment_id = p.current_assignment_id
             WHERE a.specimen_uid IS NOT NULL
               AND TRIM(a.specimen_uid) <> ''
             GROUP BY a.specimen_uid
            """
        ).fetchall()
    except sqlite3.Error:
        return {}
    return {_text(row["uid"]): int(row["n"] or 0) for row in rows if _text(row["uid"])}


def _specimen_rows_by_uid(conn: sqlite3.Connection) -> dict[str, Specimen]:
    try:
        rows = conn.execute("SELECT * FROM specimens").fetchall()
    except sqlite3.Error:
        return {}
    out: dict[str, Specimen] = {}
    for row in rows:
        sp = Specimen.from_row(row)
        uid = _text(sp.uid)
        if uid:
            out[uid] = sp
    return out


def _uid_context(uid: str, sp: Specimen | None = None) -> dict:
    parsed = parse_uid(uid) or {}
    ad_hoc_label = "未归属" if str(uid or "").startswith("~") else ""
    if not parsed:
        parts = [part for part in str(uid or "").split("-") if part]
        if len(parts) >= 4 and parts[-1].isdigit():
            parsed = {
                "province": parts[0],
                "site": parts[1],
                "speciesId": parts[2],
                "dateSegment": parts[-1],
            }
    return {
        "province": _text(getattr(sp, "province", "")) or _text(parsed.get("province")),
        "site": _text(getattr(sp, "site", "")) or _text(parsed.get("site")),
        "station": _text(getattr(sp, "station", "")) or _text(parsed.get("station")),
        "sample_label": _text(getattr(sp, "id", ""))
        or _text(parsed.get("speciesId"))
        or ad_hoc_label,
        "storage": _text(getattr(sp, "storage", "")) or _text(parsed.get("storage")),
        "collection_date": _text(getattr(sp, "collection_date", ""))
        or _text(parsed.get("dateSegment")),
    }


def _taxonomy_status(sp: Specimen | None) -> str:
    if sp is None:
        return "未建标本记录"
    if _specimen_identifications(sp):
        return "已鉴定"
    return "未鉴定"


def collect_sample_processing_summary(dirs: list[str], root: str) -> list[dict]:
    """Summarise sample/photo/grouping workflow state without calling it a checklist."""
    rows: list[dict] = []
    for ws_dir, conn in _iter_dbs(dirs):
        label = _label(ws_dir, root)
        specimens = _specimen_rows_by_uid(conn)
        grouping = _grouping_processing_counts(conn)
        photo_counts = _current_assignment_counts(conn)
        all_uids = sorted(set(specimens) | set(grouping) | set(photo_counts))
        for uid in all_uids:
            sp = specimens.get(uid)
            ctx = _uid_context(uid, sp)
            g = grouping.get(uid, {})
            photo_count = int(photo_counts.get(uid) or 0)
            group_jpg_count = int(g.get("group_jpg_count") or 0)
            rows.append({
                "workspace_label": label,
                "uid": uid,
                "province": ctx["province"],
                "site": ctx["site"],
                "station": ctx["station"],
                "sample_label": ctx["sample_label"],
                "storage": ctx["storage"],
                "collection_date": ctx["collection_date"],
                "taxonomy_status": _taxonomy_status(sp),
                "photo_count": photo_count,
                "group_count": int(g.get("group_count") or 0),
                "done_group_count": int(g.get("done_group_count") or 0),
                "pending_group_count": int(g.get("pending_group_count") or 0),
                "tiff_count": int(g.get("tiff_count") or 0),
                "zip_count": int(g.get("zip_count") or 0),
                "group_jpg_count": group_jpg_count,
            })
    return sorted(
        rows,
        key=lambda r: (
            r.get("workspace_label", ""),
            r.get("site", ""),
            r.get("station", ""),
            r.get("sample_label", ""),
            r.get("uid", ""),
        ),
    )


_SAMPLE_PROCESSING_HEADERS: tuple[tuple[str, str], ...] = (
    ("断面/来源", "workspace_label"),
    ("地区", "province"),
    ("样地", "site"),
    ("站位", "station"),
    ("物种编号", "sample_label"),
    ("保存方式", "storage"),
    ("采集日期", "collection_date"),
    ("分类状态", "taxonomy_status"),
    ("当前照片数", "photo_count"),
    ("分组数", "group_count"),
    ("已完成分组", "done_group_count"),
    ("待处理分组", "pending_group_count"),
    ("TIFF 数", "tiff_count"),
    ("ZIP 数", "zip_count"),
    ("分组 JPG 数", "group_jpg_count"),
    ("UID/样品编号", "uid"),
)


def export_sample_processing_summary(
    dirs: list[str],
    root: str,
    out_path: str | None = None,
) -> Path:
    rows = collect_sample_processing_summary(dirs, root)
    if out_path is None:
        out = _default_out(root, "样品处理概况", "xlsx")
    else:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

    headers = [label for label, _key in _SAMPLE_PROCESSING_HEADERS]
    data_rows = [
        [row.get(key, "") for _label, key in _SAMPLE_PROCESSING_HEADERS]
        for row in rows
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "样品处理概况"
    ws.freeze_panes = "A2"
    _apply_header_row(ws, headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, headers, data_rows)
    wb.save(str(out))
    return out.resolve()


def _default_out(root: str, suffix: str, ext: str) -> Path:
    """``<root>/_data/exports/<basename(root)>_<suffix>.<ext>``，并 mkdir 父目录。"""
    root_path = Path(root)
    base = os.path.basename(os.path.normpath(root)) or root_path.name
    out = root_path / "_data" / "exports" / f"{base}_{suffix}.{ext}"
    out.parent.mkdir(parents=True, exist_ok=True)
    return out


# ── 0. 项目看板汇总 ─────────────────────────────────────────────────────────────

def _is_rna_storage(storage: Optional[str]) -> bool:
    return bool(storage) and str(storage).upper().startswith("R")


def _is_alcohol_storage(storage: Optional[str]) -> bool:
    """Current storage codes describe alcohol fixation unless R-prefixed RNA."""
    return bool(storage) and not _is_rna_storage(storage)


def _person_key(value: Optional[str]) -> str:
    s = str(value or "").strip()
    return s or "未填写"


def _counter_rows(counter: Counter) -> list[dict]:
    return [
        {"name": name, "count": count}
        for name, count in sorted(counter.items(), key=lambda kv: (-kv[1], kv[0]))
    ]


def collect_project_dashboard(dirs: list[str], root: str) -> dict:
    """Return dashboard-ready counts across workspaces.

    The result is deliberately plain JSON-shaped data so both Qt views and a
    future collaboration host API can use it directly.
    """
    totals = Counter()
    photographers: Counter = Counter()
    collectors: Counter = Counter()
    identifiers: Counter = Counter()
    label_printers: Counter = Counter()
    photo_assigners: Counter = Counter()
    by_workspace: list[dict] = []

    for ws_dir, conn in _iter_dbs(dirs):
        label = _label(ws_dir, root)
        ws = Counter()
        ws["workspace_count"] = 1

        try:
            rows = conn.execute(
                "SELECT uid, storage, collector, photographer, identifier FROM specimens"
            ).fetchall()
        except sqlite3.Error:
            rows = []
        for row in rows:
            ws["specimen_count"] += 1
            if _is_rna_storage(row["storage"]):
                ws["rna_specimen_count"] += 1
            elif _is_alcohol_storage(row["storage"]):
                ws["alcohol_specimen_count"] += 1
            photographers[_person_key(row["photographer"])] += 1
            collectors[_person_key(row["collector"])] += 1
            identifiers[_person_key(row["identifier"])] += 1

        try:
            prow = conn.execute(
                """
                SELECT COUNT(*) AS total,
                       SUM(CASE WHEN current_assignment_id IS NULL THEN 1 ELSE 0 END) AS unassigned
                  FROM photos
                """
            ).fetchone()
            ws["photo_count"] += int(prow["total"] or 0)
            ws["unassigned_photo_count"] += int(prow["unassigned"] or 0)
        except sqlite3.Error:
            pass

        grouping_cols = _table_columns(conn, "grouping")
        if "uid" in grouping_cols:
            select = ", ".join(
                name if name in grouping_cols else f"NULL AS {name}"
                for name in ("composed_tiff_path", "archive_zip")
            )
            try:
                rows = conn.execute(f"SELECT {select} FROM grouping").fetchall()
                for row in rows:
                    if str(row["composed_tiff_path"] or "").strip():
                        ws["tiff_count"] += 1
                    if str(row["archive_zip"] or "").strip():
                        ws["zip_count"] += 1
            except sqlite3.Error:
                pass

        try:
            assigns = conn.execute(
                """
                SELECT assigned_by, COUNT(*) AS n
                  FROM photo_assignments
                 WHERE is_current=1
                 GROUP BY assigned_by
                """
            ).fetchall()
            for row in assigns:
                photo_assigners[_person_key(row["assigned_by"])] += int(row["n"] or 0)
        except sqlite3.Error:
            pass

        try:
            prints = conn.execute(
                """
                SELECT actor, bucket, COUNT(*) AS events, SUM(label_count) AS labels
                  FROM label_print_events
                 GROUP BY actor, bucket
                """
            ).fetchall()
            for row in prints:
                events = int(row["events"] or 0)
                labels = int(row["labels"] or 0)
                bucket = str(row["bucket"] or "")
                ws["label_print_event_count"] += events
                ws["label_print_count"] += labels
                if bucket == "tissue":
                    ws["tissue_label_print_count"] += labels
                else:
                    ws["sample_label_print_count"] += labels
                label_printers[_person_key(row["actor"])] += labels
        except sqlite3.Error:
            pass

        totals.update(ws)
        by_workspace.append({
            "workspace_label": label,
            "workspace_dir": ws_dir,
            "specimen_count": ws["specimen_count"],
            "alcohol_specimen_count": ws["alcohol_specimen_count"],
            "rna_specimen_count": ws["rna_specimen_count"],
            "photo_count": ws["photo_count"],
            "unassigned_photo_count": ws["unassigned_photo_count"],
            "tiff_count": ws["tiff_count"],
            "zip_count": ws["zip_count"],
            "label_print_event_count": ws["label_print_event_count"],
            "sample_label_print_count": ws["sample_label_print_count"],
            "tissue_label_print_count": ws["tissue_label_print_count"],
            "label_print_count": ws["label_print_count"],
        })

    return {
        "root": root,
        "totals": {
            "workspace_count": totals["workspace_count"],
            "specimen_count": totals["specimen_count"],
            "alcohol_specimen_count": totals["alcohol_specimen_count"],
            "rna_specimen_count": totals["rna_specimen_count"],
            "photo_count": totals["photo_count"],
            "unassigned_photo_count": totals["unassigned_photo_count"],
            "tiff_count": totals["tiff_count"],
            "zip_count": totals["zip_count"],
            "label_print_event_count": totals["label_print_event_count"],
            "sample_label_print_count": totals["sample_label_print_count"],
            "tissue_label_print_count": totals["tissue_label_print_count"],
            "label_print_count": totals["label_print_count"],
        },
        "by_workspace": sorted(by_workspace, key=lambda r: r["workspace_label"]),
        "by_person": {
            "photographer": _counter_rows(photographers),
            "collector": _counter_rows(collectors),
            "identifier": _counter_rows(identifiers),
            "photo_assigner": _counter_rows(photo_assigners),
            "label_printer": _counter_rows(label_printers),
        },
    }


# ── 1. 标本汇总 ─────────────────────────────────────────────────────────────────

def export_specimen_summary(
    dirs: list[str],
    root: str,
    out_path: str | None = None,
) -> Path:
    """合并各工作区的 specimens，导出带 断面 前置列的 34 列标本汇总 Excel。"""
    all_specs: list[Specimen] = []
    for ws_dir, conn in _iter_dbs(dirs):
        try:
            rows = conn.execute("SELECT * FROM specimens ORDER BY uid").fetchall()
        except sqlite3.Error:
            continue
        for r in rows:
            sp = Specimen.from_row(r)
            sp.owner_project_dir = ws_dir
            all_specs.append(sp)

    if out_path is None:
        out = _default_out(root, "标本汇总", "xlsx")
    else:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

    return export_excel(
        all_specs,
        out,
        extra_leading=[("断面", lambda s: _label(s.owner_project_dir or "", root))],
    )


# ── 2. 采集站位汇总 ─────────────────────────────────────────────────────────────

def export_collection_summary(
    dirs: list[str],
    root: str,
    out_path: str | None = None,
) -> Path:
    """合并各工作区的 collection_records，写带 断面 前置列的「采集站位汇总」表。"""
    headers = ["断面", *_COLLECTION_COLUMNS]
    data_rows: list[list] = []
    for ws_dir, conn in _iter_dbs(dirs):
        try:
            records = list_records(conn)
        except sqlite3.Error:
            continue
        label = _label(ws_dir, root)
        for rec in records:
            row: list = [label]
            for col in _COLLECTION_COLUMNS:
                val = rec.get(col)
                if col in ("lon", "lat"):
                    # 数字按数字写，缺失写空串（CLAUDE.md：空经纬度不写 0）
                    row.append(val if val not in (None, "") else "")
                else:
                    row.append("" if val is None else val)
            data_rows.append(row)

    if out_path is None:
        out = _default_out(root, "采集站位汇总", "xlsx")
    else:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采集站位汇总"
    ws.freeze_panes = "A2"
    _apply_header_row(ws, headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, headers, data_rows)

    wb.save(str(out))
    return out.resolve()


# ── 3. 质控报告 ─────────────────────────────────────────────────────────────────

def _entity_from_loc(province, site, station) -> str:
    """以 / 连接 (province, site, station) 中非空的段，作为缺经纬度站位的实体名。"""
    return "/".join(str(x) for x in (province, site, station) if x not in (None, ""))


def collect_qc_findings(dirs: list[str], root: str) -> list[dict]:
    """跨工作区计算质控发现。每条 finding = ``{workspace_label, category, entity, detail}``。

    类别：
      - "缺经纬度站位"：collection_records 中 lon 或 lat 为 NULL 的行。
      - "有标本无成片"：无 status∈(composed,organized) 的 grouping 行的 specimen uid。
      - "分类不完整"：scientific_name 或 family 为空的 specimen。
      - "跨断面UID冲突"：同一 uid 出现在 ≥2 个不同工作区标签下。
      - "文件夹存在但非工作区"：root 下扫描到的、has_data=False 的叶节点目录。
    """
    findings: list[dict] = []
    uid_to_labels: dict[str, set[str]] = {}

    for ws_dir, conn in _iter_dbs(dirs):
        label = _label(ws_dir, root)

        # 缺经纬度站位
        try:
            rows = conn.execute(
                "SELECT province, site, station FROM collection_records "
                "WHERE lon IS NULL OR lat IS NULL"
            ).fetchall()
            for r in rows:
                findings.append({
                    "workspace_label": label,
                    "category": "缺经纬度站位",
                    "entity": _entity_from_loc(r["province"], r["site"], r["station"]),
                    "detail": "经度/纬度缺失",
                })
        except sqlite3.Error:
            pass

        # 有成片的 uid 集合（grouping 表可能在裸库中缺失 → 当作无成片）
        composed_uids: set[str] = set()
        try:
            placeholders = ", ".join("?" for _ in _GROUPING_COMPOSED_STATUSES)
            grows = conn.execute(
                f"SELECT DISTINCT uid FROM grouping WHERE status IN ({placeholders})",
                _GROUPING_COMPOSED_STATUSES,
            ).fetchall()
            composed_uids = {g["uid"] for g in grows}
        except sqlite3.Error:
            composed_uids = set()

        # specimens：有标本无成片 + 分类不完整 + uid→labels 累积
        try:
            srows = conn.execute(
                "SELECT uid, scientific_name, family FROM specimens"
            ).fetchall()
        except sqlite3.Error:
            srows = []
        for s in srows:
            uid = s["uid"]
            uid_to_labels.setdefault(uid, set()).add(label)
            if uid not in composed_uids:
                findings.append({
                    "workspace_label": label,
                    "category": "有标本无成片",
                    "entity": uid,
                    "detail": "无 composed/organized 成片",
                })
            sci = s["scientific_name"]
            fam = s["family"]
            if sci in (None, "") or fam in (None, ""):
                findings.append({
                    "workspace_label": label,
                    "category": "分类不完整",
                    "entity": uid,
                    "detail": "学名或科缺失",
                })

    # 跨断面UID冲突：同一 uid 出现在 ≥2 个不同标签
    for uid, labels in sorted(uid_to_labels.items()):
        if len(labels) >= 2:
            label_list = sorted(labels)
            findings.append({
                "workspace_label": "、".join(label_list),
                "category": "跨断面UID冲突",
                "entity": uid,
                "detail": "出现于断面：" + "、".join(label_list),
            })

    # 文件夹存在但非工作区：root 下扫描到的、非工作区的叶节点
    findings.extend(_orphan_folder_findings(root))

    return findings


def _orphan_folder_findings(root: str) -> list[dict]:
    """root 下扫描到的「文件夹存在但非工作区」叶节点（has_data=False 且无子节点）。"""
    out: list[dict] = []
    try:
        tree = scan_tree(root, use_cache=False)
    except OSError:
        return out

    def collect_orphan_folder_leaf(node: dict) -> None:
        children = node.get("children", [])
        if not children and not node.get("has_data"):
            # 叶子且非工作区（root 自身若为空叶也会被纳入；用 path != root 排除根）
            if os.path.normpath(node["path"]) != os.path.normpath(root):
                out.append({
                    "workspace_label": _label(node["path"], root),
                    "category": "文件夹存在但非工作区",
                    "entity": node.get("name", ""),
                    "detail": node["path"],
                })
        for child in children:
            collect_orphan_folder_leaf(child)

    collect_orphan_folder_leaf(tree)
    return out


_QC_CATEGORY_ORDER: tuple[str, ...] = (
    "缺经纬度站位",
    "有标本无成片",
    "分类不完整",
    "跨断面UID冲突",
    "文件夹存在但非工作区",
)

_QC_SHEET_HEADERS = ("类别", "断面", "实体", "详情")


def export_qc_report(
    dirs: list[str],
    root: str,
    out_dir: str | None = None,
) -> tuple[Path, Path]:
    """写出 ``<basename>_质控报告.html`` 和 ``<basename>_质控报告.xlsx``。返回 (html, xlsx)。"""
    findings = collect_qc_findings(dirs, root)

    base = os.path.basename(os.path.normpath(root)) or Path(root).name
    if out_dir is None:
        out_base = Path(root) / "_data" / "exports"
    else:
        out_base = Path(out_dir)
    out_base.mkdir(parents=True, exist_ok=True)
    html_path = out_base / f"{base}_质控报告.html"
    xlsx_path = out_base / f"{base}_质控报告.xlsx"

    # 按类别分组（保持固定顺序，未知类别附后）
    by_cat: dict[str, list[dict]] = {}
    for f in findings:
        by_cat.setdefault(f["category"], []).append(f)
    ordered_cats = [c for c in _QC_CATEGORY_ORDER if c in by_cat]
    ordered_cats += [c for c in by_cat if c not in _QC_CATEGORY_ORDER]

    _write_qc_html(html_path, base, by_cat, ordered_cats)
    _write_qc_xlsx(xlsx_path, findings)

    return html_path.resolve(), xlsx_path.resolve()


def _html_escape(text) -> str:
    s = "" if text is None else str(text)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _write_qc_html(
    html_path: Path,
    base: str,
    by_cat: dict[str, list[dict]],
    ordered_cats: list[str],
) -> None:
    total = sum(len(v) for v in by_cat.values())
    parts: list[str] = []
    parts.append("<!DOCTYPE html><html lang='zh'><head><meta charset='utf-8'>")
    parts.append(f"<title>{_html_escape(base)} 质控报告</title>")
    parts.append(
        "<style>"
        "body{font-family:'Microsoft YaHei',sans-serif;margin:24px;color:#222;}"
        "h1{font-size:20px;}h2{font-size:16px;margin-top:24px;color:#2C5F8A;}"
        "table{border-collapse:collapse;width:100%;margin-top:8px;}"
        "th,td{border:1px solid #cbd5e1;padding:6px 10px;text-align:left;font-size:13px;}"
        "th{background:#2C5F8A;color:#fff;}"
        "tr:nth-child(even){background:#eef3fa;}"
        ".summary{background:#f1f5f9;padding:10px 14px;border-radius:6px;}"
        ".summary li{margin:2px 0;}"
        "</style></head><body>"
    )
    parts.append(f"<h1>{_html_escape(base)} 质控报告</h1>")
    parts.append("<div class='summary'><strong>汇总</strong><ul>")
    parts.append(f"<li>问题总数：{total}</li>")
    for cat in ordered_cats:
        parts.append(f"<li>{_html_escape(cat)}：{len(by_cat[cat])}</li>")
    parts.append("</ul></div>")

    if not ordered_cats:
        parts.append("<p>未发现质控问题。</p>")
    for cat in ordered_cats:
        parts.append(f"<h2>{_html_escape(cat)}（{len(by_cat[cat])}）</h2>")
        parts.append("<table><thead><tr><th>断面</th><th>实体</th><th>详情</th></tr></thead><tbody>")
        for f in by_cat[cat]:
            parts.append(
                "<tr><td>{}</td><td>{}</td><td>{}</td></tr>".format(
                    _html_escape(f.get("workspace_label")),
                    _html_escape(f.get("entity")),
                    _html_escape(f.get("detail")),
                )
            )
        parts.append("</tbody></table>")
    parts.append("</body></html>")

    html_path.write_text("".join(parts), encoding="utf-8")


def _write_qc_xlsx(xlsx_path: Path, findings: list[dict]) -> None:
    headers = list(_QC_SHEET_HEADERS)
    data_rows: list[list] = []
    for f in findings:
        data_rows.append([
            f.get("category", ""),
            f.get("workspace_label", ""),
            f.get("entity", ""),
            f.get("detail", ""),
        ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质控报告"
    ws.freeze_panes = "A2"
    _apply_header_row(ws, headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, headers, data_rows)

    wb.save(str(xlsx_path))
