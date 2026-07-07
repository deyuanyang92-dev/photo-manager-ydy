"""export_service.py — Specimen data export: Excel (34-col), CSV, Darwin Core.

Oracle:
  - server.js:595-721  (_buildSpecimensWorkbook — 34-column layout)
  - db_manager.py      (darwin_core view definition)
  - Specimen dataclass (field names)

Three public functions:
  export_excel(specimens, path, columns=None)   — openpyxl workbook, blue header
  export_csv(specimens, path, columns=None)     — UTF-8 with BOM (Excel-compatible)
  export_darwin_core(db, path)                  — reads darwin_core VIEW from SQLite
"""
from __future__ import annotations

import csv
import sqlite3
from datetime import date
from pathlib import Path
from typing import Optional, Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.specimen import Specimen
from app.utils.naming import specimen_date_seg

# ── Storage-code helpers (mirrors server.js _spPresDetail / _spIsRNA) ─────────

from app.services.project_settings_service import BUILTIN_STORAGES

_STORAGE_LABELS: dict[str, str] = {
    str(entry["code"]): str(entry["detail"]) for entry in BUILTIN_STORAGES
}


def _pres_detail(storage: Optional[str]) -> str:
    """Return human-readable preservation description from storage code."""
    if not storage:
        return ""
    return _STORAGE_LABELS.get(storage.upper(), storage)


def _is_rna(storage: Optional[str]) -> bool:
    """Return True when the storage code indicates RNAlater treatment (R prefix)."""
    if not storage:
        return False
    return str(storage).upper().startswith("R")


def _meta_score(sp: Specimen) -> int:
    """Return metadata completeness percentage (0-100).

    Mirrors server.js _spMetaScore: counts 5 key fields.
    """
    fields = [sp.scientific_name, sp.family, sp.collector, sp.lon, sp.lat]
    filled = sum(1 for f in fields if f is not None and str(f).strip() != "")
    return round(filled / len(fields) * 100)


def _taxon_complete(sp: Specimen) -> str:
    """Return checkmark if scientific_name and family are set, else cross."""
    return "✓" if (sp.scientific_name and sp.family) else "✗"


def _date_seg(sp: Specimen) -> str:
    return specimen_date_seg(sp.collection_date, sp.photo_date)


# ── Column definitions (34 columns, mirrors Sheet 1 of _buildSpecimensWorkbook) ─

#: Master column list (header label, accessor callable).
#: The accessor receives a Specimen and returns a scalar value.
COLUMNS: list[tuple[str, callable]] = [
    ("标本唯一编号",      lambda s: s.uid or ""),
    ("物种拼音编号",      lambda s: s.id or ""),
    ("物种中名",          lambda s: s.scientific_name_cn or ""),
    ("物种拉丁名",        lambda s: s.scientific_name or ""),
    ("类群中名",          lambda s: s.taxon_group_cn or ""),
    ("类群拉丁名",        lambda s: s.taxon_group or ""),
    ("目中名",            lambda s: s.order_cn or ""),
    ("目拉丁名",          lambda s: s.order_name or ""),
    ("科中名",            lambda s: s.family_cn or ""),
    ("科拉丁名",          lambda s: s.family or ""),
    ("属中名",            lambda s: s.genus_cn or ""),
    ("属拉丁名",          lambda s: s.genus or ""),
    ("省份代码",          lambda s: s.province or ""),
    ("样地代码",          lambda s: s.site or ""),
    ("站位",              lambda s: s.station or ""),
    ("采集地全称",        lambda s: s.geo_area or (
        (s.province or "")
        + ("·" + s.site if s.site else "")
        + ("·" + s.station if s.station else "")
    )),
    ("经度",              lambda s: float(s.lon) if s.lon is not None else ""),
    ("纬度",              lambda s: float(s.lat) if s.lat is not None else ""),
    ("保存方式代码",      lambda s: s.storage or ""),
    ("固定方式全文",      lambda s: _pres_detail(s.storage)),
    ("采集日期",          lambda s: s.collection_date or ""),
    ("拍照日期",          lambda s: s.photo_date or ""),
    ("日期段",            _date_seg),
    ("采集人",            lambda s: s.collector or ""),
    ("拍摄人",            lambda s: s.photographer or ""),
    ("鉴定人",            lambda s: s.identifier or ""),
    ("角度标记",          lambda s: s.angle or ""),
    ("分类完整",          _taxon_complete),
    ("RNA标本",           lambda s: "✓" if _is_rna(s.storage) else "✗"),
    ("Metadata完整度(%)", _meta_score),
    ("备注",              lambda s: s.notes or ""),
    ("拍照备注",          lambda s: s.photo_notes or ""),
    ("元数据标志",        lambda s: s.metadata or 0),
    ("置顶",              lambda s: "✓" if s.pinned else "✗"),
]

# Verify we have exactly 34 columns
assert len(COLUMNS) == 34, f"Expected 34 export columns, got {len(COLUMNS)}"

COLUMN_HEADERS: list[str] = [h for h, _ in COLUMNS]


# ── openpyxl style constants ───────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="2C5F8A")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")
_HDR_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALT_FILL = PatternFill("solid", fgColor="EEF3FA")


def _apply_header_row(ws, headers: list[str]) -> None:
    """Write a styled header row (row 1) to *ws*."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border = _HDR_BORDER
    ws.row_dimensions[1].height = 22


def _auto_col_widths(ws, headers: list[str], data_rows: list[list]) -> None:
    """Set column widths based on header and data content (max 50, min 8)."""
    for col_idx, header in enumerate(headers, start=1):
        max_len = sum(2 if ord(c) > 127 else 1 for c in header)
        for row in data_rows:
            val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 1, 8), 50)


def _build_data_rows(
    specimens: Sequence[Specimen],
    columns: list[tuple[str, callable]],
) -> list[list]:
    """Convert specimen list to list of row lists using column accessors."""
    rows = []
    for sp in specimens:
        row = []
        for _header, accessor in columns:
            try:
                row.append(accessor(sp))
            except Exception:
                row.append("")
        rows.append(row)
    return rows


def _resolve_columns(
    columns: Optional[list[str]],
) -> list[tuple[str, callable]]:
    """Return the filtered column definitions matching requested header labels.

    If *columns* is None or empty, return all 34 columns.
    Unknown column names are silently skipped.
    """
    if not columns:
        return COLUMNS
    col_map = {h: acc for h, acc in COLUMNS}
    return [(h, col_map[h]) for h in columns if h in col_map]


# ── Public API ─────────────────────────────────────────────────────────────────

def export_excel(
    specimens: Sequence[Specimen],
    path: str | Path,
    columns: Optional[list[str]] = None,
    extra_leading: Optional[list[tuple[str, callable]]] = None,
) -> Path:
    """Export specimens to an Excel file (.xlsx) at *path*.

    Parameters
    ----------
    specimens:
        Iterable of Specimen instances to export.
    path:
        Destination file path (will be created or overwritten).
    columns:
        Optional list of column header strings to include.
        Defaults to all 34 columns.
    extra_leading:
        Optional list of ``(header, accessor)`` pairs prepended *before* the
        resolved columns. Each accessor takes a Specimen and returns a scalar
        (same contract as the master columns; a raising accessor yields "").
        When None/empty the output is byte-for-byte identical to before — this
        is a purely additive parameter and must never alter the default 34-col
        layout (red line: oracle server.js:595-721).

    Returns
    -------
    Path
        Resolved path of the written file.

    Oracle: server.js:595-721 (_buildSpecimensWorkbook Sheet 1)
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    active_cols = _resolve_columns(columns)
    if extra_leading:
        active_cols = list(extra_leading) + active_cols
    headers = [h for h, _ in active_cols]
    data_rows = _build_data_rows(specimens, active_cols)

    wb = openpyxl.Workbook()
    wb.properties.creator = "拍照工作台"

    ws = wb.active
    ws.title = "标本汇总"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _apply_header_row(ws, headers)

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        # Alternating row fill (odd data rows = white, even = light blue)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL

    _auto_col_widths(ws, headers, data_rows)

    # Metadata sheet
    ws2 = wb.create_sheet("导出信息")
    ws2["A1"] = "导出日期"
    ws2["B1"] = date.today().isoformat()
    ws2["A2"] = "标本数"
    ws2["B2"] = len(data_rows)
    ws2["A3"] = "列数"
    ws2["B3"] = len(headers)

    wb.save(str(path))
    return path.resolve()


def export_csv(
    specimens: Sequence[Specimen],
    path: str | Path,
    columns: Optional[list[str]] = None,
) -> Path:
    """Export specimens to a CSV file with UTF-8 BOM (Excel-compatible).

    Parameters
    ----------
    specimens:
        Iterable of Specimen instances to export.
    path:
        Destination file path (will be created or overwritten).
    columns:
        Optional list of column header strings to include.
        Defaults to all 34 columns.

    Returns
    -------
    Path
        Resolved path of the written file.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    active_cols = _resolve_columns(columns)
    headers = [h for h, _ in active_cols]
    data_rows = _build_data_rows(specimens, active_cols)

    # utf-8-sig = UTF-8 with BOM; Excel on Windows opens correctly
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)

    return path.resolve()


# ── WoRMS Match-Taxa annotated export ─────────────────────────────────────────
#
# Writes the user's ORIGINAL table verbatim (all columns, original order — zero
# loss) and APPENDS selected WoRMS columns derived from each row's match result.
# Never merges WoRMS data into an original column, so a `*Cn` (Chinese-name)
# column simply passes through as one of the original columns (red line).

def _match_cell_text(v) -> str:
    """Match-string: None → '', everything else str()."""
    return "" if v is None else str(v)


def _m_best(result: dict, field: str) -> str:
    return _match_cell_text((result.get("best") or {}).get(field))


_RANK_FIELDS = ("kingdom", "phylum", "class", "order", "family", "genus")


def _m_rank(result: dict, rank_key: str) -> str:
    """Rank name from the (optional) full chain, else the best record's inline field."""
    chain = result.get("chain")
    if chain:
        for node in chain:
            if str(node.get("rank", "")).lower() == rank_key:
                return _match_cell_text(node.get("scientificname"))
    return _match_cell_text((result.get("best") or {}).get(rank_key))


def _m_classification(result: dict) -> str:
    chain = result.get("chain")
    if chain:
        return " > ".join(_match_cell_text(n.get("scientificname")) for n in chain if n.get("scientificname"))
    best = result.get("best") or {}
    return " > ".join(_match_cell_text(best.get(r)) for r in _RANK_FIELDS if best.get(r))


_ENV_FLAGS = (("isMarine", "海洋"), ("isBrackish", "半咸水"),
              ("isFreshwater", "淡水"), ("isTerrestrial", "陆生"))


def _m_environment(result: dict) -> str:
    best = result.get("best") or {}
    return ",".join(label for k, label in _ENV_FLAGS if best.get(k))


_RESOLUTION_CN = {"matched": "已匹配", "near": "待确认", "ambiguous": "多候选", "none": "无匹配"}


def _m_resolution(result: dict) -> str:
    r = result.get("resolution", "")
    return _RESOLUTION_CN.get(r, r)


#: WoRMS append-column catalog: (key, header label, extractor(result) -> str).
#: ``key`` is the stable identifier the UI passes in ``append_cols``; each entry
#: carries a Chinese AND an English header (output language is user-selectable),
#: plus the extractor.  Mirrors the WoRMS website's output-column choices.
#: NOTE: WoRMS "Qualitystatus" is NOT exposed by the public REST
#: AphiaRecordsByMatchNames endpoint, so it cannot be a column here; "分类状态 /
#: Taxon status" (``status``) is the available taxon-status field.
MATCH_APPEND_COLUMNS: list[tuple[str, str, str, callable]] = [
    ("aphia_id",           "AphiaID",      "AphiaID",          lambda r: _m_best(r, "AphiaID")),
    ("matched_name",       "匹配名",        "ScientificName",   lambda r: _m_best(r, "scientificname")),
    ("accepted_name",      "接受名",        "Accepted name",    lambda r: _match_cell_text((r.get("best") or {}).get("valid_name")
                                                                            or (r.get("best") or {}).get("scientificname"))),
    ("accepted_aphia_id",  "接受AphiaID",   "Accepted AphiaID", lambda r: _match_cell_text((r.get("best") or {}).get("valid_AphiaID")
                                                                            or (r.get("best") or {}).get("AphiaID"))),
    ("authority",          "命名人",        "Authority",        lambda r: _m_best(r, "authority")),
    ("accepted_authority", "接受名命名人",   "Accepted Authority", lambda r: _m_best(r, "valid_authority")),
    ("status",             "分类状态",       "Taxon status",     lambda r: _m_best(r, "status")),
    ("unacceptreason",     "不接受原因",     "Unaccept reason",  lambda r: _m_best(r, "unacceptreason")),
    ("match_type",         "命中类型",       "Match type",       lambda r: _m_best(r, "match_type")),
    ("rank",               "阶元",          "Rank",             lambda r: _m_best(r, "rank")),
    ("kingdom",            "界",            "Kingdom",          lambda r: _m_rank(r, "kingdom")),
    ("phylum",             "门",            "Phylum",           lambda r: _m_rank(r, "phylum")),
    ("class",              "纲",            "Class",            lambda r: _m_rank(r, "class")),
    ("order",              "目",            "Order",            lambda r: _m_rank(r, "order")),
    ("family",             "科",            "Family",           lambda r: _m_rank(r, "family")),
    ("genus",              "属",            "Genus",            lambda r: _m_rank(r, "genus")),
    ("classification",     "完整分类链",      "Classification",   _m_classification),
    ("lsid",               "LSID",         "LSID",             lambda r: _m_best(r, "lsid")),
    ("tsn",                "TSN",          "TSN",              lambda r: _m_best(r, "tsn")),
    ("environment",        "生境",          "Environment",      _m_environment),
    ("citation",           "引用",          "Citation",         lambda r: _m_best(r, "citation")),
    ("url",                "WoRMS链接",     "URL",              lambda r: _m_best(r, "url")),
    ("resolution",         "复核状态",       "Match status",     _m_resolution),
]

#: Output-header language: "zh" (中文), "en" (English), "both" (双语).
OUTPUT_LANGS = ("zh", "en", "both")


def _col_header(zh: str, en: str, lang: str) -> str:
    if lang == "en":
        return en
    if lang == "both" and zh != en:
        return f"{zh} ({en})"
    return zh


def _resolve_match_cols(append_cols: Optional[list[str]]) -> list[tuple[str, str, str, callable]]:
    """Filter the append catalog to the requested keys (None/empty → all)."""
    if not append_cols:
        return MATCH_APPEND_COLUMNS
    by_key = {k: (k, zh, en, fn) for k, zh, en, fn in MATCH_APPEND_COLUMNS}
    return [by_key[k] for k in append_cols if k in by_key]


def _append_headers(cols, lang: str) -> list[str]:
    return [_col_header(zh, en, lang) for _k, zh, en, _fn in cols]


def _annotated_rows(headers, rows, results, cols) -> list[list]:
    """Build [original cells…, appended WoRMS cells…] per row, index-aligned."""
    data_rows: list[list] = []
    for i, row in enumerate(rows):
        result = results[i] if i < len(results) else {}
        out = [row.get(h, "") for h in headers]
        for _key, _zh, _en, fn in cols:
            try:
                out.append(fn(result))
            except Exception:
                out.append("")
        data_rows.append(out)
    return data_rows


def _match_counts(results) -> dict:
    counts: dict[str, int] = {}
    for r in results:
        k = r.get("resolution", "")
        counts[k] = counts.get(k, 0) + 1
    return counts


def export_annotated_xlsx(
    headers: list[str],
    rows: list[dict],
    results: list[dict],
    append_cols: Optional[list[str]],
    path: str | Path,
    lang: str = "zh",
) -> Path:
    """Write the original table + appended WoRMS columns to an .xlsx file.

    Parameters
    ----------
    headers / rows:
        The original table as returned by ``coord_import_service.read_table``
        (preserved verbatim, in order — zero loss).
    results:
        Per-row match results from ``WormsService.match_names`` (index-aligned to
        *rows*).  May carry an optional ``chain`` list for full classification.
    append_cols:
        Stable keys from :data:`MATCH_APPEND_COLUMNS` to append; None → all.
    lang:
        Header language for the appended WoRMS columns — "zh" / "en" / "both".
        Original columns are always kept verbatim.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    cols = _resolve_match_cols(append_cols)
    out_headers = list(headers) + _append_headers(cols, lang)
    data_rows = _annotated_rows(headers, rows, results, cols)

    wb = openpyxl.Workbook()
    wb.properties.creator = "拍照工作台"
    ws = wb.active
    ws.title = "匹配结果"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(len(out_headers), 1))}1"

    _apply_header_row(ws, out_headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(out_headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, out_headers, data_rows)

    counts = _match_counts(results)
    ws2 = wb.create_sheet("匹配信息")
    info = [
        ("导出日期", date.today().isoformat()),
        ("总行数", len(rows)),
        ("已匹配", counts.get("matched", 0)),
        ("待确认", counts.get("near", 0)),
        ("多候选", counts.get("ambiguous", 0)),
        ("无匹配", counts.get("none", 0)),
    ]
    for i, (k, v) in enumerate(info, start=1):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)

    wb.save(str(path))
    return path.resolve()


def export_annotated_csv(
    headers: list[str],
    rows: list[dict],
    results: list[dict],
    append_cols: Optional[list[str]],
    path: str | Path,
    lang: str = "zh",
) -> Path:
    """Write the original table + appended WoRMS columns to a UTF-8-BOM CSV.

    ``lang`` sets the appended-column header language ("zh"/"en"/"both").
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    cols = _resolve_match_cols(append_cols)
    out_headers = list(headers) + _append_headers(cols, lang)
    data_rows = _annotated_rows(headers, rows, results, cols)

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(out_headers)
        for row in data_rows:
            writer.writerow(row)

    return path.resolve()


def export_darwin_core(db: sqlite3.Connection, path: str | Path) -> Path:
    """Export the darwin_core VIEW to a CSV file.

    Reads directly from the SQLite darwin_core view (created by db_manager);
    column set is taken dynamically from the view, so view changes flow through
    here automatically. The view exposes the original 12 oracle terms
    (occurrenceID, scientificName, family, genus, order, decimalLongitude,
    decimalLatitude, eventDate, recordedBy, identifiedBy, locality,
    verbatimPreservation) plus standards-aligned terms joined from
    collection_records / constants: basisOfRecord, preparations, geodeticDatum,
    countryCode, occurrenceStatus, habitat, waterBody,
    minimum/maximumDepthInMeters, samplingProtocol, sampleSizeValue,
    sampleSizeUnit, samplingEffort, and dynamicProperties (JSON of 采集性质 +
    environmental measurements). See db_manager._DARWIN_CORE_SQL.

    Parameters
    ----------
    db:
        Open SQLite connection with the darwin_core view available.
    path:
        Destination .csv file path.

    Returns
    -------
    Path
        Resolved path of the written file.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    cursor = db.execute("SELECT * FROM darwin_core")
    col_names = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(col_names)
        for row in rows:
            writer.writerow(list(row))

    return path.resolve()


# ── Inventory (物种名录) export ────────────────────────────────────────────────
#
# 加法式辅助函数,服务于调查汇总视图的物种名录面板 (spec survey-summary-view T4)。
# inventory 是 taxon_inventory_service.aggregate_taxon_inventory 返回的 list[dict]
# (scientific_name / scientific_name_cn / family / genus / order_name / sites /
# count_per_site / total_count)。34 列标本导出 (export_excel / export_csv) 接收的是
# Specimen 实例,不适用于按物种聚合的汇总表,故此处的专用函数复用同一套 openpyxl
# 样式与 UTF-8-BOM CSV 约定,但完全不触碰标本导出路径。
#
# 红线:无 species / species_cn 列 → 用 scientific_name / scientific_name_cn。

#: 物种名录列定义:(表头, accessor(row_dict) -> 标量)。
INVENTORY_COLUMNS: list[tuple[str, callable]] = [
    ("学名",       lambda r: r.get("scientific_name", "")),
    ("中文名",     lambda r: r.get("scientific_name_cn", "")),
    ("目",         lambda r: r.get("order_name", "")),
    ("科",         lambda r: r.get("family", "")),
    ("属",         lambda r: r.get("genus", "")),
    ("出现断面",   lambda r: ", ".join(r.get("sites", []) or [])),
    ("各断面数量", lambda r: ", ".join(
        f"{site}={n}" for site, n in sorted((r.get("count_per_site") or {}).items())
    )),
    ("合计编号数", lambda r: r.get("total_count", 0)),
]

INVENTORY_HEADERS: list[str] = [h for h, _ in INVENTORY_COLUMNS]


def _inventory_rows(
    inventory: Sequence[dict],
    columns: list[tuple[str, callable]],
) -> list[list]:
    """inventory list[dict] → 行列表 (与 _build_data_rows 对应,但作用于 dict)。"""
    rows: list[list] = []
    for row in (inventory or []):
        out = []
        for _h, fn in columns:
            try:
                out.append(fn(row))
            except Exception:
                out.append("")
        rows.append(out)
    return rows


def export_inventory_excel(
    inventory: Sequence[dict],
    path: str | Path,
) -> Path:
    """将物种名录汇总写入 .xlsx (样式与 34 列标本导出一致)。

    *inventory* 为 taxon_inventory_service.aggregate_taxon_inventory 的返回值。
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = INVENTORY_HEADERS
    data_rows = _inventory_rows(inventory, INVENTORY_COLUMNS)

    wb = openpyxl.Workbook()
    wb.properties.creator = "拍照工作台"
    ws = wb.active
    ws.title = "物种名录"
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _apply_header_row(ws, headers)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL
    _auto_col_widths(ws, headers, data_rows)

    # 导出信息 sheet
    sites_seen: set[str] = set()
    for r in (inventory or []):
        sites_seen.update(r.get("sites", []) or [])
    ws2 = wb.create_sheet("导出信息")
    ws2["A1"] = "导出日期"
    ws2["B1"] = date.today().isoformat()
    ws2["A2"] = "物种数"
    ws2["B2"] = len(data_rows)
    ws2["A3"] = "断面数"
    ws2["B3"] = len(sites_seen)

    wb.save(str(path))
    return path.resolve()


def export_inventory_csv(
    inventory: Sequence[dict],
    path: str | Path,
) -> Path:
    """将物种名录汇总写入 UTF-8-BOM CSV (Excel 友好)。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = INVENTORY_HEADERS
    data_rows = _inventory_rows(inventory, INVENTORY_COLUMNS)

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)

    return path.resolve()


def export_inventory_darwin_core(
    workspaces: Sequence[str],
    path: str | Path,
) -> Path:
    """跨多工作区合并 darwin_core 视图行,写入一张 CSV。

    只读打开每个 ``<workspace>/_data/project.db``;db 缺失 / 被锁 / 缺少
    darwin_core 视图的工作区静默跳过 (与 taxon_inventory_service 同样的容错)。
    输出列取自首个能返回列描述的工作区 —— 所有工作区 db 共用同一 schema,
    因此列对齐。
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    col_names: list[str] = []
    combined: list[list] = []
    for ws in (workspaces or []):
        db_path = Path(ws) / "_data" / "project.db"
        if not db_path.exists():
            continue
        try:
            conn = sqlite3.connect(str(db_path))
            try:
                cur = conn.execute("SELECT * FROM darwin_core")
                if not col_names:
                    col_names = [d[0] for d in cur.description]
                combined.extend(list(r) for r in cur.fetchall())
            finally:
                conn.close()
        except sqlite3.Error:
            # 损坏 / 非 SQLite / 视图缺失 —— 跳过该断面,不阻断整体导出。
            continue

    if not col_names:
        # 没有任何工作区提供列描述时,回退到 oracle darwin_core 的 12 个核心术语,
        # 保证导出文件仍然是一份结构合法的 DwC CSV (仅有表头,无数据行)。
        col_names = [
            "occurrenceID", "scientificName", "family", "genus", "order",
            "decimalLongitude", "decimalLatitude", "eventDate", "recordedBy",
            "identifiedBy", "locality", "verbatimPreservation",
        ]

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(col_names)
        for row in combined:
            writer.writerow(row)

    return path.resolve()
