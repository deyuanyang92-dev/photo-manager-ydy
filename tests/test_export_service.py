"""Tests for export_service.py.

Covers:
  - COLUMNS list has exactly 34 entries
  - export_excel: produces a readable xlsx, correct sheet name, headers, row count
  - export_excel: column filtering (subset)
  - export_csv: produces UTF-8 BOM CSV with correct headers and row count
  - export_csv: column filtering (subset)
  - export_darwin_core: reads darwin_core view; produces correct DwC headers
  - Helpers: _pres_detail, _is_rna, _meta_score, _taxon_complete
  - Boundary: empty specimen list → file still created, only header row
  - Boundary: specimen with None fields → no crash, blank cell
"""
from __future__ import annotations

import csv
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

from app.db import db_manager
from app.models.specimen import Specimen
from app.services.export_service import (
    COLUMN_HEADERS,
    COLUMNS,
    _is_rna,
    _meta_score,
    _pres_detail,
    _taxon_complete,
    export_csv,
    export_darwin_core,
    export_excel,
)


# ── Fixtures ───────────────────────────────────────────────────────────────────

def _make_specimen(**kwargs) -> Specimen:
    """Build a Specimen with defaults for required fields."""
    defaults = dict(
        uid="FJ-XM-B2-DLC001-T95E-20260601",
        id="DLC001",
        province="FJ",
        site="XM",
        station="B2",
        storage="T95E",
        collection_date="20260601",
        photo_date="20260601",
        scientific_name="Homo sapiens",
        scientific_name_cn="人",
        taxon_group="Mammalia",
        taxon_group_cn="哺乳纲",
        order_name="Primates",
        order_cn="灵长目",
        family="Hominidae",
        family_cn="人科",
        genus="Homo",
        genus_cn="人属",
        lon=119.5,
        lat=26.1,
        geo_area="FJ·XM·B2",
        collector="张三",
        photographer="李四",
        identifier="王五",
        notes="test note",
        photo_notes="test photo note",
        angle="dorsal",
        metadata=1,
        pinned=0,
    )
    defaults.update(kwargs)
    return Specimen(**defaults)


@pytest.fixture
def specimens() -> list[Specimen]:
    return [
        _make_specimen(),
        _make_specimen(
            uid="FJ-XM-B2-DLC002-T95E-20260601",
            id="DLC002",
            scientific_name="Pan troglodytes",
            scientific_name_cn="黑猩猩",
            lon=120.0,
            lat=26.5,
        ),
    ]


@pytest.fixture
def tmp_dir(tmp_path):
    return tmp_path


@pytest.fixture(autouse=True)
def reset_db():
    db_manager.close_all()
    yield
    db_manager.close_all()


# ── Column count invariant ─────────────────────────────────────────────────────

class TestColumnCount:
    def test_exactly_34_columns(self):
        assert len(COLUMNS) == 34

    def test_headers_match_columns(self):
        assert COLUMN_HEADERS == [h for h, _ in COLUMNS]

    def test_no_duplicate_headers(self):
        assert len(COLUMN_HEADERS) == len(set(COLUMN_HEADERS))


# ── Helper functions ───────────────────────────────────────────────────────────

class TestHelpers:
    def test_pres_detail_known_code(self):
        assert "梯度酒精" in _pres_detail("T95E")

    def test_pres_detail_unknown_code(self):
        assert _pres_detail("UNKNOWN") == "UNKNOWN"

    def test_pres_detail_none(self):
        assert _pres_detail(None) == ""

    def test_is_rna_r_prefix(self):
        assert _is_rna("RD75E") is True
        assert _is_rna("RT95E") is True

    def test_is_rna_non_r(self):
        assert _is_rna("T95E") is False
        assert _is_rna("D75E") is False

    def test_is_rna_none(self):
        assert _is_rna(None) is False

    def test_meta_score_all_filled(self):
        sp = _make_specimen()
        assert _meta_score(sp) == 100

    def test_meta_score_none_fields(self):
        sp = _make_specimen(scientific_name=None, family=None, lon=None, lat=None, collector=None)
        assert _meta_score(sp) == 0

    def test_meta_score_partial(self):
        sp = _make_specimen(lon=None, lat=None)
        # 3 out of 5 filled → 60%
        assert _meta_score(sp) == 60

    def test_taxon_complete_full(self):
        sp = _make_specimen()
        assert _taxon_complete(sp) == "✓"

    def test_taxon_complete_missing_family(self):
        sp = _make_specimen(family=None)
        assert _taxon_complete(sp) == "✗"


# ── export_excel ───────────────────────────────────────────────────────────────

class TestExportExcel:
    def test_creates_file(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        assert out.exists()

    def test_sheet_name(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        assert "标本汇总" in wb.sheetnames

    def test_header_row_count(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        headers = [ws.cell(1, c).value for c in range(1, 35)]
        assert headers == COLUMN_HEADERS

    def test_data_row_count(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        # row 1 = header; rows 2..N = data
        assert ws.max_row == len(specimens) + 1

    def test_uid_in_first_column(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        uids = [ws.cell(r, 1).value for r in range(2, len(specimens) + 2)]
        assert specimens[0].uid in uids

    def test_column_filter(self, specimens, tmp_dir):
        subset = ["标本唯一编号", "物种拉丁名"]
        out = export_excel(specimens, tmp_dir / "filtered.xlsx", columns=subset)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        headers = [ws.cell(1, c).value for c in range(1, 3)]
        assert headers == subset
        # Only 2 data columns
        assert ws.max_column == 2

    def test_empty_specimens(self, tmp_dir):
        out = export_excel([], tmp_dir / "empty.xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        # Only header row
        assert ws.max_row == 1

    def test_none_fields_no_crash(self, tmp_dir):
        sp = Specimen(uid="test-uid")  # all optional fields = None
        out = export_excel([sp], tmp_dir / "none_fields.xlsx")
        assert out.exists()

    def test_metadata_sheet_exists(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        assert "导出信息" in wb.sheetnames

    def test_metadata_sheet_row_count(self, specimens, tmp_dir):
        out = export_excel(specimens, tmp_dir / "out.xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["导出信息"]
        # B2 = specimen count
        assert ws["B2"].value == len(specimens)

    def test_creates_parent_dir(self, specimens, tmp_dir):
        nested = tmp_dir / "nested" / "deep" / "out.xlsx"
        out = export_excel(specimens, nested)
        assert out.exists()


# ── export_excel: extra_leading (additive, red-line stability) ──────────────────

class TestExportExcelExtraLeading:
    def test_export_excel_unchanged_without_extra_leading(self, specimens, tmp_dir):
        """Red-line: omitting/None extra_leading must not change the 34-column output.

        We read back with openpyxl and assert the "标本汇总" header row equals the
        canonical COLUMN_HEADERS and max_column == 34. Chosen over raw byte
        comparison because the "导出信息" sheet stamps date.today() into B1, so two
        exports run on the same day are equal in content but the bytes are not a
        contract we control (openpyxl zip metadata). The header/column-count
        readback is the reliable, meaningful invariant for the red line.
        """
        p1 = tmp_dir / "default_a.xlsx"
        p2 = tmp_dir / "default_b.xlsx"
        out1 = export_excel(specimens, p1)
        out2 = export_excel(specimens, p2)

        for out in (out1, out2):
            wb = openpyxl.load_workbook(str(out))
            ws = wb["标本汇总"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            assert headers == COLUMN_HEADERS
            assert ws.max_column == 34
            assert ws.max_row == len(specimens) + 1

    def test_export_excel_extra_leading_prepends_column(self, specimens, tmp_dir):
        specs = [
            _make_specimen(owner_project_dir="/projects/断面A"),
            _make_specimen(
                uid="FJ-XM-B2-DLC002-T95E-20260601",
                id="DLC002",
                owner_project_dir="/projects/断面B",
            ),
        ]
        out = export_excel(
            specs,
            tmp_dir / "extra.xlsx",
            extra_leading=[("断面", lambda s: s.owner_project_dir or "")],
        )
        wb = openpyxl.load_workbook(str(out))
        ws = wb["标本汇总"]
        # Leading column prepended before the 34-column master list.
        assert ws.cell(1, 1).value == "断面"
        assert ws.cell(1, 2).value == COLUMN_HEADERS[0]  # "标本唯一编号"
        assert ws.max_column == 35
        # A-column data cells carry owner_project_dir values.
        a_vals = [ws.cell(r, 1).value for r in range(2, len(specs) + 2)]
        assert a_vals == ["/projects/断面A", "/projects/断面B"]


# ── export_csv ─────────────────────────────────────────────────────────────────

class TestExportCsv:
    def test_creates_file(self, specimens, tmp_dir):
        out = export_csv(specimens, tmp_dir / "out.csv")
        assert out.exists()

    def test_header_row(self, specimens, tmp_dir):
        out = export_csv(specimens, tmp_dir / "out.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            header = next(reader)
        assert header == COLUMN_HEADERS

    def test_data_row_count(self, specimens, tmp_dir):
        out = export_csv(specimens, tmp_dir / "out.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        # 1 header + N data rows
        assert len(rows) == len(specimens) + 1

    def test_column_filter(self, specimens, tmp_dir):
        subset = ["标本唯一编号", "采集人"]
        out = export_csv(specimens, tmp_dir / "filtered.csv", columns=subset)
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            header = next(csv.reader(fh))
        assert header == subset

    def test_empty_specimens(self, tmp_dir):
        out = export_csv([], tmp_dir / "empty.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        # Only header row
        assert len(rows) == 1
        assert rows[0] == COLUMN_HEADERS

    def test_none_fields_no_crash(self, tmp_dir):
        sp = Specimen(uid="test-uid")
        out = export_csv([sp], tmp_dir / "none_fields.csv")
        assert out.exists()

    def test_utf8_bom(self, specimens, tmp_dir):
        """File must start with UTF-8 BOM (0xEF BB BF) for Excel compatibility."""
        out = export_csv(specimens, tmp_dir / "bom.csv")
        raw = out.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf"

    def test_uid_value_in_first_data_column(self, specimens, tmp_dir):
        out = export_csv(specimens, tmp_dir / "out.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        data = rows[1:]  # skip header
        uids = [r[0] for r in data]
        assert specimens[0].uid in uids


# ── export_darwin_core ─────────────────────────────────────────────────────────

class TestExportDarwinCore:
    def _make_db(self, tmp_path) -> sqlite3.Connection:
        """Create a minimal project db with one specimen."""
        proj_dir = tmp_path / "proj"
        proj_dir.mkdir()
        conn = db_manager.open_project_db(str(proj_dir), create=True)
        conn.execute("""
            INSERT INTO specimens (uid, scientific_name, family, genus, order_name,
                lon, lat, collection_date, collector, identifier,
                province, site, station, storage)
            VALUES ('FJ-XM-B2-DLC001-T95E-20260601','Homo sapiens','Hominidae','Homo',
                'Primates', 119.5, 26.1, '20260601', '张三', '王五',
                'FJ', 'XM', 'B2', 'T95E')
        """)
        conn.commit()
        return conn

    def test_creates_file(self, tmp_path):
        db = self._make_db(tmp_path)
        out = export_darwin_core(db, tmp_path / "dwc.csv")
        assert out.exists()

    def test_dwc_header_fields(self, tmp_path):
        db = self._make_db(tmp_path)
        out = export_darwin_core(db, tmp_path / "dwc.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            header = next(csv.reader(fh))
        required_fields = {
            "occurrenceID", "scientificName", "family", "genus",
            "order", "decimalLongitude", "decimalLatitude",
            "eventDate", "recordedBy", "identifiedBy",
            "locality", "verbatimPreservation",
        }
        assert required_fields.issubset(set(header))

    def test_dwc_row_count(self, tmp_path):
        db = self._make_db(tmp_path)
        out = export_darwin_core(db, tmp_path / "dwc.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        # 1 header + 1 data row
        assert len(rows) == 2

    def test_dwc_occurrence_id(self, tmp_path):
        db = self._make_db(tmp_path)
        out = export_darwin_core(db, tmp_path / "dwc.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert rows[0]["occurrenceID"] == "FJ-XM-B2-DLC001-T95E-20260601"

    def test_dwc_locality_concatenation(self, tmp_path):
        db = self._make_db(tmp_path)
        out = export_darwin_core(db, tmp_path / "dwc.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert rows[0]["locality"] == "FJ·XM·B2"

    def test_dwc_empty_db(self, tmp_path):
        """Empty specimens table should yield only a header row."""
        proj_dir = tmp_path / "empty_proj"
        proj_dir.mkdir()
        db = db_manager.open_project_db(str(proj_dir), create=True)
        out = export_darwin_core(db, tmp_path / "empty_dwc.csv")
        with open(str(out), encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        assert len(rows) == 1  # header only
