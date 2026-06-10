"""test_worms_match_dialog.py — offscreen smoke for the Match-Taxa wizard.

Guards construction + the decoupled logic methods (load_file / name_list /
selected_append_cols / set_results / resolve_row / export) without GUI events.
"""
from __future__ import annotations

import os

import openpyxl
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PyQt6.QtWidgets import QApplication  # noqa: E402

from app.widgets.worms_match_dialog import WormsMatchDialog  # noqa: E402


@pytest.fixture(scope="module")
def qt_app():
    return QApplication.instance() or QApplication([])


class _FakeService:
    """Stand-in WormsService — no network, deterministic results."""

    def match_names(self, names, *, marine_only=False, auto_accept_near=False, progress_cb=None):
        out = []
        for n in names:
            nn = (n or "").strip()
            if not nn:
                out.append({"input": n, "candidates": [], "resolution": "none", "best": None})
                continue
            best = {"AphiaID": 1, "scientificname": nn, "authority": "Auth, 1900",
                    "valid_name": nn, "match_type": "exact", "rank": "Species"}
            out.append({"input": n, "candidates": [best], "resolution": "matched", "best": best})
        if progress_cb:
            progress_cb(len(names), len(names))
        return out


def _write_csv(path, text):
    path.write_text(text, encoding="utf-8-sig")
    return str(path)


def test_construct_and_load(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "names.csv", "编号,学名,备注\n1,Abra alba,x\n2,Cancer pagurus,y\n")
    dlg.load_file(csv)
    assert dlg.name_column() == "学名"          # auto-guessed
    assert dlg.name_list() == ["Abra alba", "Cancer pagurus"]
    # all WoRMS append columns checked by default
    assert "aphia_id" in dlg.selected_append_cols()


def test_results_and_export_roundtrip(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "n.csv", "学名\nAbra alba\n")
    dlg.load_file(csv)
    results = _FakeService().match_names(dlg.name_list())
    dlg.set_results(results)

    out = dlg.export(str(tmp_path / "out.xlsx"))
    ws = openpyxl.load_workbook(out).active
    hdr = [c.value for c in ws[1]]
    assert hdr[0] == "学名"                      # original column preserved first
    assert "AphiaID" in hdr                       # WoRMS column appended
    assert ws.cell(row=2, column=1).value == "Abra alba"


def test_no_header_reload(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "nh.csv", "Abra alba\nCancer pagurus\n")
    dlg.load_file(csv)
    assert len(dlg._rows) == 1                  # has-header ON → row 0 is header
    dlg._header_cb.setChecked(False)            # toggled signal reloads file
    assert len(dlg._rows) == 2                  # no-header → first row kept
    dlg._name_combo.setCurrentText(dlg._headers[0])
    assert dlg.name_list() == ["Abra alba", "Cancer pagurus"]


def test_authorities_from_author_column(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "a.csv", '学名,命名人\nAbra alba,"(W. Wood, 1802)"\n')
    dlg.load_file(csv)
    assert dlg.match_authority() is False
    assert dlg.authorities() is None            # off → no authorities
    dlg._authority_cb.setChecked(True)
    assert dlg.authorities() == ["(W. Wood, 1802)"]


def test_match_rank_mapping(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    assert dlg.match_rank() is None             # default 种 / ScientificName
    dlg._rank_combo.setCurrentIndex(1)          # 属 / Genus
    assert dlg.match_rank() == "Genus"
    assert dlg.limit_taxon() is None
    dlg._limit_edit.setText("Porifera")
    assert dlg.limit_taxon() == "Porifera"


def test_output_lang_selector_drives_export(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "n.csv", "学名\nAbra alba\n")
    dlg.load_file(csv)
    dlg.set_results(_FakeService().match_names(["Abra alba"]))
    assert dlg.output_lang() == "zh"               # default
    dlg._lang_combo.setCurrentIndex(1)             # English
    assert dlg.output_lang() == "en"

    out = dlg.export(str(tmp_path / "en.xlsx"))
    ws = openpyxl.load_workbook(out).active
    hdr = [c.value for c in ws[1]]
    assert "ScientificName" in hdr                 # English appended header
    assert "匹配名" not in hdr


def test_resolve_row_marks_none(qt_app, tmp_path):
    dlg = WormsMatchDialog(_FakeService())
    csv = _write_csv(tmp_path / "n.csv", "学名\nAbra alba\n")
    dlg.load_file(csv)
    dlg.set_results(_FakeService().match_names(["Abra alba"]))
    dlg.resolve_row(0, None, mark_none=True)
    assert dlg._results[0]["resolution"] == "none"
    assert dlg._results[0]["best"] is None
