"""test_worms_match_export.py — annotated Match-Taxa file writers.

Covers export_annotated_xlsx / export_annotated_csv:
  - all original columns preserved, in original order (zero-loss).
  - only the user-selected WoRMS columns are appended.
  - a Chinese-name (*Cn) original column passes through verbatim, exactly once,
    and no appended column is a *Cn field (red line).
  - CSV is written UTF-8 with BOM (Excel-compatible).
"""
from __future__ import annotations

import openpyxl

from app.services.export_service import (
    export_annotated_xlsx,
    export_annotated_csv,
)


def _result(name, best=None, resolution="matched", candidates=None):
    return {
        "input": name,
        "best": best,
        "resolution": resolution,
        "candidates": candidates or [],
    }


def test_xlsx_preserves_all_original_columns_and_order(tmp_path):
    headers = ["编号", "学名", "备注"]
    rows = [{"编号": "1", "学名": "Abra alba", "备注": "x"}]
    results = [_result("Abra alba", best={"AphiaID": 138474, "scientificname": "Abra alba"})]

    out = tmp_path / "o.xlsx"
    export_annotated_xlsx(headers, rows, results, None, out)

    ws = openpyxl.load_workbook(out).active
    hdr = [c.value for c in ws[1]]
    assert hdr[:3] == ["编号", "学名", "备注"]
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=2).value == "Abra alba"
    assert ws.cell(row=2, column=3).value == "x"


def test_xlsx_appends_selected_columns_only(tmp_path):
    headers = ["学名"]
    rows = [{"学名": "Abra alba"}]
    results = [_result("Abra alba", best={
        "AphiaID": 138474, "scientificname": "Abra alba",
        "authority": "(W. Wood, 1802)", "status": "accepted",
    })]

    out = tmp_path / "o.xlsx"
    export_annotated_xlsx(headers, rows, results, ["aphia_id", "authority"], out)

    ws = openpyxl.load_workbook(out).active
    hdr = [c.value for c in ws[1]]
    assert hdr == ["学名", "AphiaID", "命名人"]
    assert ws.cell(row=2, column=2).value == "138474"
    assert ws.cell(row=2, column=3).value == "(W. Wood, 1802)"


def test_writer_passes_cn_column_through_once(tmp_path):
    headers = ["学名", "物种中名"]
    rows = [{"学名": "Abra alba", "物种中名": "白阿布拉蛤"}]
    results = [_result("Abra alba", best={"AphiaID": 1, "scientificname": "Abra alba"})]

    out = tmp_path / "o.xlsx"
    export_annotated_xlsx(headers, rows, results, None, out)

    ws = openpyxl.load_workbook(out).active
    hdr = [c.value for c in ws[1]]
    assert hdr.count("物种中名") == 1
    cn_idx = hdr.index("物种中名") + 1
    assert ws.cell(row=2, column=cn_idx).value == "白阿布拉蛤"
    appended = hdr[2:]
    assert not any(str(h).endswith("Cn") for h in appended)


def test_csv_is_utf8_sig_bom(tmp_path):
    headers = ["学名"]
    rows = [{"学名": "Abra alba"}]
    results = [_result("Abra alba", best=None, resolution="none")]

    out = tmp_path / "o.csv"
    export_annotated_csv(headers, rows, results, ["aphia_id"], out)

    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf"
    text = raw.decode("utf-8-sig")
    assert "学名" in text and "AphiaID" in text
