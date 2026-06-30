"""Tests for project_summary_service.py — cross-workspace aggregation export.

These are pure-logic, READ-ONLY exports over multiple workspaces' project.db.
Each test materialises real schema-correct dbs via
``open_project_db(dir, create=True)`` then INSERTs rows directly.

Covers:
  - specimen summary merges specimens from 2 workspaces; 断面 column per row
  - collection summary merges collection_records from 2 workspaces; 断面 column
  - QC findings detect each category (missing coords, no composed grouping,
    incomplete taxonomy, cross-workspace UID conflict)
  - output files land under <root>/_data/exports/
  - a workspace whose db is missing is skipped without raising
  - _label fallback for dirs not under root
"""
from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl
import pytest

from app.db import db_manager
from app.services.activity_audit_service import record_label_print_event
from app.services import project_summary_service as pss
from app.services.photo_asset_service import assign_photo


# ── helpers ────────────────────────────────────────────────────────────────────

def _make_workspace(root: Path, name: str):
    """Create a real schema-correct workspace dir + project.db; return (dir, conn)."""
    ws_dir = root / name
    ws_dir.mkdir(parents=True, exist_ok=True)
    conn = db_manager.open_project_db(str(ws_dir), create=True)
    return str(ws_dir), conn


def _insert_specimen(conn, uid, **cols):
    keys = ["uid"] + list(cols.keys())
    vals = [uid] + list(cols.values())
    placeholders = ", ".join("?" for _ in keys)
    conn.execute(
        f"INSERT OR REPLACE INTO specimens ({', '.join(keys)}) VALUES ({placeholders})",
        vals,
    )
    conn.commit()


def _insert_record(conn, **cols):
    keys = list(cols.keys())
    vals = list(cols.values())
    placeholders = ", ".join("?" for _ in keys)
    conn.execute(
        f"INSERT INTO collection_records ({', '.join(keys)}) VALUES ({placeholders})",
        vals,
    )
    conn.commit()


def _insert_grouping(conn, uid, group_index, status):
    conn.execute(
        "INSERT INTO grouping (uid, group_index, status) VALUES (?, ?, ?)",
        (uid, group_index, status),
    )
    conn.commit()


def _insert_grouping_detail(
    conn,
    uid,
    group_index,
    *,
    status="composed",
    jpg_paths=None,
    tiff_path="",
    zip_path="",
):
    conn.execute(
        """
        INSERT INTO grouping
        (uid, group_index, status, jpg_paths, composed_tiff_path, archive_zip)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            uid,
            group_index,
            status,
            json.dumps(jpg_paths or [], ensure_ascii=False),
            tiff_path,
            zip_path,
        ),
    )
    conn.commit()


def _insert_photo(conn, photo_id, *, assigned=False, assigned_by=""):
    conn.execute(
        """
        INSERT INTO photos (
          photo_id, original_filename, first_seen_at, lifecycle_status,
          current_assignment_id
        ) VALUES (?, ?, '2026-06-01T00:00:00+00:00', ?, NULL)
        """,
        (photo_id, f"{photo_id}.jpg", "assigned" if assigned else "incoming"),
    )
    conn.commit()
    if assigned:
        assign_photo(conn, photo_id, "UID-A1", assigned_by=assigned_by)


@pytest.fixture(autouse=True)
def _clean_db_cache():
    db_manager.close_all()
    yield
    db_manager.close_all()


# ── _label ─────────────────────────────────────────────────────────────────────

def test_label_relpath_under_root(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    ws = root / "断面A"
    ws.mkdir()
    assert pss._label(str(ws), str(root)) == "断面A"


def test_label_root_itself_uses_basename(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    assert pss._label(str(root), str(root)) == "survey"


def test_label_dir_not_under_root_falls_back_to_basename(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    other = tmp_path / "elsewhere" / "projX"
    other.mkdir(parents=True)
    # relpath would be "../elsewhere/projX" → fall back to basename
    assert pss._label(str(other), str(root)) == "projX"


# ── dashboard ─────────────────────────────────────────────────────────────────

def test_project_dashboard_counts_storage_photos_people_and_prints(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    dir_b, conn_b = _make_workspace(root, "断面B")

    _insert_specimen(
        conn_a,
        "UID-A1",
        storage="D95E",
        collector="张三",
        photographer="李四",
        identifier="王五",
    )
    _insert_specimen(
        conn_a,
        "UID-A2",
        storage="RD75E",
        collector="张三",
        photographer="赵六",
        identifier="",
    )
    _insert_specimen(
        conn_b,
        "UID-B1",
        storage="T95E",
        collector="钱七",
        photographer="李四",
        identifier="王五",
    )

    _insert_photo(conn_a, "P1", assigned=True, assigned_by="李四")
    _insert_photo(conn_a, "P2", assigned=False)
    record_label_print_event(
        conn_a,
        specimen_uids=["UID-A1", "UID-A2"],
        actor="周八",
        bucket="sample",
        copies=2,
    )
    record_label_print_event(
        conn_a,
        specimen_uids=["UID-A2"],
        actor="周八",
        bucket="tissue",
        copies=1,
    )

    dashboard = pss.collect_project_dashboard([dir_a, dir_b], str(root))
    totals = dashboard["totals"]
    assert totals["workspace_count"] == 2
    assert totals["specimen_count"] == 3
    assert totals["alcohol_specimen_count"] == 2
    assert totals["rna_specimen_count"] == 1
    assert totals["photo_count"] == 2
    assert totals["unassigned_photo_count"] == 1
    assert totals["sample_label_print_count"] == 4
    assert totals["tissue_label_print_count"] == 1

    people = dashboard["by_person"]
    assert {"name": "李四", "count": 2} in people["photographer"]
    assert {"name": "张三", "count": 2} in people["collector"]
    assert {"name": "未填写", "count": 1} in people["identifier"]
    assert {"name": "李四", "count": 1} in people["photo_assigner"]
    assert {"name": "周八", "count": 5} in people["label_printer"]

    by_ws = {r["workspace_label"]: r for r in dashboard["by_workspace"]}
    assert by_ws["断面A"]["specimen_count"] == 2
    assert by_ws["断面B"]["specimen_count"] == 1


# ── specimen summary ───────────────────────────────────────────────────────────

def test_specimen_summary_merges_and_labels(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    dir_b, conn_b = _make_workspace(root, "断面B")
    _insert_specimen(conn_a, "UID-A1", scientific_name="Aaa", family="Fam")
    _insert_specimen(conn_a, "UID-A2", scientific_name="Bbb", family="Fam")
    _insert_specimen(conn_b, "UID-B1", scientific_name="Ccc", family="Fam")

    out = pss.export_specimen_summary([dir_a, dir_b], str(root))
    assert out.exists()

    wb = openpyxl.load_workbook(str(out))
    ws = wb["标本汇总"]
    assert ws.cell(1, 1).value == "断面"
    # uid column is the second column now (first is 断面)
    rows = {}
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        uid = ws.cell(r, 2).value
        if uid:
            rows[uid] = label
    assert rows["UID-A1"] == "断面A"
    assert rows["UID-A2"] == "断面A"
    assert rows["UID-B1"] == "断面B"


def test_specimen_summary_lands_under_exports(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_specimen(conn_a, "UID-A1", scientific_name="Aaa", family="Fam")
    out = pss.export_specimen_summary([dir_a], str(root))
    assert out.parent == (root / "_data" / "exports").resolve()
    assert out.name == "survey_标本汇总.xlsx"


def test_specimen_summary_skips_missing_db(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_specimen(conn_a, "UID-A1", scientific_name="Aaa", family="Fam")
    # a dir with no project.db
    missing = root / "断面X"
    missing.mkdir()

    out = pss.export_specimen_summary([dir_a, str(missing)], str(root))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["标本汇总"]
    uids = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
    assert "UID-A1" in uids


def test_collect_specimen_summary_rows_keeps_duplicate_uids_separate(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    dir_b, conn_b = _make_workspace(root, "断面B")
    _insert_specimen(conn_a, "UID-DUP", scientific_name="Aaa", family="Fam")
    _insert_specimen(conn_b, "UID-DUP", scientific_name="Bbb", family="Fam")
    _insert_grouping(conn_a, "UID-DUP", 0, "composed")
    _insert_grouping(conn_b, "UID-DUP", 0, "pending")
    conn_a.execute(
        "INSERT OR REPLACE INTO project_settings (setting_key, value_json) VALUES (?, ?)",
        ("project_meta", json.dumps({"name": "A项目", "project_code": "PA"}, ensure_ascii=False)),
    )
    conn_b.execute(
        "INSERT OR REPLACE INTO project_settings (setting_key, value_json) VALUES (?, ?)",
        ("project_meta", json.dumps({"name": "B项目", "project_code": "PB"}, ensure_ascii=False)),
    )
    conn_a.commit()
    conn_b.commit()

    payload = pss.collect_specimen_summary_rows([dir_a, dir_b], str(root), use_cache=False)

    assert len(payload["specimens"]) == 2
    key_a = pss.specimen_grouping_key(str(Path(dir_a).resolve()), "UID-DUP")
    key_b = pss.specimen_grouping_key(str(Path(dir_b).resolve()), "UID-DUP")
    assert payload["grouping"][key_a]["status"] == "已合成"
    assert payload["grouping"][key_a]["projCode"] == "PA"
    assert payload["grouping"][key_b]["status"] == "待合成"
    assert payload["grouping"][key_b]["projCode"] == "PB"


def test_collect_specimen_summary_rows_includes_result_file_counts(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_specimen(conn_a, "UID-A1", storage="RD75E", scientific_name="Aaa", family="Fam")
    _insert_grouping_detail(
        conn_a,
        "UID-A1",
        0,
        jpg_paths=["/p/a.jpg", "/p/b.jpg"],
        tiff_path="/p/results/UID-A1-1.tif",
        zip_path="/p/results/UID-A1-1.zip",
    )
    _insert_grouping_detail(
        conn_a,
        "UID-A1",
        1,
        status="pending",
        jpg_paths=["/p/c.jpg"],
    )

    payload = pss.collect_specimen_summary_rows([dir_a], str(root), use_cache=False)
    key = pss.specimen_grouping_key(str(Path(dir_a).resolve()), "UID-A1")
    facts = payload["grouping"][key]

    assert facts["count"] == 1
    assert facts["group_count"] == 2
    assert facts["status"] == "部分合成"
    assert facts["jpg_count"] == 3
    assert facts["tiff_count"] == 1
    assert facts["zip_count"] == 1
    assert facts["tiff_names"] == "UID-A1-1.tif"
    assert payload["dashboard"]["totals"]["tiff_count"] == 1


# ── collection summary ─────────────────────────────────────────────────────────

def test_collection_summary_merges_and_labels(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    dir_b, conn_b = _make_workspace(root, "断面B")
    _insert_record(conn_a, province="FJ", site="XM", station="B2",
                   collection_date="20260601", lon=119.5, lat=26.1)
    _insert_record(conn_b, province="ZJ", site="SM", station="C1",
                   collection_date="20260602", lon=121.0, lat=29.0)

    out = pss.export_collection_summary([dir_a, dir_b], str(root))
    assert out.exists()
    assert out.parent == (root / "_data" / "exports").resolve()
    assert out.name == "survey_采集站位汇总.xlsx"

    wb = openpyxl.load_workbook(str(out))
    ws = wb["采集站位汇总"]
    assert ws.cell(1, 1).value == "断面"
    labels = set()
    stations = set()
    for r in range(2, ws.max_row + 1):
        labels.add(ws.cell(r, 1).value)
        # station is one of the columns; collect station values by scanning
    # 断面 labels present
    assert "断面A" in labels
    assert "断面B" in labels


def test_collection_summary_lonlat_numeric_or_blank(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_record(conn_a, province="FJ", site="XM", station="B2",
                   collection_date="20260601", lon=119.5, lat=26.1)
    _insert_record(conn_a, province="FJ", site="XM", station="B3",
                   collection_date="20260601")  # no lon/lat

    out = pss.export_collection_summary([dir_a], str(root))
    wb = openpyxl.load_workbook(str(out))
    ws = wb["采集站位汇总"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lon_col = headers.index("lon") + 1
    lon_vals = [ws.cell(r, lon_col).value for r in range(2, ws.max_row + 1)]
    assert 119.5 in lon_vals
    # missing-coord row: written as "" → openpyxl reads an empty cell back as None,
    # never 0 (CLAUDE.md: empty lon/lat must not become 0).
    assert None in lon_vals
    assert 0 not in lon_vals


# ── QC findings ────────────────────────────────────────────────────────────────

def test_qc_findings_detect_all_categories(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    dir_b, conn_b = _make_workspace(root, "断面B")

    # missing-coord record
    _insert_record(conn_a, province="FJ", site="XM", station="NOCOORD",
                   collection_date="20260601")
    # a record with coords (should NOT be flagged)
    _insert_record(conn_a, province="FJ", site="XM", station="OK",
                   collection_date="20260601", lon=119.5, lat=26.1)

    # specimen with no composed grouping
    _insert_specimen(conn_a, "UID-NOCOMP", scientific_name="Aaa", family="Fam")
    # specimen WITH a composed grouping (should NOT be flagged as no-composed)
    _insert_specimen(conn_a, "UID-COMP", scientific_name="Bbb", family="Fam")
    _insert_grouping(conn_a, "UID-COMP", 0, "composed")

    # taxonomy-incomplete specimen (missing scientific_name)
    _insert_specimen(conn_a, "UID-BADTAX", scientific_name="", family="Fam")

    # SAME uid in two workspaces → cross-workspace conflict
    _insert_specimen(conn_a, "UID-DUP", scientific_name="Ddd", family="Fam")
    _insert_grouping(conn_a, "UID-DUP", 0, "composed")
    _insert_specimen(conn_b, "UID-DUP", scientific_name="Ddd", family="Fam")
    _insert_grouping(conn_b, "UID-DUP", 0, "composed")

    findings = pss.collect_qc_findings([dir_a, dir_b], str(root))
    cats = {f["category"] for f in findings}
    assert "缺经纬度站位" in cats
    assert "有标本无成片" in cats
    assert "分类不完整" in cats
    assert "跨断面UID冲突" in cats

    # missing-coord finding entity contains the station
    miss = [f for f in findings if f["category"] == "缺经纬度站位"]
    assert any("NOCOORD" in f["entity"] for f in miss)
    assert all("OK" not in f["entity"] for f in miss)

    # no-composed finding lists UID-NOCOMP but not UID-COMP
    nocomp = [f for f in findings if f["category"] == "有标本无成片"]
    nocomp_uids = {f["entity"] for f in nocomp}
    assert "UID-NOCOMP" in nocomp_uids
    assert "UID-COMP" not in nocomp_uids

    # taxonomy finding lists UID-BADTAX
    badtax = [f for f in findings if f["category"] == "分类不完整"]
    assert any(f["entity"] == "UID-BADTAX" for f in badtax)

    # conflict finding lists UID-DUP with both labels
    conflict = [f for f in findings if f["category"] == "跨断面UID冲突"]
    assert any(f["entity"] == "UID-DUP" for f in conflict)
    detail = next(f["detail"] for f in conflict if f["entity"] == "UID-DUP")
    assert "断面A" in detail and "断面B" in detail


def test_qc_report_writes_both_files_under_exports(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_record(conn_a, province="FJ", site="XM", station="NOCOORD",
                   collection_date="20260601")
    _insert_specimen(conn_a, "UID-A1", scientific_name="", family="")

    html_path, xlsx_path = pss.export_qc_report([dir_a], str(root))
    assert html_path.exists()
    assert xlsx_path.exists()
    exports = (root / "_data" / "exports").resolve()
    assert html_path.parent == exports
    assert xlsx_path.parent == exports
    assert html_path.name == "survey_质控报告.html"
    assert xlsx_path.name == "survey_质控报告.xlsx"
    # HTML mentions a category heading
    text = html_path.read_text(encoding="utf-8")
    assert "缺经纬度站位" in text or "分类不完整" in text


def test_qc_report_skips_missing_db(tmp_path):
    root = tmp_path / "survey"
    root.mkdir()
    dir_a, conn_a = _make_workspace(root, "断面A")
    _insert_specimen(conn_a, "UID-A1", scientific_name="", family="")
    missing = root / "断面X"
    missing.mkdir()
    # should not raise
    findings = pss.collect_qc_findings([dir_a, str(missing)], str(root))
    assert any(f["category"] == "分类不完整" for f in findings)
