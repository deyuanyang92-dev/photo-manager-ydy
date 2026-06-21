"""tests/test_collection_record_io.py — 采集记录 Excel/CSV 导出导入（zone 模型）.

两套国标表样：潮间带 H.39 / 潮下带 H.30。列序来自 crs.ZONE_COLUMNS。
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest

from app.db.db_manager import ensure_schema
from app.services import collection_record_io as io
from app.services import collection_record_service as crs


@pytest.fixture()
def db():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    ensure_schema(conn)
    yield conn
    conn.close()


def _xlsx_header(path: str) -> list[str]:
    """读首个 sheet 表头。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [("" if c is None else str(c)).strip() for c in next(ws.iter_rows(values_only=True))]


def _xlsx_sheet_names(path: str) -> list[str]:
    import openpyxl
    return openpyxl.load_workbook(path, read_only=True).sheetnames


# ── 模板列序 parity（国标表样）──────────────────────────────────────────────

class TestTemplateParity:
    def test_intertidal_template_has_H39_columns(self, db, tmp_path):
        out = tmp_path / "it.xlsx"
        io.export_template(db, str(out), zone="intertidal", blank_rows=0)
        hdr = _xlsx_header(str(out))
        for must in ("样方号", "气温(℃)", "定量瓶数", "定性瓶数", "样品厚度(cm)",
                     "潮区", "底质", "气象"):
            assert must in hdr, f"潮间带模板缺列 {must}"
        # 不应含潮下带专属
        for nope in ("航次", "放绳长度(m)", "网型", "采泥器面积(m²)"):
            assert nope not in hdr, f"潮间带模板不应含 {nope}"

    def test_subtidal_template_has_H30_columns(self, db, tmp_path):
        out = tmp_path / "st.xlsx"
        io.export_template(db, str(out), zone="subtidal", blank_rows=0)
        hdr = _xlsx_header(str(out))
        for must in ("航次", "船号", "放绳长度(m)", "采泥器面积(m²)", "网型",
                     "网宽(m)", "拖网距离(m)", "拖网起始", "拖网结束",
                     "采泥样品总数", "拖网样品总数", "样品厚度(cm)", "底质", "气象"):
            assert must in hdr, f"潮下带模板缺列 {must}"
        for nope in ("样方号", "气温(℃)", "定量瓶数", "潮区"):
            assert nope not in hdr, f"潮下带模板不应含 {nope}"

    def test_all_zone_makes_two_sheets(self, db, tmp_path):
        out = tmp_path / "all.xlsx"
        io.export_template(db, str(out), zone="all", blank_rows=0)
        names = _xlsx_sheet_names(str(out))
        assert names == ["潮间带", "潮下带"]


# ── export 按区筛选 ──────────────────────────────────────────────────────────

class TestExportFiltering:
    def test_single_zone_exports_only_that_zone(self, db, tmp_path):
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "B2",
                               "collection_date": "20260518", "zone": "intertidal",
                               "quadrate_no": "B2-Q1"})
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "H1",
                               "collection_date": "20260518", "zone": "subtidal",
                               "wire_out": "30"})
        out = tmp_path / "st.xlsx"
        n = io.export_template(db, str(out), zone="subtidal", blank_rows=0)
        assert n == 1
        header, rows = io._read_xlsx(str(out))
        assert rows[0][header.index("站位")] == "H1"
        assert rows[0][header.index("放绳长度(m)")] == "30"

    def test_legacy_null_records_route_to_intertidal_sheet(self, db, tmp_path):
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "B2",
                               "collection_date": "20260518"})  # 无 zone
        out = tmp_path / "all.xlsx"
        n = io.export_template(db, str(out), zone="all", blank_rows=0)
        assert n == 1   # 仅在潮间带 sheet 出现一次，不重复
        names = _xlsx_sheet_names(str(out))
        assert names == ["潮间带", "潮下带"]
        # 潮间带 sheet 含该行，潮下带不含
        import openpyxl
        wb = openpyxl.load_workbook(str(out), read_only=True, data_only=True)
        it_rows = list(wb["潮间带"].iter_rows(values_only=True))
        st_rows = list(wb["潮下带"].iter_rows(values_only=True))
        assert any(r[2] == "B2" for r in it_rows[1:])   # 站位列
        assert not any(r[2] == "B2" for r in st_rows[1:])

    def test_blank_rows_prefill_province_site(self, db, tmp_path):
        out = tmp_path / "it.xlsx"
        io.export_template(db, str(out), zone="intertidal", province="ZJ",
                           site="SMW", blank_rows=3)
        header, rows = io._read_xlsx(str(out))
        # 第一个空行（无记录时即首数据行）地区/样地预填
        assert rows[0][header.index("地区")] == "ZJ"
        assert rows[0][header.index("样地")] == "SMW"


# ── roundtrip（zone + 专属字段不丢）──────────────────────────────────────────

class TestRoundtrip:
    def test_intertidal_roundtrip(self, db, tmp_path):
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "B2",
                               "collection_date": "20260518", "zone": "intertidal",
                               "quadrate_no": "B2-Q3", "air_temp": "26",
                               "quant_bottles": "2", "tidal_zone": "中潮区",
                               "collector": "杨德援"})
        out = tmp_path / "it.xlsx"
        io.export_template(db, str(out), zone="intertidal", blank_rows=0)
        db.execute("DELETE FROM collection_records"); db.commit()
        rep = io.import_file(db, str(out))
        assert rep.ok and rep.imported == 1
        rec = crs.lookup_record(db, "ZJ", "SMW", "B2", "20260518")
        assert rec["zone"] == "intertidal"
        for k, v in {"quadrate_no": "B2-Q3", "air_temp": "26",
                     "quant_bottles": "2", "tidal_zone": "中潮区",
                     "collector": "杨德援"}.items():
            assert rec[k] == v, f"{k} 丢失: {rec.get(k)!r}"

    def test_subtidal_roundtrip(self, db, tmp_path):
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "H1",
                               "collection_date": "20260601", "zone": "subtidal",
                               "cruise": "2026春季航次", "vessel": "科学三号",
                               "wire_out": "32", "sampler_area": "0.1",
                               "net_type": "阿氏网", "trawl_distance": "500"})
        out = tmp_path / "st.xlsx"
        io.export_template(db, str(out), zone="subtidal", blank_rows=0)
        db.execute("DELETE FROM collection_records"); db.commit()
        rep = io.import_file(db, str(out))
        assert rep.ok and rep.imported == 1
        rec = crs.lookup_record(db, "ZJ", "SMW", "H1", "20260601")
        assert rec["zone"] == "subtidal"
        for k, v in {"cruise": "2026春季航次", "vessel": "科学三号",
                     "wire_out": "32", "sampler_area": "0.1",
                     "net_type": "阿氏网", "trawl_distance": "500"}.items():
            assert rec[k] == v, f"{k} 丢失: {rec.get(k)!r}"

    def test_all_zone_two_sheet_roundtrip(self, db, tmp_path):
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "B2",
                               "collection_date": "20260518", "zone": "intertidal",
                               "air_temp": "26"})
        crs.upsert_record(db, {"province": "ZJ", "site": "SMW", "station": "H1",
                               "collection_date": "20260601", "zone": "subtidal",
                               "wire_out": "30"})
        out = tmp_path / "all.xlsx"
        io.export_template(db, str(out), zone="all", blank_rows=0)
        db.execute("DELETE FROM collection_records"); db.commit()
        rep = io.import_file(db, str(out))
        assert rep.imported == 2
        it = crs.lookup_record(db, "ZJ", "SMW", "B2", "20260518")
        st = crs.lookup_record(db, "ZJ", "SMW", "H1", "20260601")
        assert it["zone"] == "intertidal" and it["air_temp"] == "26"
        assert st["zone"] == "subtidal" and st["wire_out"] == "30"


# ── CSV / 异常 / 英文 key / 旧表头兼容 ────────────────────────────────────────

class TestImportMisc:
    def test_import_csv_intertidal_headers(self, db, tmp_path):
        csv_path = tmp_path / "in.csv"
        csv_path.write_text(
            "地区,样地,站位,采集日期,样方号,底质,采集人\n"
            "FJ,XM,B2,20260601,B2-Q1,岩相,李四\n",
            encoding="utf-8-sig",
        )
        rep = io.import_file(db, str(csv_path))
        assert rep.imported == 1
        rec = crs.lookup_record(db, "FJ", "XM", "B2", "20260601")
        assert rec["habitat"] == "岩相" and rec["collector"] == "李四"
        assert rec["quadrate_no"] == "B2-Q1"
        assert rec["zone"] == "intertidal"

    def test_import_csv_subtidal_headers(self, db, tmp_path):
        csv_path = tmp_path / "st.csv"
        csv_path.write_text(
            "地区,样地,站位,采集日期,航次,放绳长度(m),网型\n"
            "FJ,XM,H1,20260601,春季航次,30,阿氏网\n",
            encoding="utf-8-sig",
        )
        rep = io.import_file(db, str(csv_path))
        assert rep.imported == 1
        rec = crs.lookup_record(db, "FJ", "XM", "H1", "20260601")
        assert rec["zone"] == "subtidal"
        assert rec["cruise"] == "春季航次" and rec["wire_out"] == "30"

    def test_import_skips_rows_missing_key_fields(self, db, tmp_path):
        csv_path = tmp_path / "in.csv"
        csv_path.write_text(
            "地区,样地,站位,采集日期,底质\n"
            "FJ,XM,,20260601,泥滩\n"      # missing 站位 → skip
            "FJ,XM,B2,20260601,岩相\n",   # ok
            encoding="utf-8-sig",
        )
        rep = io.import_file(db, str(csv_path))
        assert rep.imported == 1
        assert rep.skipped == 1

    def test_import_unrecognized_header_fails_gracefully(self, db, tmp_path):
        csv_path = tmp_path / "bad.csv"
        csv_path.write_text("foo,bar,baz\n1,2,3\n", encoding="utf-8-sig")
        rep = io.import_file(db, str(csv_path))
        assert rep.ok is False
        assert rep.errors

    def test_import_english_keys(self, db, tmp_path):
        csv_path = tmp_path / "en.csv"
        csv_path.write_text(
            "province,site,station,collection_date,habitat\n"
            "GD,雷州,S09,20260519,沙滩\n",
            encoding="utf-8-sig",
        )
        rep = io.import_file(db, str(csv_path))
        assert rep.imported == 1
        rec = crs.lookup_record(db, "GD", "雷州", "S09", "20260519")
        assert rec["habitat"] == "沙滩"
        # 无 zone 判别表头 → fallback intertidal
        assert rec["zone"] == "intertidal"

    def test_import_legacy_headers_still_work(self, db, tmp_path):
        """旧表头「生境/天气/水温/采集方法」仍能读（UI 改名前的模板兼容）。"""
        csv_path = tmp_path / "legacy.csv"
        csv_path.write_text(
            "地区,样地,站位,采集日期,生境,天气,水温\n"
            "GD,雷州,S09,20260519,沙滩,晴,25\n",
            encoding="utf-8-sig",
        )
        rep = io.import_file(db, str(csv_path))
        assert rep.imported == 1
        rec = crs.lookup_record(db, "GD", "雷州", "S09", "20260519")
        assert rec["habitat"] == "沙滩"
        assert rec["weather"] == "晴"
        assert rec["water_temp"] == "25"
